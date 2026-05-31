"""
V5+V6 Combined: Use CDCB 121k to learn trait behavior distributions,
then enrich V5 model with that knowledge.

Key insight from user: we don't need the dam from CDCB.
We need to learn HOW each PTA behaves: distributions, bounds, medians,
correlations, and feed that as domain knowledge into V5.
"""
import warnings
warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd
import openpyxl, csv
from pathlib import Path
from collections import defaultdict
from sklearn.model_selection import KFold
from sklearn.metrics import mean_absolute_error, r2_score
from lightgbm import LGBMRegressor

DOWNLOADS = Path("C:/Users/DiegoGuerra/Downloads")
BULLS_CSV = Path("C:/Users/DiegoGuerra/gen-genie-advisor/sms-engine/data/bulls.csv")

def sf(v):
    if v is None: return None
    v = str(v).strip().strip('"')
    if not v: return None
    try:
        x = float(v)
        return x if not np.isnan(x) else None
    except: return None

def read_csv_safe(path):
    for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
        try:
            with open(path, 'r', encoding=enc) as f:
                return list(csv.DictReader(f))
        except: continue

# ============================================================
# PHASE 1: Learn PTA behavior from CDCB 130k cows
# ============================================================
print("=" * 90)
print("  PHASE 1: Learning PTA distributions from CDCB 130k cows")
print("=" * 90)

cow_rows = read_csv_safe(DOWNLOADS / "Cow_Report (4).csv")
bull_rows = read_csv_safe(DOWNLOADS / "Bull_Report (1).csv")
print(f"  Cows: {len(cow_rows)}, Bulls: {len(bull_rows)}")

CDCB_TRAITS = {
    'NM$': 'NM$_PTA', 'MILK': 'MLK_PTA', 'FAT': 'FAT_PTA', 'PRO': 'PRO_PTA',
    'PL': 'PL_PTA', 'SCS': 'SCS_PTA', 'DPR': 'DPR_PTA', 'HCR': 'HCR_PTA',
    'CCR': 'CCR_PTA', 'LIV': 'LIV_PTA', 'MAST': 'MAS_PTA', 'FSAV': 'FS_PTA',
}
TRAITS = list(CDCB_TRAITS.keys())

# Learn distributions from cows (daughters)
cow_distributions = {}
for trait, col in CDCB_TRAITS.items():
    vals = [sf(r.get(col)) for r in cow_rows]
    vals = [v for v in vals if v is not None]
    arr = np.array(vals)
    cow_distributions[trait] = {
        'mean': np.mean(arr), 'std': np.std(arr), 'median': np.median(arr),
        'p5': np.percentile(arr, 5), 'p10': np.percentile(arr, 10),
        'p25': np.percentile(arr, 25), 'p75': np.percentile(arr, 75),
        'p90': np.percentile(arr, 90), 'p95': np.percentile(arr, 95),
        'min': np.min(arr), 'max': np.max(arr), 'iqr': np.percentile(arr, 75) - np.percentile(arr, 25),
        'skew': float(pd.Series(arr).skew()), 'kurtosis': float(pd.Series(arr).kurtosis()),
    }
    print(f"  {trait:>6}: mean={cow_distributions[trait]['mean']:>8.1f}  "
          f"std={cow_distributions[trait]['std']:>7.1f}  "
          f"P5={cow_distributions[trait]['p5']:>8.1f}  P95={cow_distributions[trait]['p95']:>8.1f}")

# Learn distributions from bulls (sires)
bull_distributions = {}
for trait, col in CDCB_TRAITS.items():
    vals = [sf(r.get(col)) for r in bull_rows]
    vals = [v for v in vals if v is not None]
    arr = np.array(vals)
    bull_distributions[trait] = {
        'mean': np.mean(arr), 'std': np.std(arr), 'median': np.median(arr),
        'p5': np.percentile(arr, 5), 'p25': np.percentile(arr, 25),
        'p75': np.percentile(arr, 75), 'p95': np.percentile(arr, 95),
        'iqr': np.percentile(arr, 75) - np.percentile(arr, 25),
    }

# Learn sire->daughter transfer function from CDCB
# For each sire with daughters, compute: mean(daughters) vs sire PTA
print("\nLearning sire->daughter transfer...")
bull_idx = {r['ANIMAL'].strip('"').upper(): r for r in bull_rows}
sire_daughter_vals = defaultdict(lambda: defaultdict(list))

for r in cow_rows:
    sire_id = r['SIRE'].strip('"').upper()
    for trait, col in CDCB_TRAITS.items():
        dv = sf(r.get(col))
        if dv is not None:
            sire_daughter_vals[sire_id][trait].append(dv)

# Transfer coefficients: daughter_mean = alpha * sire_PTA + beta
transfer_coefs = {}
for trait in TRAITS:
    sire_ptas = []
    daughter_means = []
    daughter_stds = []
    ns = []
    for sire_id, trait_vals in sire_daughter_vals.items():
        if trait not in trait_vals or len(trait_vals[trait]) < 5:
            continue
        sire = bull_idx.get(sire_id)
        if not sire:
            continue
        sv = sf(sire.get(CDCB_TRAITS[trait]))
        if sv is None:
            continue
        sire_ptas.append(sv)
        daughter_means.append(np.mean(trait_vals[trait]))
        daughter_stds.append(np.std(trait_vals[trait]))
        ns.append(len(trait_vals[trait]))

    if len(sire_ptas) > 10:
        sp = np.array(sire_ptas)
        dm = np.array(daughter_means)
        ds = np.array(daughter_stds)
        # Linear fit: daughter = alpha * sire + beta
        A = np.vstack([sp, np.ones(len(sp))]).T
        alpha, beta = np.linalg.lstsq(A, dm, rcond=None)[0]
        corr = np.corrcoef(sp, dm)[0, 1]
        transfer_coefs[trait] = {
            'alpha': alpha, 'beta': beta, 'corr': corr,
            'residual_std': np.std(dm - (alpha * sp + beta)),
            'mean_daughter_std': np.mean(ds),
            'n_sires': len(sire_ptas),
        }
        print(f"  {trait:>6}: alpha={alpha:.3f} beta={beta:>7.1f} corr={corr:.3f} "
              f"res_std={transfer_coefs[trait]['residual_std']:.1f} n_sires={len(sire_ptas)}")

# ============================================================
# PHASE 2: Enrich V5 features with CDCB-learned knowledge
# ============================================================
print(f"\n{'='*90}")
print("  PHASE 2: V5 + CDCB Knowledge")
print("=" * 90)

# Load May2026 trios
wb = openpyxl.load_workbook(DOWNLOADS / "May 2026.xlsx", read_only=True, data_only=True)
ws = wb.active
may_rows_xl = list(ws.iter_rows(values_only=True))
wb.close()
hdr = [str(c).strip() if c else f'col_{i}' for i, c in enumerate(may_rows_xl[0])]

dams = {}
for i in range(1, 21):
    f = DOWNLOADS / f"DAM{i}.xlsx"
    if not f.exists(): continue
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb.active
    dr = list(ws.iter_rows(values_only=True))
    wb.close()
    dh = [str(c).strip() if c else '' for j, c in enumerate(dr[0])]
    ni = next(j for j, h in enumerate(dh) if 'naab' in h.lower() and 'code' in h.lower())
    for r in dr[1:]:
        if r[ni]: dams[str(r[ni]).strip()] = {dh[j]: r[j] for j in range(min(len(dh), len(r)))}

bulls_by_regname, bulls_by_name = {}, {}
with open(BULLS_CSV, 'r', encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        rn = row.get('Registration Name', '').strip().upper()
        nm = row.get('Name', '').strip().upper()
        if rn: bulls_by_regname[rn] = row
        if nm: bulls_by_name[nm] = row

BULL_COL = {'NM$': 'NM$', 'MILK': 'PTAM', 'FAT': 'PTAF', 'PRO': 'PTAP',
            'PL': 'PL', 'SCS': 'SCS', 'DPR': 'DPR', 'HCR': 'HCR',
            'CCR': 'CCR', 'LIV': 'LIV', 'MAST': 'MAST', 'FSAV': 'F SAV'}
DAM_COL = {'NM$': 'Net Merit', 'MILK': 'PTA Milk', 'FAT': 'PTA Fat', 'PRO': 'PTA Pro',
           'PL': 'PL', 'SCS': 'SCS', 'DPR': 'PTA DPR', 'HCR': 'HCR',
           'CCR': 'CCR', 'LIV': 'PTA LIV', 'MAST': 'Mastitis', 'FSAV': 'Feed Saved'}
DAUGH_COL = {'NM$': 'NM$', 'MILK': 'MILK', 'FAT': 'FAT', 'PRO': 'PRO',
             'PL': 'PL', 'SCS': 'SCS', 'DPR': 'DPR', 'HCR': 'HCR',
             'CCR': 'CCR', 'LIV': 'LIV', 'MAST': 'MAST', 'FSAV': 'FSAV'}

records = []
for r in may_rows_xl[1:]:
    rd = {hdr[i]: r[i] for i in range(min(len(hdr), len(r)))}
    sn = str(rd.get('SIRENAME', '')).strip().upper()
    sire = bulls_by_regname.get(sn) or bulls_by_name.get(sn)
    if not sire: continue
    damreg = str(rd.get('DAMREGNUM2', '')).strip()
    dam = dams.get(damreg)
    if not dam: continue

    rec = {}
    for t in TRAITS:
        rec[f'd_{t}'] = sf(rd.get(DAUGH_COL[t]))
        rec[f's_{t}'] = sf(sire.get(BULL_COL[t]))
        rec[f'm_{t}'] = sf(dam.get(DAM_COL[t]))
    records.append(rec)

test_df = pd.DataFrame(records)
print(f"May2026 trios: {len(test_df)}")

# Compare V5 vs V5+CDCB knowledge
lgbm_kw = dict(n_estimators=500, max_depth=6, learning_rate=0.03,
               subsample=0.8, colsample_bytree=0.7, min_child_samples=3,
               reg_alpha=0.1, reg_lambda=1.0, random_state=42, verbose=-1, n_jobs=-1)

kf = KFold(n_splits=5, shuffle=True, random_state=42)

print(f"\n{'Trait':>6} | {'N':>5} | {'PA R2':>7} | {'V5 R2':>7} | {'V5+CDCB R2':>10} | "
      f"{'V5 MAE':>8} {'V5+CDCB MAE':>11} | {'Gain':>6}")
print("-" * 95)

all_results = []

for trait in TRAITS:
    target = f'd_{trait}'
    s_col = f's_{trait}'
    m_col = f'm_{trait}'

    needed = [target, s_col, m_col]
    mask = test_df[needed].notna().all(axis=1)
    sub = test_df[mask].reset_index(drop=True)
    if len(sub) < 50: continue

    y = sub[target].values
    sv = sub[s_col].values.astype(float)
    mv = sub[m_col].values.astype(float)
    pa = (sv + mv) / 2
    pa_r2 = r2_score(y, pa)

    # V5 base features
    base_feats = [sv, mv, pa, sv - mv, sv * mv, sv**2, mv**2]
    for ot in TRAITS:
        if ot == trait: continue
        osv = sub[f's_{ot}'].fillna(0).values.astype(float)
        omv = sub[f'm_{ot}'].fillna(0).values.astype(float)
        base_feats.extend([osv, omv])
    X_v5 = np.column_stack(base_feats)

    # CDCB-enriched features
    cdcb_feats = []
    cd = cow_distributions[trait]
    bd = bull_distributions[trait]
    tc = transfer_coefs.get(trait, {})

    # Percentile position of PA in CDCB cow distribution
    pa_percentile = np.array([(v - cd['mean']) / cd['std'] for v in pa])
    # Percentile position of sire in CDCB bull distribution
    sire_percentile = np.array([(v - bd['mean']) / bd['std'] for v in sv])
    # Percentile position of dam in CDCB cow distribution
    dam_percentile = np.array([(v - cd['mean']) / cd['std'] for v in mv])

    # Transfer function prediction
    if tc:
        transfer_pred = tc['alpha'] * sv + tc['beta']
        transfer_residual_expected = np.full(len(sv), tc['residual_std'])
    else:
        transfer_pred = sv / 2
        transfer_residual_expected = np.full(len(sv), cd['std'])

    # How extreme is this PA relative to population?
    pa_extreme = np.abs(pa - cd['median']) / cd['iqr']
    sire_extreme = np.abs(sv - bd['median']) / bd['iqr'] if bd['iqr'] > 0 else np.zeros(len(sv))

    # Is PA in tails? (potential regression to mean)
    pa_in_upper_tail = (pa > cd['p90']).astype(float)
    pa_in_lower_tail = (pa < cd['p10']).astype(float)

    # Skewness-adjusted prediction
    skew_adj = np.where(pa > cd['median'], -cd['skew'] * 0.1, cd['skew'] * 0.1)

    # Cross-trait CDCB features: percentile position in related traits
    cross_cdcb = []
    for ot in TRAITS:
        if ot == trait: continue
        ocd = cow_distributions[ot]
        osv = sub[f's_{ot}'].fillna(0).values.astype(float)
        omv = sub[f'm_{ot}'].fillna(0).values.astype(float)
        opa = (osv + omv) / 2
        cross_cdcb.append((opa - ocd['mean']) / ocd['std'] if ocd['std'] > 0 else np.zeros(len(opa)))

    X_enriched = np.column_stack([
        X_v5,
        pa_percentile, sire_percentile, dam_percentile,
        transfer_pred, transfer_residual_expected,
        pa_extreme, sire_extreme,
        pa_in_upper_tail, pa_in_lower_tail,
        skew_adj,
        *cross_cdcb,
    ])

    # CV both
    v5_r2s, v5_maes = [], []
    vc_r2s, vc_maes = [], []
    for tr, te in kf.split(X_v5):
        m1 = LGBMRegressor(**lgbm_kw); m1.fit(X_v5[tr], y[tr])
        p1 = m1.predict(X_v5[te])
        v5_r2s.append(r2_score(y[te], p1)); v5_maes.append(mean_absolute_error(y[te], p1))

        m2 = LGBMRegressor(**lgbm_kw); m2.fit(X_enriched[tr], y[tr])
        p2 = m2.predict(X_enriched[te])
        vc_r2s.append(r2_score(y[te], p2)); vc_maes.append(mean_absolute_error(y[te], p2))

    v5_r2 = np.mean(v5_r2s); v5_mae = np.mean(v5_maes)
    vc_r2 = np.mean(vc_r2s); vc_mae = np.mean(vc_maes)
    gain = (vc_r2 - v5_r2) / max(abs(v5_r2), 0.001) * 100

    print(f"  {trait:>5} | {len(y):>5} | {pa_r2:>7.4f} | {v5_r2:>7.4f} | {vc_r2:>10.4f} | "
          f"{v5_mae:>8.3f} {vc_mae:>11.3f} | {gain:>+5.1f}%")

    all_results.append({
        'Trait': trait, 'N': len(y),
        'PA_R2': round(pa_r2, 4), 'V5_R2': round(v5_r2, 4),
        'V5_CDCB_R2': round(vc_r2, 4), 'V5_MAE': round(v5_mae, 4),
        'V5_CDCB_MAE': round(vc_mae, 4), 'Gain_%': round(gain, 2),
    })

print(f"\n{'='*90}")
avg_v5 = np.mean([r['V5_R2'] for r in all_results])
avg_vc = np.mean([r['V5_CDCB_R2'] for r in all_results])
avg_gain = np.mean([r['Gain_%'] for r in all_results])
v_wins = sum(1 for r in all_results if r['V5_CDCB_R2'] > r['V5_R2'])

print(f"  V5 R2 medio:      {avg_v5:.4f}")
print(f"  V5+CDCB R2 medio: {avg_vc:.4f}")
print(f"  Gain medio:       {avg_gain:+.1f}%")
print(f"  V5+CDCB vence V5: {v_wins}/{len(all_results)} traits")

pd.DataFrame(all_results).to_csv(
    Path("C:/Users/DiegoGuerra/Documents/Projetos/genetic_prediction/dsii_v6_results/v5_cdcb_comparison.csv"),
    index=False)
print("  Salvo em dsii_v6_results/v5_cdcb_comparison.csv")
