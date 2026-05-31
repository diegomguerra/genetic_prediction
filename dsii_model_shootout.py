"""
DSII Model Shootout — Test ALL promising models to reach R2 ~ 0.9
Uses V5 feature engineering (proven best) + tests 10 model architectures.
"""
import warnings
warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd
import openpyxl, csv
from pathlib import Path
from collections import defaultdict

from sklearn.model_selection import KFold
from sklearn.metrics import mean_absolute_error, r2_score
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline

# Models
from sklearn.linear_model import Ridge, BayesianRidge, ElasticNet
from sklearn.svm import SVR
from sklearn.neural_network import MLPRegressor
from sklearn.ensemble import (RandomForestRegressor, GradientBoostingRegressor,
                               StackingRegressor, VotingRegressor)
from sklearn.gaussian_process import GaussianProcessRegressor
from sklearn.gaussian_process.kernels import RBF, ConstantKernel, Matern
from lightgbm import LGBMRegressor
from xgboost import XGBRegressor
from catboost import CatBoostRegressor

DOWNLOADS = Path("C:/Users/DiegoGuerra/Downloads")
BULLS_CSV = Path("C:/Users/DiegoGuerra/gen-genie-advisor/sms-engine/data/bulls.csv")
OUTPUT_DIR = Path("C:/Users/DiegoGuerra/Documents/Projetos/genetic_prediction/dsii_shootout")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

def sf(v):
    if v is None: return None
    try:
        x = float(v)
        return x if not np.isnan(x) else None
    except: return None

# ============================================================
# DOMAIN KNOWLEDGE (same as V5)
# ============================================================
GENETIC_SD = {
    'TPI': 250, 'NM$': 275, 'CM$': 280,
    'MILK': 675, 'FAT': 29, 'FAT%': 0.05, 'PRO': 19, 'PRO%': 0.02, 'CFP': 25,
    'PL': 1.85, 'SCS': 0.14, 'DPR': 1.3, 'HCR': 1.4, 'CCR': 1.65,
    'LIV': 1.2, 'FI': 1.0, 'MAST': 1.0, 'FSAV': 50,
    'PTAT': 0.70, 'UDC': 0.75, 'FLC': 0.65,
    'SCE': 1.5, 'DCE': 1.2, 'SSB': 1.0, 'DSB': 1.0,
}
HERITABILITY = {
    'TPI': 0.30, 'NM$': 0.30, 'CM$': 0.30,
    'MILK': 0.25, 'FAT': 0.25, 'FAT%': 0.50, 'PRO': 0.25, 'PRO%': 0.50,
    'CFP': 0.30, 'PL': 0.08, 'SCS': 0.12, 'DPR': 0.04, 'LIV': 0.05,
    'FI': 0.06, 'HCR': 0.04, 'CCR': 0.04, 'MAST': 0.04, 'FSAV': 0.15,
    'PTAT': 0.30, 'UDC': 0.25, 'FLC': 0.15,
    'SCE': 0.08, 'DCE': 0.06, 'SSB': 0.06, 'DSB': 0.04,
}
BREED_IDEAL = {
    'MILK': 1800, 'FAT': 90, 'PRO': 60, 'CFP': 70,
    'DPR': 1.5, 'CCR': 2.0, 'HCR': 2.0,
    'PL': 5.0, 'LIV': 2.0, 'UDC': 1.5, 'FLC': 1.0,
    'MAST': -1.0, 'SCS': 2.60, 'SCE': 2.0,
}
LOWER_IS_BETTER = {'SCS', 'SCE', 'SSB', 'DSB', 'DCE', 'MAST'}
INBREEDING_DEPRESSION = {
    'NM$': -25.0, 'TPI': -20.0, 'CM$': -25.0,
    'MILK': -28.5, 'FAT': -1.0, 'PRO': -0.9,
    'PL': -0.35, 'DPR': -0.03, 'SCS': 0.007, 'LIV': -0.15,
}
GENETIC_CORRELATIONS = {
    ('MILK', 'DPR'): -0.35, ('MILK', 'CCR'): -0.30, ('MILK', 'SCS'): 0.10,
    ('MILK', 'PL'): -0.15, ('FAT', 'PRO'): 0.65, ('FAT', 'DPR'): -0.25,
    ('PRO', 'DPR'): -0.30, ('DPR', 'CCR'): 0.55, ('DPR', 'PL'): 0.45,
    ('SCS', 'PL'): -0.30, ('SCS', 'MAST'): 0.70,
    ('PTAT', 'UDC'): 0.40, ('PTAT', 'FLC'): 0.30,
    ('NM$', 'TPI'): 0.85, ('NM$', 'CM$'): 0.95,
    ('MILK', 'FAT'): 0.55, ('MILK', 'PRO'): 0.85,
    ('PL', 'LIV'): 0.65, ('DPR', 'FI'): 0.70,
    ('SCE', 'SSB'): 0.60, ('DCE', 'DSB'): 0.55,
}
DOMINANCE_RATIOS = {
    'MILK': 0.12, 'FAT': 0.10, 'PRO': 0.12,
    'SCS': 0.14, 'PL': 0.15, 'DPR': 0.20, 'CCR': 0.15,
    'UDC': 0.11, 'FLC': 0.08, 'LIV': 0.12,
}

TRAIT_MAP = [
    ('TPI',  'GTPI',      'TPI',         'TPI'),
    ('NM$',  'NM$',       'Net Merit',   'NM$'),
    ('CM$',  'CM$',       'CM$',         'CM$'),
    ('MILK', 'MILK',      'PTA Milk',    'PTAM'),
    ('FAT',  'FAT',       'PTA Fat',     'PTAF'),
    ('FAT%', '%F',        '% Fat',       'PTAF%'),
    ('PRO',  'PRO',       'PTA Pro',     'PTAP'),
    ('PRO%', '%P',        '% Pro',       'PTAP%'),
    ('CFP',  'CFP',       'CFP',         'CFP'),
    ('PL',   'PL',        'PL',          'PL'),
    ('SCS',  'SCS',       'SCS',         'SCS'),
    ('DPR',  'DPR',       'PTA DPR',     'DPR'),
    ('LIV',  'LIV',       'PTA LIV',     'LIV'),
    ('FI',   'FI',        'Fertil Index', 'FI'),
    ('HCR',  'HCR',       'HCR',         'HCR'),
    ('CCR',  'CCR',       'CCR',         'CCR'),
    ('MAST', 'MAST',      'Mastitis',    'MAST'),
    ('FSAV', 'FSAV',      'Feed Saved',  'F SAV'),
    ('PTAT', 'PTAT',      'PTA Type',    'PTAT'),
    ('UDC',  'UDC',       'UDC',         'UDC'),
    ('FLC',  'FLC',       'FLC',         'FLC'),
    ('SCE',  'SCE',       'SCE',         'SCE'),
    ('DCE',  'DCE',       'DCE',         'DCE'),
    ('SSB',  'SSB',       'SSB',         'SSB'),
    ('DSB',  'DSB',       'DSB',         'DSB'),
]
ALL_TRAITS = [t[0] for t in TRAIT_MAP]


def build_features(sire_ptas, dam_ptas, trait):
    s = sire_ptas.get(trait)
    d = dam_ptas.get(trait)
    if s is None or d is None:
        return None
    f = {}
    sd = GENETIC_SD.get(trait, 1)
    h2 = HERITABILITY.get(trait, 0.15)
    pa = (s + d) / 2

    f['sire'] = s; f['dam'] = d; f['pa'] = pa
    f['delta_g'] = (s - d) / 2; f['diff'] = s - d
    f['sire_z'] = s / sd if sd else 0; f['dam_z'] = d / sd if sd else 0
    f['sire_sq'] = s * s; f['dam_sq'] = d * d; f['sxd'] = s * d
    f['h2_pa'] = pa * (0.5 + h2)

    ideal = BREED_IDEAL.get(trait)
    if ideal is not None:
        defic = max(0, ((d - ideal) / sd) if trait in LOWER_IS_BETTER else ((ideal - d) / sd))
        f['deficiency'] = defic; f['dg_x_def'] = f['delta_g'] * defic
    else:
        f['deficiency'] = 0; f['dg_x_def'] = 0

    f['ib_effect'] = INBREEDING_DEPRESSION.get(trait, 0) * 0.085
    dom = DOMINANCE_RATIOS.get(trait, 0)
    f['dom_pot'] = dom * abs(s - d) / sd if sd else 0

    for ot in ALL_TRAITS:
        if ot == trait: continue
        sv = sire_ptas.get(ot); dv = dam_ptas.get(ot)
        if sv is not None: f[f's_{ot}'] = sv
        if dv is not None: f[f'd_{ot}'] = dv

    for (t1, t2), corr in GENETIC_CORRELATIONS.items():
        if t1 == trait:
            ov = sire_ptas.get(t2); dov = dam_ptas.get(t2)
            if ov is not None and dov is not None:
                f[f'gc_{t2}'] = ((ov + dov) / 2) * corr
        elif t2 == trait:
            ov = sire_ptas.get(t1); dov = dam_ptas.get(t1)
            if ov is not None and dov is not None:
                f[f'gc_{t1}'] = ((ov + dov) / 2) * corr

    milk_pa = ((sire_ptas.get('MILK', 0) or 0) + (dam_ptas.get('MILK', 0) or 0)) / 2
    fat_pa = ((sire_ptas.get('FAT', 0) or 0) + (dam_ptas.get('FAT', 0) or 0)) / 2
    pro_pa = ((sire_ptas.get('PRO', 0) or 0) + (dam_ptas.get('PRO', 0) or 0)) / 2
    f['prod_press'] = milk_pa / 675 + fat_pa / 29 + pro_pa / 19
    f['sire_nm'] = sire_ptas.get('NM$', 0) or 0
    f['dam_nm'] = dam_ptas.get('NM$', 0) or 0

    return f


# ============================================================
# LOAD DATA
# ============================================================
print("=" * 100)
print("  DSII MODEL SHOOTOUT — 10 Architectures")
print("=" * 100)

print("\nLoading data...")
wb = openpyxl.load_workbook(DOWNLOADS / "May 2026.xlsx", read_only=True, data_only=True)
ws = wb.active
may_rows = list(ws.iter_rows(values_only=True))
wb.close()
hdr = [str(c).strip() if c else f'col_{i}' for i, c in enumerate(may_rows[0])]

dams = {}
for i in range(1, 21):
    fp = DOWNLOADS / f"DAM{i}.xlsx"
    if not fp.exists(): continue
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    ws = wb.active
    dr = list(ws.iter_rows(values_only=True))
    wb.close()
    dh = [str(c).strip() if c else '' for j, c in enumerate(dr[0])]
    ni = next(j for j, h in enumerate(dh) if 'naab' in h.lower() and 'code' in h.lower())
    for r in dr[1:]:
        if r[ni]: dams[str(r[ni]).strip()] = {dh[j]: r[j] for j in range(min(len(dh), len(r)))}

bulls_rn, bulls_nm = {}, {}
with open(BULLS_CSV, 'r', encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        rn = row.get('Registration Name', '').strip().upper()
        nm = row.get('Name', '').strip().upper()
        if rn: bulls_rn[rn] = row
        if nm: bulls_nm[nm] = row

records = []
for r in may_rows[1:]:
    rd = {hdr[i]: r[i] for i in range(min(len(hdr), len(r)))}
    sn = str(rd.get('SIRENAME', '')).strip().upper()
    sire = bulls_rn.get(sn) or bulls_nm.get(sn)
    if not sire: continue
    damreg = str(rd.get('DAMREGNUM2', '')).strip()
    dam = dams.get(damreg)
    if not dam: continue

    sire_ptas, dam_ptas = {}, {}
    rec = {'sirename': sn}
    for tn, dcol, mcol, bcol in TRAIT_MAP:
        dv = sf(rd.get(dcol)); sv = sf(sire.get(bcol))
        mv = sf(dam.get(mcol)) if mcol else None
        rec[f'd_{tn}'] = dv
        if sv is not None: sire_ptas[tn] = sv
        if mv is not None: dam_ptas[tn] = mv
    rec['_sp'] = sire_ptas; rec['_dp'] = dam_ptas
    records.append(rec)

df = pd.DataFrame(records)
print(f"  Trios: {len(df)}")

# ============================================================
# MODEL DEFINITIONS
# ============================================================
def get_models():
    return {
        'Ridge': Pipeline([('sc', StandardScaler()), ('m', Ridge(alpha=1.0))]),
        'BayesRidge': Pipeline([('sc', StandardScaler()), ('m', BayesianRidge())]),
        'ElasticNet': Pipeline([('sc', StandardScaler()),
                                ('m', ElasticNet(alpha=0.01, l1_ratio=0.3, max_iter=10000))]),
        'SVR': Pipeline([('sc', StandardScaler()),
                         ('m', SVR(kernel='rbf', C=10, epsilon=0.01, gamma='scale'))]),
        'MLP': Pipeline([('sc', StandardScaler()),
                         ('m', MLPRegressor(hidden_layer_sizes=(128, 64, 32),
                                           activation='relu', solver='adam',
                                           alpha=0.01, learning_rate='adaptive',
                                           max_iter=1000, early_stopping=True,
                                           validation_fraction=0.15,
                                           random_state=42))]),
        'RF': RandomForestRegressor(n_estimators=500, max_depth=12,
                                     min_samples_leaf=3, random_state=42, n_jobs=-1),
        'GBR': GradientBoostingRegressor(n_estimators=500, max_depth=5,
                                          learning_rate=0.03, min_samples_leaf=3,
                                          subsample=0.8, random_state=42),
        'XGB': XGBRegressor(n_estimators=500, max_depth=6, learning_rate=0.03,
                            subsample=0.8, colsample_bytree=0.7, reg_alpha=0.1,
                            reg_lambda=1.0, min_child_weight=3,
                            random_state=42, verbosity=0),
        'LGBM': LGBMRegressor(n_estimators=500, max_depth=6, learning_rate=0.03,
                               subsample=0.8, colsample_bytree=0.7,
                               min_child_samples=3, reg_alpha=0.1, reg_lambda=1.0,
                               random_state=42, verbose=-1),
        'CatBoost': CatBoostRegressor(iterations=500, depth=6, learning_rate=0.03,
                                       l2_leaf_reg=3, random_seed=42, verbose=0),
    }

# ============================================================
# SHOOTOUT
# ============================================================
kf = KFold(n_splits=5, shuffle=True, random_state=42)
model_names = list(get_models().keys())

# Header
hdr_str = f"{'Trait':>6} | {'N':>5} | {'PA':>6}"
for mn in model_names:
    hdr_str += f" | {mn:>9}"
hdr_str += " | Best"
print(f"\n{hdr_str}")
print("-" * len(hdr_str))

all_results = []
model_wins = defaultdict(int)

for tn, _, mcol, _ in TRAIT_MAP:
    if mcol is None: continue
    target = f'd_{tn}'
    mask = df[target].notna() & df.apply(lambda r: tn in r['_sp'] and tn in r['_dp'], axis=1)
    sub = df[mask].reset_index(drop=True)
    if len(sub) < 50: continue

    y = sub[target].values

    # Build features
    feat_rows = [build_features(row['_sp'], row['_dp'], tn) for _, row in sub.iterrows()]
    valid = [f is not None for f in feat_rows]
    feat_df = pd.DataFrame([f for f in feat_rows if f is not None])
    y_v = y[valid]

    # Sire consistency
    sub_v = sub[valid].reset_index(drop=True)
    sire_groups = defaultdict(list)
    for idx in range(len(sub_v)):
        sire_groups[sub_v.at[idx, 'sirename']].append((idx, y_v[idx]))

    sire_loo = np.full(len(y_v), np.nan)
    for sire, daughters in sire_groups.items():
        if len(daughters) < 2: continue
        for idx, val in daughters:
            others = [v for i, v in daughters if i != idx]
            sire_loo[idx] = np.mean(others)
    feat_df['sire_loo'] = sire_loo

    feat_df = feat_df.dropna(axis=1, thresh=int(len(feat_df) * 0.3))
    for c in feat_df.columns:
        if feat_df[c].isna().any():
            feat_df[c] = feat_df[c].fillna(feat_df[c].median())

    X = feat_df.values

    # PA
    pa_v = np.array([(sub_v.at[i, '_sp'][tn] + sub_v.at[i, '_dp'][tn]) / 2 for i in range(len(sub_v))])
    pa_r2 = r2_score(y_v, pa_v)

    # Test each model
    row_str = f"  {tn:>5} | {len(y_v):>5} | {pa_r2:>.4f}"
    trait_results = {'Trait': tn, 'N': len(y_v), 'PA_R2': round(pa_r2, 4)}
    best_r2, best_name = pa_r2, 'PA'

    models = get_models()
    for mn, model in models.items():
        try:
            r2s = []
            for tr, te in kf.split(X):
                model_clone = type(model)(**model.get_params()) if not isinstance(model, Pipeline) else Pipeline(
                    [(n, type(s)(**s.get_params()) if hasattr(s, 'get_params') else s) for n, s in model.steps])

                # For Pipeline, need to clone differently
                import copy
                m = copy.deepcopy(model)
                m.fit(X[tr], y_v[tr])
                pred = m.predict(X[te])
                r2s.append(r2_score(y_v[te], pred))

            avg_r2 = np.mean(r2s)
            trait_results[f'{mn}_R2'] = round(avg_r2, 4)
            row_str += f" | {avg_r2:>9.4f}"

            if avg_r2 > best_r2:
                best_r2 = avg_r2
                best_name = mn
        except Exception as e:
            trait_results[f'{mn}_R2'] = None
            row_str += f" |    ERROR"

    trait_results['Best'] = best_name
    trait_results['Best_R2'] = round(best_r2, 4)
    model_wins[best_name] += 1

    row_str += f" | {best_name}"
    print(row_str)
    all_results.append(trait_results)

# ============================================================
# SUMMARY
# ============================================================
print(f"\n{'='*100}")
print("  RESUMO POR MODELO")
print("=" * 100)

for mn in ['PA'] + model_names:
    if mn == 'PA':
        vals = [r['PA_R2'] for r in all_results]
    else:
        vals = [r.get(f'{mn}_R2', 0) for r in all_results if r.get(f'{mn}_R2') is not None]
    if vals:
        print(f"  {mn:>10}: R2 medio={np.mean(vals):.4f}  "
              f"min={min(vals):.4f}  max={max(vals):.4f}  "
              f"wins={model_wins.get(mn, 0)}")

print(f"\n  MELHOR MODELO POR TRAIT:")
for r in all_results:
    print(f"    {r['Trait']:>6}: {r['Best']:>10} R2={r['Best_R2']:.4f}")

# Best possible: pick best model per trait
best_per_trait = [r['Best_R2'] for r in all_results]
print(f"\n  ORACLE (melhor modelo por trait): R2 medio = {np.mean(best_per_trait):.4f}")

# Save
pd.DataFrame(all_results).to_csv(OUTPUT_DIR / "shootout_results.csv", index=False)
print(f"  Salvo em: {OUTPUT_DIR / 'shootout_results.csv'}")
