"""
Predição Genética DSII v3 — Felipe Santana
Roda a predição para IAs (sêmen) e Embriões já acasalados.
Fonte: 'Cópia de Predicao Genetica Felipe Santana 24 04 2026.xlsx'
"""

import os, math, csv, sys
from dataclasses import dataclass, field
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERRO: instale openpyxl: pip install openpyxl")
    sys.exit(1)

# ===========================================================================
# PARÂMETROS DO MODELO (mesmos do v3 original)
# ===========================================================================
INBREEDING_DEPRESSION = {
    "NM$": -25.0, "MILK": -28.5, "FAT": -1.0, "PROT": -0.9,
    "PL": -0.35, "DPR": -0.03, "SCS": +0.007, "LIV": -0.15,
    "CCR": -0.025, "HCR": -0.02,
}

GENETIC_SD = {
    "MILK": 675, "FAT": 29, "PROT": 19, "CFP": 25,
    "SCS": 0.14, "PL": 1.85, "LIV": 1.2, "HLIV": 0.8,
    "DPR": 1.3, "CCR": 1.65, "HCR": 1.4, "EFC": 5.0,
    "UDC": 0.75, "FLC": 0.65, "BWC": 1.0,
    "NM$": 275, "TPI": 250,
    "SCE": 1.5, "SSB": 1.0, "DSB": 1.0,
    "MF": 0.5, "DA": 0.5, "KET": 0.5, "MAST": 1.0, "MET": 0.5, "RP": 0.5,
    "STA": 1.0, "STR": 1.0, "BD": 1.0, "DF": 1.0, "RA": 1.0, "TW": 1.0,
    "RLS": 1.0, "RLR": 1.0, "FA": 1.0, "FLS": 1.0,
    "FUA": 1.0, "RUH": 1.0, "RUW": 1.0, "UC": 1.0, "UD": 1.0,
    "FTP": 1.0, "RTP": 1.0, "TL": 1.0,
    "TIPO": 0.75,
}

# Pesos econômicos NM$ 2024 (CDCB)
ECON_WEIGHTS = {
    "FAT": 2.60, "PROT": 1.40, "PL": 35.0, "SCS": -120.0,
    "DPR": 28.0, "LIV": 25.0, "UDC": 30.0, "BWC": -6.0,
    "CCR": 15.0, "HCR": 12.0, "MILK": 0.00, "FLC": 15.0,
    "MF": -10.0, "DA": -10.0, "KET": -10.0, "MAST": -15.0,
    "MET": -10.0, "RP": -10.0, "SCE": -20.0,
}

BREED_IDEAL = {
    "MILK": 1800, "FAT": 90, "PROT": 60, "CFP": 70,
    "DPR": 1.5, "CCR": 2.0, "HCR": 2.0, "PL": 5.0, "LIV": 2.0,
    "UDC": 1.5, "FLC": 1.0, "SCS": 2.60,
}

LOWER_IS_BETTER = {"SCS", "BWC", "SCE", "SSB", "DSB"}

CRITICAL_THRESHOLDS = {
    "DPR": (-2.0, "min"), "PL": (-1.0, "min"), "SCS": (3.30, "max"),
    "UDC": (-1.0, "min"), "FLC": (-1.0, "min"), "SCE": (5.0, "max"),
}

GENETIC_CORRELATIONS = {
    ("MILK", "DPR"): -0.35, ("MILK", "CCR"): -0.30,
    ("FAT", "PROT"): 0.65, ("DPR", "CCR"): 0.55,
    ("DPR", "PL"): 0.45, ("SCS", "PL"): -0.30,
    ("SCS", "UDC"): -0.25, ("MILK", "SCS"): 0.15,
}

DOMINANCE_RATIOS = {"DPR": 0.20, "PL": 0.15, "SCS": 0.14, "MILK": 0.12, "PROT": 0.12}

BASE_F = 0.0625  # 6.25% base


# ---------------------------------------------------------------------------
# HHP$ (Holistic Health Profit) — Select Sires formula
# ---------------------------------------------------------------------------
def compute_hhp(ptas):
    """Calcula HHP$ a partir dos PTAs preditos da cria."""
    fat = ptas.get("FAT", 0)
    prot = ptas.get("PROT", 0)
    pl = ptas.get("PL", 0)
    liv = ptas.get("LIV", 0)
    scs = ptas.get("SCS", 0)
    dpr = ptas.get("DPR", 0)
    ccr = ptas.get("CCR", 0)
    rfi = ptas.get("RFI", 0)
    sta = ptas.get("STA", 0)
    df = ptas.get("DF", 0)
    ruw = ptas.get("RUW", 0)
    ud = ptas.get("UD", 0)
    rtp = ptas.get("RTP", 0)
    tl = ptas.get("TL", 0)
    mast = ptas.get("MAST", 0)

    hhp = (
        4.91 * fat +
        6.01 * prot +
        12.83 * pl +
        10.69 * liv +
        (-158.56) * (scs - 3.0) +
        19.3 * dpr +
        15.84 * ccr +
        (-0.19) * rfi +
        (-13.32) * sta +
        (-8.88) * df +
        8.88 * ruw +
        13.32 * ud +
        (-14.8) * (abs(rtp) - 0.65) +
        (-26.64) * (abs(tl) - 0.5) +
        25.37 * mast
    )
    return round(hhp, 0)

# ===========================================================================
# DATA CLASSES
# ===========================================================================
@dataclass
class Animal:
    id: str = ""
    naab: str = ""
    name: str = ""
    num: str = ""
    reg: str = ""
    tpi: float = 0.0
    nm_dollar: float = 0.0
    sire_name: str = ""
    sire_naab: str = ""
    mgs_naab: str = ""
    ptas: dict = field(default_factory=dict)


@dataclass
class Prediction:
    sire: Animal = None
    dam: Animal = None
    tipo: str = ""  # "IA" ou "Embrião"

    nm_pa: float = 0.0
    tpi_pa: float = 0.0
    nm_corrected: float = 0.0
    tpi_corrected: float = 0.0
    tpi_ic_lower: float = 0.0
    tpi_ic_upper: float = 0.0
    delta_nm: float = 0.0
    delta_tpi: float = 0.0

    expected_f: float = 0.0
    endo_depression: float = 0.0

    delta_g: dict = field(default_factory=dict)
    trait_pred: dict = field(default_factory=dict)
    dsii: float = 0.0

    nm_sd: float = 0.0
    nm_ic_lower: float = 0.0
    nm_ic_upper: float = 0.0

    prob_above_avg: float = 0.0
    prob_top25: float = 0.0
    prob_top10: float = 0.0

    critical_penalty: float = 0.0
    antagonism_penalty: float = 0.0
    dominance_bonus: float = 0.0
    distance_to_ideal: float = 0.0

    hhp_dollar: float = 0.0

    final_score: float = 0.0
    rank: int = 0


# ===========================================================================
# UTILITY FUNCTIONS
# ===========================================================================
def safe_float(val, default=0.0):
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", ".")
    try:
        return float(s)
    except:
        return default


def normal_cdf(z):
    return 0.5 * (1.0 + math.erf(z / math.sqrt(2.0)))


# ===========================================================================
# LOAD DATA
# ===========================================================================
def load_data(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # --- TOUROS ---
    ws_t = wb["Prova dos touros"]
    bulls = {}
    for r in range(2, ws_t.max_row + 1):
        naab = str(ws_t.cell(r, 1).value or "").strip()
        if not naab:
            continue
        b = Animal()
        b.naab = naab
        b.id = naab
        b.name = str(ws_t.cell(r, 2).value or "")
        b.reg = str(ws_t.cell(r, 3).value or "")
        b.tpi = safe_float(ws_t.cell(r, 4).value)
        b.nm_dollar = safe_float(ws_t.cell(r, 5).value)
        b.ptas = {
            "MILK": safe_float(ws_t.cell(r, 8).value),
            "FAT": safe_float(ws_t.cell(r, 9).value),
            "PROT": safe_float(ws_t.cell(r, 11).value),
            "SCS": safe_float(ws_t.cell(r, 14).value),
            "DPR": safe_float(ws_t.cell(r, 15).value),
            "CCR": safe_float(ws_t.cell(r, 16).value),
            "PL": safe_float(ws_t.cell(r, 17).value),
            "LIV": safe_float(ws_t.cell(r, 18).value),
            "MF": safe_float(ws_t.cell(r, 20).value),
            "DA": safe_float(ws_t.cell(r, 21).value),
            "KET": safe_float(ws_t.cell(r, 22).value),
            "MAST": safe_float(ws_t.cell(r, 23).value),
            "MET": safe_float(ws_t.cell(r, 24).value),
            "RP": safe_float(ws_t.cell(r, 25).value),
            "HCR": safe_float(ws_t.cell(r, 26).value),
            "EFC": safe_float(ws_t.cell(r, 29).value),
            "TIPO": safe_float(ws_t.cell(r, 34).value),
            "UDC": safe_float(ws_t.cell(r, 35).value),
            "FLC": safe_float(ws_t.cell(r, 36).value),
            "STA": safe_float(ws_t.cell(r, 37).value),
            "STR": safe_float(ws_t.cell(r, 38).value),
            "BD": safe_float(ws_t.cell(r, 39).value),
            "DF": safe_float(ws_t.cell(r, 40).value),
            "RA": safe_float(ws_t.cell(r, 41).value),
            "TW": safe_float(ws_t.cell(r, 42).value),
            "RLS": safe_float(ws_t.cell(r, 43).value),
            "RLR": safe_float(ws_t.cell(r, 44).value),
            "FA": safe_float(ws_t.cell(r, 45).value),
            "FLS": safe_float(ws_t.cell(r, 46).value),
            "FUA": safe_float(ws_t.cell(r, 47).value),
            "RUH": safe_float(ws_t.cell(r, 48).value),
            "RUW": safe_float(ws_t.cell(r, 49).value),
            "UC": safe_float(ws_t.cell(r, 50).value),
            "UD": safe_float(ws_t.cell(r, 51).value),
            "FTP": safe_float(ws_t.cell(r, 52).value),
            "RTP": safe_float(ws_t.cell(r, 53).value),
            "TL": safe_float(ws_t.cell(r, 54).value),
            "SCE": safe_float(ws_t.cell(r, 55).value),
        }
        # Pedigree from col 58
        ped = str(ws_t.cell(r, 58).value or "")
        parts = [p.strip() for p in ped.split("x")] if "x" in ped else [ped]
        b.sire_name = parts[0] if len(parts) > 0 else ""
        b.mgs_naab = parts[1] if len(parts) > 1 else ""
        bulls[naab] = b

    # --- FÊMEAS ---
    ws_f = wb["Prova das f\u00eameas"]
    females = {}
    for r in range(2, ws_f.max_row + 1):
        num = ws_f.cell(r, 1).value
        if num is None:
            continue
        num_str = str(num).strip()
        d = Animal()
        d.id = num_str
        d.num = num_str
        d.naab = str(ws_f.cell(r, 3).value or "")
        d.sire_naab = d.naab
        d.sire_name = str(ws_f.cell(r, 5).value or "")
        d.mgs_naab = str(ws_f.cell(r, 6).value or "")
        d.nm_dollar = safe_float(ws_f.cell(r, 8).value)
        d.tpi = safe_float(ws_f.cell(r, 17).value)
        d.ptas = {
            "MILK": safe_float(ws_f.cell(r, 9).value),
            "FAT": safe_float(ws_f.cell(r, 10).value),
            "PROT": safe_float(ws_f.cell(r, 11).value),
            "SCS": safe_float(ws_f.cell(r, 12).value),
            "PL": safe_float(ws_f.cell(r, 14).value),
            "DPR": safe_float(ws_f.cell(r, 15).value),
            "TIPO": safe_float(ws_f.cell(r, 18).value),
            "LIV": safe_float(ws_f.cell(r, 25).value),
            "UDC": safe_float(ws_f.cell(r, 42).value),
            "FLC": safe_float(ws_f.cell(r, 41).value),
            "HCR": safe_float(ws_f.cell(r, 34).value),  # TCN (taxa concepcao novilha)
            "CCR": safe_float(ws_f.cell(r, 35).value),  # TCV (taxa concepcao vaca)
            "MF": safe_float(ws_f.cell(r, 27).value),
            "DA": safe_float(ws_f.cell(r, 28).value),
            "KET": safe_float(ws_f.cell(r, 29).value),
            "MAST": safe_float(ws_f.cell(r, 30).value),
            "MET": safe_float(ws_f.cell(r, 31).value),
            "RP": safe_float(ws_f.cell(r, 32).value),
            "SCE": safe_float(ws_f.cell(r, 33).value),
            "STA": safe_float(ws_f.cell(r, 43).value),
            "STR": safe_float(ws_f.cell(r, 44).value),
            "BD": safe_float(ws_f.cell(r, 45).value),
            "DF": safe_float(ws_f.cell(r, 46).value),
            "RA": safe_float(ws_f.cell(r, 47).value),
            "TW": safe_float(ws_f.cell(r, 48).value),
            "RLS": safe_float(ws_f.cell(r, 49).value),
            "RLR": safe_float(ws_f.cell(r, 50).value),
            "FA": safe_float(ws_f.cell(r, 51).value),
            "FLS": safe_float(ws_f.cell(r, 52).value),
            "FUA": safe_float(ws_f.cell(r, 53).value),
            "RUH": safe_float(ws_f.cell(r, 54).value),
            "RUW": safe_float(ws_f.cell(r, 55).value),
            "UC": safe_float(ws_f.cell(r, 56).value),
            "UD": safe_float(ws_f.cell(r, 57).value),
            "FTP": safe_float(ws_f.cell(r, 58).value),
            "RTP": safe_float(ws_f.cell(r, 59).value),
            "TL": safe_float(ws_f.cell(r, 60).value),
        }
        females[num_str] = d

    # --- IAs (Sêmen) ---
    # Try different sheet names for compatibility
    ia_sheet_name = "Semen" if "Semen" in wb.sheetnames else "IA"
    ws_ia = wb[ia_sheet_name]
    matings_ia = []
    for r in range(2, ws_ia.max_row + 1):
        fem_num = str(ws_ia.cell(r, 1).value or "").strip()
        bull_naab = str(ws_ia.cell(r, 4).value or "").strip()
        if fem_num and bull_naab:
            matings_ia.append((fem_num, bull_naab))

    # --- Embriões ---
    emb_sheet_name = "Embriao" if "Embriao" in wb.sheetnames else "Embrião"
    ws_emb = wb[emb_sheet_name]
    matings_emb = []
    for r in range(2, ws_emb.max_row + 1):
        fem_num = str(ws_emb.cell(r, 1).value or "").strip()  # N° animal (produto)
        mae_num = str(ws_emb.cell(r, 4).value or "").strip()  # Mãe
        bull_naab = str(ws_emb.cell(r, 6).value or "").strip()
        if mae_num and bull_naab and bull_naab != "None":
            matings_emb.append((fem_num, mae_num, bull_naab))

    return bulls, females, matings_ia, matings_emb


# ===========================================================================
# PREDICTION ENGINE (mesma lógica do v3)
# ===========================================================================
def compute_dam_deficiency(dam, pop_avgs):
    profile = {}
    for trait, avg_val in pop_avgs.items():
        dam_val = dam.ptas.get(trait, 0)
        sd = GENETIC_SD.get(trait, 1)
        if sd == 0:
            sd = 1
        if trait in LOWER_IS_BETTER:
            deficiency = (dam_val - avg_val) / sd
        else:
            deficiency = (avg_val - dam_val) / sd
        profile[trait] = max(0, deficiency)
    return profile


def compute_dsii(sire, dam, deficiency_profile):
    total = 0.0
    for trait in sire.ptas:
        if trait not in dam.ptas:
            continue
        delta_g = (sire.ptas[trait] - dam.ptas.get(trait, 0)) / 2.0
        if trait in LOWER_IS_BETTER:
            delta_g = -delta_g
        econ_w = abs(ECON_WEIGHTS.get(trait, 1.0))
        deficiency = deficiency_profile.get(trait, 0)
        amplification = 1.0 + min(2.0, deficiency * 0.8)
        contribution = delta_g * econ_w * amplification
        total += max(0, contribution)
    return total


def compute_critical_penalty(trait_pred):
    penalty = 0.0
    for trait, (thresh, direction) in CRITICAL_THRESHOLDS.items():
        val = trait_pred.get(trait, 0)
        if direction == "min" and val < thresh:
            penalty += (thresh - val) * 10
        elif direction == "max" and val > thresh:
            penalty += (val - thresh) * 10
    return penalty


def compute_antagonism(sire, dam):
    penalty = 0.0
    for (t1, t2), corr in GENETIC_CORRELATIONS.items():
        if corr >= 0:
            continue
        s1 = sire.ptas.get(t1, 0)
        s2 = sire.ptas.get(t2, 0)
        d1 = dam.ptas.get(t1, 0)
        d2 = dam.ptas.get(t2, 0)
        pa1 = (s1 + d1) / 2.0
        pa2 = (s2 + d2) / 2.0
        sd1 = GENETIC_SD.get(t1, 1)
        sd2 = GENETIC_SD.get(t2, 1)
        z1 = pa1 / sd1 if sd1 else 0
        z2 = pa2 / sd2 if sd2 else 0
        if z1 > 0.5 and z2 < -0.5:
            penalty += abs(corr) * abs(z1) * abs(z2) * 5
    return penalty


def compute_distance_to_ideal(trait_pred):
    dist_sq = 0.0
    count = 0
    for trait, ideal in BREED_IDEAL.items():
        val = trait_pred.get(trait, None)
        if val is None:
            continue
        sd = GENETIC_SD.get(trait, 1)
        if sd == 0:
            sd = 1
        if trait in LOWER_IS_BETTER:
            diff = (val - ideal) / sd
        else:
            diff = (ideal - val) / sd
        if diff > 0:
            dist_sq += diff ** 2
            count += 1
    return math.sqrt(dist_sq) if dist_sq > 0 else 0


def estimate_inbreeding(sire, dam):
    f = BASE_F
    # Simple pedigree check
    if sire.naab and dam.sire_naab and sire.naab == dam.sire_naab:
        f += 0.125
    if sire.naab and dam.mgs_naab and sire.naab == dam.mgs_naab:
        f += 0.0625
    if sire.sire_name and dam.sire_name and sire.sire_name == dam.sire_name:
        f += 0.03125
    return f


def predict_mating(sire, dam, deficiency_profile, pop_avg_nm, pop_sd_nm):
    p = Prediction(sire=sire, dam=dam)

    # 1. Parent Average
    p.nm_pa = (sire.nm_dollar + dam.nm_dollar) / 2.0
    p.tpi_pa = (sire.tpi + dam.tpi) / 2.0

    # 2. Endogamia
    p.expected_f = estimate_inbreeding(sire, dam)
    f_pct = p.expected_f * 100.0
    p.endo_depression = INBREEDING_DEPRESSION["NM$"] * f_pct
    p.nm_corrected = p.nm_pa + p.endo_depression

    # TPI
    tpi_endo = -20.0 * f_pct
    p.tpi_corrected = p.tpi_pa + tpi_endo
    p.delta_tpi = p.tpi_corrected - dam.tpi
    sigma2_tpi = GENETIC_SD["TPI"] ** 2
    var_tpi = 0.5 * (1.0 - p.expected_f) * sigma2_tpi + 0.20 * sigma2_tpi
    tpi_sd = math.sqrt(var_tpi)
    p.tpi_ic_lower = p.tpi_corrected - 1.96 * tpi_sd
    p.tpi_ic_upper = p.tpi_corrected + 1.96 * tpi_sd

    # 3. Trait predictions
    for trait in sire.ptas:
        if trait not in dam.ptas:
            continue
        pa = (sire.ptas[trait] + dam.ptas.get(trait, 0)) / 2.0
        dep = INBREEDING_DEPRESSION.get(trait, 0) * f_pct
        p.trait_pred[trait] = pa + dep

    # 4. DeltaG
    for trait in sire.ptas:
        d_val = dam.ptas.get(trait, 0)
        pa = p.trait_pred.get(trait, (sire.ptas[trait] + d_val) / 2.0)
        p.delta_g[trait] = pa - d_val

    p.delta_nm = p.nm_corrected - dam.nm_dollar

    # 4b. HHP$
    p.hhp_dollar = compute_hhp(p.trait_pred)

    # 5. DSII
    p.dsii = compute_dsii(sire, dam, deficiency_profile)

    # 6. Mendelian Sampling
    sigma2_nm = GENETIC_SD["NM$"] ** 2
    var_mendelian = 0.5 * (1.0 - p.expected_f) * sigma2_nm
    var_uncertainty = 0.20 * sigma2_nm
    total_var = var_mendelian + var_uncertainty
    p.nm_sd = math.sqrt(total_var)
    p.nm_ic_lower = p.nm_corrected - 1.96 * p.nm_sd
    p.nm_ic_upper = p.nm_corrected + 1.96 * p.nm_sd

    # 7. Probabilidades
    if p.nm_sd > 0:
        z_avg = (p.nm_corrected - pop_avg_nm) / p.nm_sd
        z_25 = (pop_avg_nm + 0.674 * pop_sd_nm - p.nm_corrected) / p.nm_sd
        z_10 = (pop_avg_nm + 1.282 * pop_sd_nm - p.nm_corrected) / p.nm_sd
        p.prob_above_avg = normal_cdf(z_avg)
        p.prob_top25 = 1.0 - normal_cdf(z_25)
        p.prob_top10 = 1.0 - normal_cdf(z_10)

    # 8. Penalidades
    p.critical_penalty = compute_critical_penalty(p.trait_pred)
    p.antagonism_penalty = compute_antagonism(sire, dam)

    # 9. Dominância
    dom_bonus = 0.0
    for trait, ratio in DOMINANCE_RATIOS.items():
        dom_bonus += ratio * (1.0 - p.expected_f) * 0.5
    p.dominance_bonus = dom_bonus

    # 10. Distância ao ideal
    p.distance_to_ideal = compute_distance_to_ideal(p.trait_pred)

    return p


def compute_final_scores(predictions):
    if not predictions:
        return

    # Normalizar componentes
    dsiis = [p.dsii for p in predictions]
    nms = [p.nm_corrected for p in predictions]
    dists = [p.distance_to_ideal for p in predictions]
    ants = [p.antagonism_penalty for p in predictions]
    crits = [p.critical_penalty for p in predictions]
    probs = [p.prob_top25 for p in predictions]

    def norm(vals, higher_better=True):
        mn, mx = min(vals), max(vals)
        rng = mx - mn if mx != mn else 1
        if higher_better:
            return [(v - mn) / rng * 100 for v in vals]
        else:
            return [(mx - v) / rng * 100 for v in vals]

    dsii_n = norm(dsiis)
    nm_n = norm(nms)
    dist_n = norm(dists, False)
    ant_n = norm(ants, False)
    crit_n = norm(crits, False)
    prob_n = norm(probs)

    for i, p in enumerate(predictions):
        p.final_score = (
            0.35 * dsii_n[i] +
            0.20 * nm_n[i] +
            0.15 * dist_n[i] +
            0.10 * ant_n[i] +
            0.10 * crit_n[i] +
            0.05 * (1.0 - p.expected_f) * 100 +
            0.05 * prob_n[i] +
            p.dominance_bonus
        )

    # Rank
    predictions.sort(key=lambda x: -x.final_score)
    for i, p in enumerate(predictions):
        p.rank = i + 1


# ===========================================================================
# MAIN
# ===========================================================================
def main():
    INPUT_FILE = os.path.join(os.path.dirname(__file__), "..",  "..", "Downloads",
                              "Cópia de Predicao Genetica Felipe Santana 24 04 2026.xlsx")
    if not os.path.exists(INPUT_FILE):
        INPUT_FILE = "C:/Users/DiegoGuerra/Downloads/Cópia de Predicao Genetica Felipe Santana 24 04 2026.xlsx"

    OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

    print("=" * 100)
    print("  PREDIÇÃO GENÉTICA DSII v3 — FELIPE SANTANA")
    print("=" * 100)

    bulls, females, matings_ia, matings_emb = load_data(INPUT_FILE)
    print(f"\n  Touros carregados: {len(bulls)}")
    print(f"  Fêmeas carregadas: {len(females)}")
    print(f"  IAs (sêmen): {len(matings_ia)}")
    print(f"  Embriões: {len(matings_emb)}")

    # Population averages (das fêmeas)
    all_traits = set()
    for f in females.values():
        all_traits.update(f.ptas.keys())

    pop_avgs = {}
    for trait in all_traits:
        vals = [f.ptas.get(trait, 0) for f in females.values() if trait in f.ptas]
        pop_avgs[trait] = sum(vals) / len(vals) if vals else 0

    all_nm = [f.nm_dollar for f in females.values()]
    pop_avg_nm = sum(all_nm) / len(all_nm) if all_nm else 0
    pop_sd_nm = (sum((x - pop_avg_nm) ** 2 for x in all_nm) / len(all_nm)) ** 0.5 if all_nm else 275

    print(f"\n  NM$ médio das fêmeas: ${pop_avg_nm:.0f} (SD=${pop_sd_nm:.0f})")

    # --- PREDIÇÃO IAs ---
    print(f"\n{'='*100}")
    print("  RODANDO PREDIÇÃO — IAs (SÊMEN)")
    print(f"{'='*100}")

    predictions_ia = []
    skipped_ia = 0
    for fem_num, bull_naab in matings_ia:
        if fem_num not in females:
            skipped_ia += 1
            continue
        if bull_naab not in bulls:
            skipped_ia += 1
            continue
        dam = females[fem_num]
        sire = bulls[bull_naab]
        deficiency = compute_dam_deficiency(dam, pop_avgs)
        p = predict_mating(sire, dam, deficiency, pop_avg_nm, pop_sd_nm)
        p.tipo = "IA"
        predictions_ia.append(p)

    compute_final_scores(predictions_ia)
    print(f"  Predições calculadas: {len(predictions_ia)}")
    print(f"  Combinações sem dados (puladas): {skipped_ia}")

    # --- PREDIÇÃO EMBRIÕES ---
    print(f"\n{'='*100}")
    print("  RODANDO PREDIÇÃO — EMBRIÕES")
    print(f"{'='*100}")

    predictions_emb = []
    skipped_emb = 0
    for prod_num, mae_num, bull_naab in matings_emb:
        if mae_num not in females:
            skipped_emb += 1
            continue
        if bull_naab not in bulls:
            skipped_emb += 1
            continue
        dam = females[mae_num]
        sire = bulls[bull_naab]
        deficiency = compute_dam_deficiency(dam, pop_avgs)
        p = predict_mating(sire, dam, deficiency, pop_avg_nm, pop_sd_nm)
        p.tipo = "Embrião"
        # Store product num
        p.dam = Animal(id=dam.id, num=dam.num, naab=dam.naab, name=dam.name,
                       tpi=dam.tpi, nm_dollar=dam.nm_dollar, sire_name=dam.sire_name,
                       sire_naab=dam.sire_naab, mgs_naab=dam.mgs_naab, ptas=dam.ptas)
        p.dam.reg = prod_num  # Use reg field to store product number
        predictions_emb.append(p)

    compute_final_scores(predictions_emb)
    print(f"  Predições calculadas: {len(predictions_emb)}")
    print(f"  Combinações sem dados (puladas): {skipped_emb}")

    # --- EXPORT CSV ---
    all_traits_list = ["MILK", "FAT", "PROT", "SCS", "DPR", "CCR", "PL", "LIV",
                       "HCR", "EFC", "UDC", "FLC", "BWC", "TIPO",
                       "MF", "DA", "KET", "MAST", "MET", "RP", "SCE",
                       "STA", "STR", "BD", "DF", "RA", "TW", "RLS", "RLR", "FA", "FLS",
                       "FUA", "RUH", "RUW", "UC", "UD", "FTP", "RTP", "TL"]

    def export_csv(predictions, filename, is_embriao=False):
        fields = [
            "Rank", "Tipo", "Touro_NAAB", "Touro_Nome", "Touro_TPI", "Touro_NM$",
            "Femea_Num", "Pai_da_Femea",
        ]
        if is_embriao:
            fields.insert(7, "Produto_Num")
        fields += [
            "Femea_TPI", "Femea_NM$",
            "NM$_PA", "NM$_Corrigido", "HHP$", "NM$_IC_Lower", "NM$_IC_Upper", "Delta_NM$_sobre_Mae",
            "TPI_PA", "TPI_Corrigido", "TPI_IC_Lower", "TPI_IC_Upper", "Delta_TPI_sobre_Mae",
            "DSII", "Score_Final_v3",
            "F_Esperada_%", "Depressão_$",
            "Penalidade_Crítica", "Penalidade_Antagonismo",
            "Bônus_Dominância", "Distância_Ideal",
            "P_Acima_Média_%", "P_Top25_%", "P_Top10_%",
        ]
        for t in all_traits_list:
            fields.extend([f"PA_{t}", f"DeltaG_{t}"])

        csv_path = os.path.join(OUTPUT_DIR, filename)
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=fields, delimiter=";")
            w.writeheader()
            for p in predictions:
                row = {
                    "Rank": p.rank,
                    "Tipo": p.tipo,
                    "Touro_NAAB": p.sire.naab,
                    "Touro_Nome": p.sire.name,
                    "Touro_TPI": f"{p.sire.tpi:.0f}",
                    "Touro_NM$": f"{p.sire.nm_dollar:.0f}",
                    "Femea_Num": p.dam.num,
                    "Pai_da_Femea": p.dam.sire_name,
                    "Femea_TPI": f"{p.dam.tpi:.0f}",
                    "Femea_NM$": f"{p.dam.nm_dollar:.0f}",
                    "NM$_PA": f"{p.nm_pa:.0f}",
                    "NM$_Corrigido": f"{p.nm_corrected:.0f}",
                    "HHP$": f"{p.hhp_dollar:.0f}",
                    "NM$_IC_Lower": f"{p.nm_ic_lower:.0f}",
                    "NM$_IC_Upper": f"{p.nm_ic_upper:.0f}",
                    "Delta_NM$_sobre_Mae": f"{p.delta_nm:.0f}",
                    "TPI_PA": f"{p.tpi_pa:.0f}",
                    "TPI_Corrigido": f"{p.tpi_corrected:.0f}",
                    "TPI_IC_Lower": f"{p.tpi_ic_lower:.0f}",
                    "TPI_IC_Upper": f"{p.tpi_ic_upper:.0f}",
                    "Delta_TPI_sobre_Mae": f"{p.delta_tpi:.0f}",
                    "DSII": f"{p.dsii:.1f}",
                    "Score_Final_v3": f"{p.final_score:.1f}",
                    "F_Esperada_%": f"{p.expected_f * 100:.2f}",
                    "Depressão_$": f"{p.endo_depression:.0f}",
                    "Penalidade_Crítica": f"{p.critical_penalty:.1f}",
                    "Penalidade_Antagonismo": f"{p.antagonism_penalty:.1f}",
                    "Bônus_Dominância": f"{p.dominance_bonus:.2f}",
                    "Distância_Ideal": f"{p.distance_to_ideal:.2f}",
                    "P_Acima_Média_%": f"{p.prob_above_avg * 100:.1f}",
                    "P_Top25_%": f"{p.prob_top25 * 100:.1f}",
                    "P_Top10_%": f"{p.prob_top10 * 100:.1f}",
                }
                if is_embriao:
                    row["Produto_Num"] = p.dam.reg
                for t in all_traits_list:
                    row[f"PA_{t}"] = f"{p.trait_pred.get(t, 0):.2f}"
                    row[f"DeltaG_{t}"] = f"{p.delta_g.get(t, 0):.2f}"
                w.writerow(row)
        return csv_path

    csv_ia = export_csv(predictions_ia, "predicao_felipe_IA.csv")
    csv_emb = export_csv(predictions_emb, "predicao_felipe_embriao.csv", is_embriao=True)

    print(f"\n{'='*100}")
    print(f"  EXPORTADO:")
    print(f"    IA:      {csv_ia}")
    print(f"    Embrião: {csv_emb}")
    print(f"{'='*100}")

    # Resumo
    print(f"\n  === RESUMO IAs (Top 10) ===")
    for p in predictions_ia[:10]:
        print(f"    #{p.rank:>3} | {p.sire.name:<12} x Fêmea #{p.dam.num:<5} | "
              f"Score={p.final_score:>5.1f} | NM$={p.nm_corrected:>5.0f} | "
              f"TPI={p.tpi_corrected:>5.0f} | DSII={p.dsii:>6.0f}")

    print(f"\n  === RESUMO EMBRIÕES (Top 10) ===")
    for p in predictions_emb[:10]:
        print(f"    #{p.rank:>3} | {p.sire.name:<12} x Mãe #{p.dam.num:<5} (Prod #{p.dam.reg}) | "
              f"Score={p.final_score:>5.1f} | NM$={p.nm_corrected:>5.0f} | "
              f"TPI={p.tpi_corrected:>5.0f} | DSII={p.dsii:>6.0f}")

    # Stats por touro
    print(f"\n  === PERFORMANCE POR TOURO (IAs) ===")
    by_bull = defaultdict(list)
    for p in predictions_ia:
        by_bull[p.sire.name].append(p)
    for bname in sorted(by_bull.keys(), key=lambda x: -sum(p.final_score for p in by_bull[x])/len(by_bull[x])):
        preds = by_bull[bname]
        scores = [p.final_score for p in preds]
        nms = [p.nm_corrected for p in preds]
        print(f"    {bname:<14} | {len(preds):>3} IAs | Score={sum(scores)/len(scores):>5.1f} "
              f"(min={min(scores):.1f} max={max(scores):.1f}) | "
              f"NM$={sum(nms)/len(nms):>5.0f}")


if __name__ == "__main__":
    main()
