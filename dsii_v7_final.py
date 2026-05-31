"""
DSII v7 FINAL — Full Hybrid: V5 Domain Knowledge + CDCB Learned Distributions
===============================================================================
Integrates everything learned:
  - V5 feature engineering (deficiency, DeltaG, antagonism, interactions)
  - V5 sire consistency (leave-one-out from May2026)
  - CDCB transfer coefficients (alpha/beta from 121k sire-daughter pairs)
  - CDCB population distributions (percentiles, bounds, regression-to-mean)
  - CDCB sire performance stats (mean/std of daughters per sire from 130k cows)
  - Stacking ensemble (LightGBM + Ridge + RF)
"""
import warnings
warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd
import openpyxl, csv
from pathlib import Path
from collections import defaultdict

from sklearn.model_selection import KFold
from sklearn.linear_model import Ridge
from sklearn.ensemble import RandomForestRegressor, StackingRegressor
from sklearn.metrics import mean_absolute_error, r2_score
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from lightgbm import LGBMRegressor

DOWNLOADS = Path("C:/Users/DiegoGuerra/Downloads")
BULLS_CSV = Path("C:/Users/DiegoGuerra/gen-genie-advisor/sms-engine/data/bulls.csv")
OUTPUT_DIR = Path("C:/Users/DiegoGuerra/Documents/Projetos/genetic_prediction/dsii_v7_results")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

def sf(v):
    if v is None: return None
    v = str(v).strip().strip('"')
    if not v: return None
    try:
        x = float(v)
        return x if not np.isnan(x) else None
    except: return None

def read_csv_safe(path):
    for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
        try:
            with open(path, 'r', encoding=enc) as f:
                return list(csv.DictReader(f))
        except: continue

# ============================================================
# DOMAIN KNOWLEDGE (from V3/V5)
# ============================================================
GENETIC_SD = {
    'TPI': 250, 'NM$': 275, 'CM$': 280,
    'MILK': 675, 'FAT': 29, 'FAT%': 0.05, 'PRO': 19, 'PRO%': 0.02, 'CFP': 25,
    'PL': 1.85, 'SCS': 0.14, 'DPR': 1.3, 'HCR': 1.4, 'CCR': 1.65,
    'LIV': 1.2, 'FI': 1.0, 'MAST': 1.0, 'FSAV': 50,
    'PTAT': 0.70, 'UDC': 0.75, 'FLC': 0.65,
    'SCE': 1.5, 'DCE': 1.2, 'SSB': 1.0, 'DSB': 1.0,
}

HERITABILITY = {
    'TPI': 0.30, 'NM$': 0.30, 'CM$': 0.30,
    'MILK': 0.25, 'FAT': 0.25, 'FAT%': 0.50, 'PRO': 0.25, 'PRO%': 0.50,
    'CFP': 0.30, 'PL': 0.08, 'SCS': 0.12, 'DPR': 0.04, 'LIV': 0.05,
    'FI': 0.06, 'HCR': 0.04, 'CCR': 0.04, 'MAST': 0.04, 'FSAV': 0.15,
    'PTAT': 0.30, 'UDC': 0.25, 'FLC': 0.15,
    'SCE': 0.08, 'DCE': 0.06, 'SSB': 0.06, 'DSB': 0.04,
}

BREED_IDEAL = {
    'MILK': 1800, 'FAT': 90, 'PRO': 60, 'CFP': 70,
    'DPR': 1.5, 'CCR': 2.0, 'HCR': 2.0,
    'PL': 5.0, 'LIV': 2.0, 'UDC': 1.5, 'FLC': 1.0,
    'MAST': -1.0, 'SCS': 2.60, 'SCE': 2.0,
}

LOWER_IS_BETTER = {'SCS', 'SCE', 'SSB', 'DSB', 'DCE', 'MAST'}

INBREEDING_DEPRESSION = {
    'NM$': -25.0, 'TPI': -20.0, 'CM$': -25.0,
    'MILK': -28.5, 'FAT': -1.0, 'PRO': -0.9,
    'PL': -0.35, 'DPR': -0.03, 'SCS': 0.007, 'LIV': -0.15,
    'CCR': -0.025, 'HCR': -0.02, 'CFP': -0.8,
}

GENETIC_CORRELATIONS = {
    ('MILK', 'DPR'): -0.35, ('MILK', 'CCR'): -0.30, ('MILK', 'SCS'): 0.10,
    ('MILK', 'PL'): -0.15, ('MILK', 'UDC'): -0.15,
    ('FAT', 'PRO'): 0.65, ('FAT', 'DPR'): -0.25, ('PRO', 'DPR'): -0.30,
    ('DPR', 'CCR'): 0.55, ('DPR', 'PL'): 0.45, ('CCR', 'PL'): 0.30,
    ('UDC', 'PL'): 0.25, ('SCS', 'PL'): -0.30, ('SCS', 'MAST'): 0.70,
    ('PTAT', 'MILK'): 0.15, ('PTAT', 'UDC'): 0.40, ('PTAT', 'FLC'): 0.30,
    ('NM$', 'TPI'): 0.85, ('NM$', 'CM$'): 0.95,
    ('MILK', 'FAT'): 0.55, ('MILK', 'PRO'): 0.85,
    ('PL', 'LIV'): 0.65, ('DPR', 'FI'): 0.70,
    ('SCE', 'SSB'): 0.60, ('DCE', 'DSB'): 0.55,
}

DOMINANCE_RATIOS = {
    'MILK': 0.12, 'FAT': 0.10, 'PRO': 0.12,
    'SCS': 0.14, 'PL': 0.15, 'DPR': 0.20, 'CCR': 0.15,
    'UDC': 0.11, 'FLC': 0.08, 'LIV': 0.12,
}

# Trait mapping
TRAIT_MAP = [
    ('TPI',  'GTPI',      'TPI',         'TPI'),
    ('NM$',  'NM$',       'Net Merit',   'NM$'),
    ('CM$',  'CM$',       'CM$',         'CM$'),
    ('MILK', 'MILK',      'PTA Milk',    'PTAM'),
    ('FAT',  'FAT',       'PTA Fat',     'PTAF'),
    ('FAT%', '%F',        '% Fat',       'PTAF%'),
    ('PRO',  'PRO',       'PTA Pro',     'PTAP'),
    ('PRO%', '%P',        '% Pro',       'PTAP%'),
    ('CFP',  'CFP',       'CFP',         'CFP'),
    ('PL',   'PL',        'PL',          'PL'),
    ('SCS',  'SCS',       'SCS',         'SCS'),
    ('DPR',  'DPR',       'PTA DPR',     'DPR'),
    ('LIV',  'LIV',       'PTA LIV',     'LIV'),
    ('FI',   'FI',        'Fertil Index', 'FI'),
    ('HCR',  'HCR',       'HCR',         'HCR'),
    ('CCR',  'CCR',       'CCR',         'CCR'),
    ('MAST', 'MAST',      'Mastitis',    'MAST'),
    ('FSAV', 'FSAV',      'Feed Saved',  'F SAV'),
    ('PTAT', 'PTAT',      'PTA Type',    'PTAT'),
    ('UDC',  'UDC',       'UDC',         'UDC'),
    ('FLC',  'FLC',       'FLC',         'FLC'),
    ('SCE',  'SCE',       'SCE',         'SCE'),
    ('DCE',  'DCE',       'DCE',         'DCE'),
    ('SSB',  'SSB',       'SSB',         'SSB'),
    ('DSB',  'DSB',       'DSB',         'DSB'),
]
ALL_TRAITS = [t[0] for t in TRAIT_MAP]

# CDCB column mapping (for learning distributions)
CDCB_COL = {
    'NM$': 'NM$_PTA', 'CM$': 'CM$_PTA', 'MILK': 'MLK_PTA',
    'FAT': 'FAT_PTA', 'PRO': 'PRO_PTA', 'PL': 'PL_PTA',
    'SCS': 'SCS_PTA', 'DPR': 'DPR_PTA', 'HCR': 'HCR_PTA',
    'CCR': 'CCR_PTA', 'LIV': 'LIV_PTA', 'MAST': 'MAS_PTA', 'FSAV': 'FS_PTA',
}


# ============================================================
# PHASE 1: Learn from CDCB
# ============================================================
def learn_from_cdcb():
    print("=" * 90)
    print("  PHASE 1: Learning from CDCB (130k cows + 92k bulls)")
    print("=" * 90)

    cow_rows = read_csv_safe(DOWNLOADS / "Cow_Report (4).csv")
    bull_rows = read_csv_safe(DOWNLOADS / "Bull_Report (1).csv")
    print(f"  Cows: {len(cow_rows)}, Bulls: {len(bull_rows)}")

    # Cow (daughter) distributions
    cow_dist = {}
    for trait, col in CDCB_COL.items():
        vals = [sf(r.get(col)) for r in cow_rows]
        vals = np.array([v for v in vals if v is not None])
        cow_dist[trait] = {
            'mean': np.mean(vals), 'std': np.std(vals), 'median': np.median(vals),
            'p5': np.percentile(vals, 5), 'p10': np.percentile(vals, 10),
            'p25': np.percentile(vals, 25), 'p75': np.percentile(vals, 75),
            'p90': np.percentile(vals, 90), 'p95': np.percentile(vals, 95),
            'iqr': np.percentile(vals, 75) - np.percentile(vals, 25),
            'skew': float(pd.Series(vals).skew()),
        }

    # Bull (sire) distributions
    bull_dist = {}
    for trait, col in CDCB_COL.items():
        vals = [sf(r.get(col)) for r in bull_rows]
        vals = np.array([v for v in vals if v is not None])
        bull_dist[trait] = {
            'mean': np.mean(vals), 'std': np.std(vals), 'median': np.median(vals),
            'p25': np.percentile(vals, 25), 'p75': np.percentile(vals, 75),
            'iqr': np.percentile(vals, 75) - np.percentile(vals, 25),
        }

    # Sire->Daughter transfer coefficients
    bull_idx = {r['ANIMAL'].strip('"').upper(): r for r in bull_rows}
    sire_daughter_vals = defaultdict(lambda: defaultdict(list))
    for r in cow_rows:
        sid = r['SIRE'].strip('"').upper()
        for trait, col in CDCB_COL.items():
            dv = sf(r.get(col))
            if dv is not None:
                sire_daughter_vals[sid][trait].append(dv)

    transfer = {}
    for trait in CDCB_COL:
        sp, dm = [], []
        for sid, tv in sire_daughter_vals.items():
            if trait not in tv or len(tv[trait]) < 5: continue
            sire = bull_idx.get(sid)
            if not sire: continue
            sv = sf(sire.get(CDCB_COL[trait]))
            if sv is None: continue
            sp.append(sv); dm.append(np.mean(tv[trait]))
        if len(sp) > 10:
            sp, dm = np.array(sp), np.array(dm)
            A = np.vstack([sp, np.ones(len(sp))]).T
            alpha, beta = np.linalg.lstsq(A, dm, rcond=None)[0]
            transfer[trait] = {'alpha': alpha, 'beta': beta,
                             'residual_std': np.std(dm - (alpha * sp + beta))}
            print(f"  {trait:>6}: alpha={alpha:.3f} beta={beta:>7.1f}")

    # Sire stats from CDCB (for sire consistency enrichment)
    sire_stats_cdcb = {}
    for sid, tv in sire_daughter_vals.items():
        stats = {}
        for trait, vals in tv.items():
            if len(vals) >= 3:
                stats[trait] = {
                    'mean': np.mean(vals), 'std': np.std(vals),
                    'median': np.median(vals), 'n': len(vals),
                    'p25': np.percentile(vals, 25), 'p75': np.percentile(vals, 75),
                }
        sire_stats_cdcb[sid] = stats

    print(f"  Sires with stats: {len(sire_stats_cdcb)}")

    return cow_dist, bull_dist, transfer, sire_stats_cdcb


# ============================================================
# PHASE 2: Build V7 features
# ============================================================
def build_v7_features(sire_ptas, dam_ptas, trait, cow_dist, bull_dist,
                      transfer, sire_stats_cdcb, sire_reg=None):
    """All features from V5 + CDCB knowledge."""
    s = sire_ptas.get(trait)
    d = dam_ptas.get(trait)
    if s is None or d is None:
        return None

    f = {}
    sd = GENETIC_SD.get(trait, 1)
    h2 = HERITABILITY.get(trait, 0.15)
    pa = (s + d) / 2

    # --- V5 core features ---
    f['sire'] = s
    f['dam'] = d
    f['pa'] = pa
    f['delta_g'] = (s - d) / 2
    f['sire_dam_diff'] = s - d
    f['sire_z'] = s / sd if sd > 0 else 0
    f['dam_z'] = d / sd if sd > 0 else 0
    f['pa_z'] = pa / sd if sd > 0 else 0
    f['sire_sq'] = s * s
    f['dam_sq'] = d * d
    f['sire_x_dam'] = s * d
    f['h2_weighted_pa'] = pa * (0.5 + h2)

    # Dam deficiency
    ideal = BREED_IDEAL.get(trait)
    if ideal is not None:
        if trait in LOWER_IS_BETTER:
            deficiency = max(0, (d - ideal) / sd)
        else:
            deficiency = max(0, (ideal - d) / sd)
        f['dam_deficiency'] = deficiency
        f['delta_g_x_deficiency'] = f['delta_g'] * deficiency
    else:
        f['dam_deficiency'] = 0
        f['delta_g_x_deficiency'] = 0

    # Inbreeding
    f['inbreeding_effect'] = INBREEDING_DEPRESSION.get(trait, 0) * 0.085

    # Dominance
    dom = DOMINANCE_RATIOS.get(trait, 0)
    f['dominance_potential'] = dom * abs(s - d) / sd if sd > 0 else 0

    # --- V5 cross-trait features ---
    for other in ALL_TRAITS:
        if other == trait: continue
        sv = sire_ptas.get(other)
        dv = dam_ptas.get(other)
        if sv is not None:
            f[f'raw_s_{other}'] = sv
        if dv is not None:
            f[f'raw_d_{other}'] = dv

    # Genetic correlation weighted features
    for (t1, t2), corr in GENETIC_CORRELATIONS.items():
        if t1 == trait:
            ov = sire_ptas.get(t2)
            dov = dam_ptas.get(t2)
            if ov is not None and dov is not None:
                f[f'corr_{t2}_pa_w'] = ((ov + dov) / 2) * corr
        elif t2 == trait:
            ov = sire_ptas.get(t1)
            dov = dam_ptas.get(t1)
            if ov is not None and dov is not None:
                f[f'corr_{t1}_pa_w'] = ((ov + dov) / 2) * corr

    # Production pressure (for health traits)
    milk_pa = ((sire_ptas.get('MILK', 0) or 0) + (dam_ptas.get('MILK', 0) or 0)) / 2
    fat_pa = ((sire_ptas.get('FAT', 0) or 0) + (dam_ptas.get('FAT', 0) or 0)) / 2
    pro_pa = ((sire_ptas.get('PRO', 0) or 0) + (dam_ptas.get('PRO', 0) or 0)) / 2
    f['production_pressure'] = milk_pa / 675 + fat_pa / 29 + pro_pa / 19

    # Economic value
    econ_s = (sire_ptas.get('NM$', 0) or 0)
    econ_d = (dam_ptas.get('NM$', 0) or 0)
    f['sire_nm'] = econ_s
    f['dam_nm'] = econ_d

    # --- CDCB learned features ---
    cd = cow_dist.get(trait)
    bd = bull_dist.get(trait)
    tc = transfer.get(trait)

    if cd:
        # Where does this PA sit in the CDCB population?
        f['pa_cdcb_z'] = (pa - cd['mean']) / cd['std'] if cd['std'] > 0 else 0
        f['pa_cdcb_percentile_approx'] = (pa - cd['p5']) / (cd['p95'] - cd['p5']) if (cd['p95'] - cd['p5']) > 0 else 0.5
        f['pa_in_upper_tail'] = float(pa > cd['p90'])
        f['pa_in_lower_tail'] = float(pa < cd['p10'])
        f['pa_extreme'] = abs(pa - cd['median']) / cd['iqr'] if cd['iqr'] > 0 else 0
        # Regression to mean: extreme values should regress
        f['regression_to_mean'] = (cd['median'] - pa) * (1 - h2) * 0.1

    if bd:
        f['sire_cdcb_z'] = (s - bd['mean']) / bd['std'] if bd['std'] > 0 else 0
        f['dam_cdcb_z'] = (d - cd['mean']) / cd['std'] if cd and cd['std'] > 0 else 0

    if tc:
        # CDCB-learned transfer prediction
        f['cdcb_transfer_pred'] = tc['alpha'] * s + tc['beta']
        f['cdcb_alpha'] = tc['alpha']
        # Deviation from simple sire/2
        f['transfer_vs_half'] = (tc['alpha'] * s + tc['beta']) - (s / 2)

    # CDCB sire performance stats
    if sire_reg and sire_reg in sire_stats_cdcb:
        ss = sire_stats_cdcb[sire_reg].get(trait)
        if ss:
            f['cdcb_sire_daughter_mean'] = ss['mean']
            f['cdcb_sire_daughter_std'] = ss['std']
            f['cdcb_sire_daughter_n'] = ss['n']
            f['cdcb_sire_daughter_median'] = ss['median']

    return f


# ============================================================
# MAIN PIPELINE
# ============================================================
def run_v7():
    # Phase 1: Learn from CDCB
    cow_dist, bull_dist, transfer, sire_stats_cdcb = learn_from_cdcb()

    # Phase 2: Load May2026 trios
    print(f"\n{'='*90}")
    print("  PHASE 2: Loading May2026 trios")
    print("=" * 90)

    wb = openpyxl.load_workbook(DOWNLOADS / "May 2026.xlsx", read_only=True, data_only=True)
    ws = wb.active
    may_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = [str(c).strip() if c else f'col_{i}' for i, c in enumerate(may_rows[0])]

    dams = {}
    for i in range(1, 21):
        fp = DOWNLOADS / f"DAM{i}.xlsx"
        if not fp.exists(): continue
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        ws = wb.active
        dr = list(ws.iter_rows(values_only=True))
        wb.close()
        dh = [str(c).strip() if c else '' for j, c in enumerate(dr[0])]
        ni = next(j for j, h in enumerate(dh) if 'naab' in h.lower() and 'code' in h.lower())
        for r in dr[1:]:
            if r[ni]: dams[str(r[ni]).strip()] = {dh[j]: r[j] for j in range(min(len(dh), len(r)))}

    bulls_rn, bulls_nm = {}, {}
    with open(BULLS_CSV, 'r', encoding='utf-8-sig') as f:
        for row in csv.DictReader(f):
            rn = row.get('Registration Name', '').strip().upper()
            nm = row.get('Name', '').strip().upper()
            if rn: bulls_rn[rn] = row
            if nm: bulls_nm[nm] = row

    records = []
    for r in may_rows[1:]:
        rd = {hdr[i]: r[i] for i in range(min(len(hdr), len(r)))}
        sn = str(rd.get('SIRENAME', '')).strip().upper()
        sire = bulls_rn.get(sn) or bulls_nm.get(sn)
        if not sire: continue
        damreg = str(rd.get('DAMREGNUM2', '')).strip()
        dam = dams.get(damreg)
        if not dam: continue

        sire_reg = str(rd.get('SIREREGNUM2', '')).strip().upper()
        sire_ptas, dam_ptas = {}, {}
        rec = {'sire_reg': sire_reg, 'sirename': sn}
        for tn, dcol, mcol, bcol in TRAIT_MAP:
            dv = sf(rd.get(dcol))
            sv = sf(sire.get(bcol))
            mv = sf(dam.get(mcol)) if mcol else None
            rec[f'd_{tn}'] = dv
            if sv is not None: sire_ptas[tn] = sv
            if mv is not None: dam_ptas[tn] = mv
        rec['_sp'] = sire_ptas
        rec['_dp'] = dam_ptas
        records.append(rec)

    df = pd.DataFrame(records)
    print(f"  Trios: {len(df)}")

    # Phase 3: Train V7
    print(f"\n{'='*90}")
    print("  PHASE 3: Training V7 (V5 + CDCB Knowledge)")
    print("=" * 90)

    kf = KFold(n_splits=5, shuffle=True, random_state=42)

    print(f"\n{'Trait':>6} | {'N':>5} | {'PA R2':>7} | {'V5 R2':>7} | {'V7 R2':>7} | "
          f"{'V5 MAE':>8} {'V7 MAE':>8} | {'V5vPA':>6} {'V7vPA':>6} {'V7vV5':>6}")
    print("-" * 105)

    all_results = []

    for tn, _, mcol, _ in TRAIT_MAP:
        if mcol is None: continue
        target = f'd_{tn}'
        mask = df[target].notna()
        # Also need sire and dam for this trait
        mask = mask & df.apply(lambda r: tn in r['_sp'] and tn in r['_dp'], axis=1)
        sub = df[mask].reset_index(drop=True)
        if len(sub) < 50: continue

        y = sub[target].values

        # PA baseline
        pa_vals = np.array([(r['_sp'][tn] + r['_dp'][tn]) / 2 for _, r in sub.iterrows()])
        pa_r2 = r2_score(y, pa_vals)
        pa_mae = mean_absolute_error(y, pa_vals)

        # Build V7 features
        feat_rows = []
        for _, row in sub.iterrows():
            feats = build_v7_features(
                row['_sp'], row['_dp'], tn,
                cow_dist, bull_dist, transfer, sire_stats_cdcb,
                sire_reg=row.get('sire_reg')
            )
            feat_rows.append(feats)

        valid_mask = [f is not None for f in feat_rows]
        valid_feats = [f for f in feat_rows if f is not None]
        if len(valid_feats) < 50: continue

        feat_df = pd.DataFrame(valid_feats)
        y_v = y[valid_mask]

        # Sire consistency (leave-one-out within May2026)
        sub_v = sub[valid_mask].reset_index(drop=True)
        sire_groups = defaultdict(list)
        for idx, row in sub_v.iterrows():
            val = y_v[idx]
            sire_groups[row['sirename']].append((idx, val))

        sire_loo = np.full(len(y_v), np.nan)
        sire_std_v = np.full(len(y_v), np.nan)
        sire_n_v = np.zeros(len(y_v))
        for sire, daughters in sire_groups.items():
            if len(daughters) < 2: continue
            for idx, val in daughters:
                others = [v for i, v in daughters if i != idx]
                sire_loo[idx] = np.mean(others)
                sire_std_v[idx] = np.std([v for _, v in daughters])
                sire_n_v[idx] = len(others)

        feat_df['sire_loo'] = sire_loo
        feat_df['sire_std_local'] = sire_std_v
        feat_df['sire_n_local'] = sire_n_v

        # Drop sparse, fill NaN
        feat_df = feat_df.dropna(axis=1, thresh=int(len(feat_df) * 0.3))
        for c in feat_df.columns:
            if feat_df[c].isna().any():
                feat_df[c] = feat_df[c].fillna(feat_df[c].median())

        X = feat_df.values
        n_feats = X.shape[1]

        # V5 (same features WITHOUT CDCB)
        v5_cols = [c for c in feat_df.columns if not any(
            k in c for k in ['cdcb_', 'pa_cdcb', 'sire_cdcb', 'dam_cdcb',
                            'transfer', 'regression_to_mean', 'pa_in_upper',
                            'pa_in_lower', 'pa_extreme'])]
        X_v5 = feat_df[v5_cols].values

        # Recalculate PA for valid
        pa_v = np.array([(sub_v.at[i, '_sp'][tn] + sub_v.at[i, '_dp'][tn]) / 2
                         for i in range(len(sub_v))])

        lgbm_kw = dict(n_estimators=500, max_depth=6, learning_rate=0.03,
                       subsample=0.8, colsample_bytree=0.7, min_child_samples=3,
                       reg_alpha=0.1, reg_lambda=1.0, random_state=42, verbose=-1, n_jobs=-1)

        # V5 CV
        v5_r2s, v5_maes = [], []
        for tr, te in kf.split(X_v5):
            m = LGBMRegressor(**lgbm_kw); m.fit(X_v5[tr], y_v[tr])
            p = m.predict(X_v5[te])
            v5_r2s.append(r2_score(y_v[te], p)); v5_maes.append(mean_absolute_error(y_v[te], p))

        # V7 CV (stacking)
        estimators = [
            ('ridge', Pipeline([('scaler', StandardScaler()), ('ridge', Ridge(alpha=1.0))])),
            ('lgbm', LGBMRegressor(**lgbm_kw)),
            ('rf', RandomForestRegressor(n_estimators=300, max_depth=10,
                                          min_samples_leaf=3, random_state=42, n_jobs=-1)),
        ]
        stack = StackingRegressor(estimators=estimators,
                                  final_estimator=Ridge(alpha=0.5),
                                  cv=3, n_jobs=-1, passthrough=True)

        v7_r2s, v7_maes = [], []
        for tr, te in kf.split(X):
            stack.fit(X[tr], y_v[tr])
            p = stack.predict(X[te])
            v7_r2s.append(r2_score(y_v[te], p)); v7_maes.append(mean_absolute_error(y_v[te], p))

        v5_r2, v5_mae = np.mean(v5_r2s), np.mean(v5_maes)
        v7_r2, v7_mae = np.mean(v7_r2s), np.mean(v7_maes)

        v5_vs_pa = (pa_mae - v5_mae) / pa_mae * 100
        v7_vs_pa = (pa_mae - v7_mae) / pa_mae * 100
        v7_vs_v5 = (v5_mae - v7_mae) / v5_mae * 100

        print(f"  {tn:>5} | {len(y_v):>5} | {pa_r2:>7.4f} | {v5_r2:>7.4f} | {v7_r2:>7.4f} | "
              f"{v5_mae:>8.3f} {v7_mae:>8.3f} | {v5_vs_pa:>+5.1f}% {v7_vs_pa:>+5.1f}% {v7_vs_v5:>+5.1f}%")

        all_results.append({
            'Trait': tn, 'N': len(y_v), 'Features': n_feats,
            'PA_R2': round(r2_score(y_v, pa_v), 4),
            'V5_R2': round(v5_r2, 4), 'V5_MAE': round(v5_mae, 4),
            'V7_R2': round(v7_r2, 4), 'V7_MAE': round(v7_mae, 4),
            'V5_vs_PA_%': round(v5_vs_pa, 2), 'V7_vs_PA_%': round(v7_vs_pa, 2),
            'V7_vs_V5_%': round(v7_vs_v5, 2),
        })

    # Summary
    rd = pd.DataFrame(all_results)
    rd.to_csv(OUTPUT_DIR / "v7_final_results.csv", index=False)

    print(f"\n{'='*90}")
    print("  RESUMO FINAL V7")
    print("=" * 90)
    ap = np.mean([r['PA_R2'] for r in all_results])
    a5 = np.mean([r['V5_R2'] for r in all_results])
    a7 = np.mean([r['V7_R2'] for r in all_results])
    gm5 = np.mean([r['V5_vs_PA_%'] for r in all_results])
    gm7 = np.mean([r['V7_vs_PA_%'] for r in all_results])
    g75 = np.mean([r['V7_vs_V5_%'] for r in all_results])

    print(f"  PA  R2 medio: {ap:.4f}")
    print(f"  V5  R2 medio: {a5:.4f}  ({gm5:+.1f}% MAE vs PA)")
    print(f"  V7  R2 medio: {a7:.4f}  ({gm7:+.1f}% MAE vs PA)")
    print(f"  V7 vs V5 MAE: {g75:+.1f}%")

    v7w = sum(1 for r in all_results if r['V7_R2'] > r['V5_R2'])
    v5w = sum(1 for r in all_results if r['V5_R2'] > r['V7_R2'])
    print(f"  V7 vence V5: {v7w}/{len(all_results)} traits")
    if v5w: print(f"  V5 vence V7: {v5w} traits: {[r['Trait'] for r in all_results if r['V5_R2'] > r['V7_R2']]}")

    v7pa = sum(1 for r in all_results if r['V7_R2'] > r['PA_R2'])
    print(f"  V7 vence PA: {v7pa}/{len(all_results)} traits")

    print(f"\n  Resultados: {OUTPUT_DIR / 'v7_final_results.csv'}")
    return rd


if __name__ == '__main__':
    run_v7()
