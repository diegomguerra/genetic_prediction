"""
DSII V8 Oracle - Best model per trait based on Model Shootout results.
Uses V5 feature engineering (proven best) + selects optimal architecture per trait.
Trains on full dataset and saves models for production use.
"""
import warnings
warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd
import openpyxl, csv, pickle, sys
from pathlib import Path
from collections import defaultdict

from sklearn.model_selection import KFold
from sklearn.metrics import mean_absolute_error, r2_score
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from sklearn.linear_model import Ridge
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from lightgbm import LGBMRegressor
from xgboost import XGBRegressor

DOWNLOADS = Path("C:/Users/DiegoGuerra/Downloads")
BULLS_CSV = Path("C:/Users/DiegoGuerra/gen-genie-advisor/sms-engine/data/bulls.csv")
OUTPUT_DIR = Path("C:/Users/DiegoGuerra/Documents/Projetos/genetic_prediction/dsii_v8_results")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

def sf(v):
    if v is None: return None
    try:
        x = float(v)
        return x if not np.isnan(x) else None
    except: return None

def flush_print(*args, **kwargs):
    print(*args, **kwargs)
    sys.stdout.flush()

# ============================================================
# DOMAIN KNOWLEDGE (same as V5)
# ============================================================
GENETIC_SD = {
    'TPI': 250, 'NM$': 275, 'CM$': 280,
    'MILK': 675, 'FAT': 29, 'FAT%': 0.05, 'PRO': 19, 'PRO%': 0.02, 'CFP': 25,
    'PL': 1.85, 'SCS': 0.14, 'DPR': 1.3, 'HCR': 1.4, 'CCR': 1.65,
    'LIV': 1.2, 'FI': 1.0, 'MAST': 1.0, 'FSAV': 50,
    'PTAT': 0.70, 'UDC': 0.75, 'FLC': 0.65,
    'SCE': 1.5, 'DCE': 1.2, 'SSB': 1.0, 'DSB': 1.0,
}
HERITABILITY = {
    'TPI': 0.30, 'NM$': 0.30, 'CM$': 0.30,
    'MILK': 0.25, 'FAT': 0.25, 'FAT%': 0.50, 'PRO': 0.25, 'PRO%': 0.50,
    'CFP': 0.30, 'PL': 0.08, 'SCS': 0.12, 'DPR': 0.04, 'LIV': 0.05,
    'FI': 0.06, 'HCR': 0.04, 'CCR': 0.04, 'MAST': 0.04, 'FSAV': 0.15,
    'PTAT': 0.30, 'UDC': 0.25, 'FLC': 0.15,
    'SCE': 0.08, 'DCE': 0.06, 'SSB': 0.06, 'DSB': 0.04,
}
BREED_IDEAL = {
    'MILK': 1800, 'FAT': 90, 'PRO': 60, 'CFP': 70,
    'DPR': 1.5, 'CCR': 2.0, 'HCR': 2.0,
    'PL': 5.0, 'LIV': 2.0, 'UDC': 1.5, 'FLC': 1.0,
    'MAST': -1.0, 'SCS': 2.60, 'SCE': 2.0,
}
LOWER_IS_BETTER = {'SCS', 'SCE', 'SSB', 'DSB', 'DCE', 'MAST'}
INBREEDING_DEPRESSION = {
    'NM$': -25.0, 'TPI': -20.0, 'CM$': -25.0,
    'MILK': -28.5, 'FAT': -1.0, 'PRO': -0.9,
    'PL': -0.35, 'DPR': -0.03, 'SCS': 0.007, 'LIV': -0.15,
}
GENETIC_CORRELATIONS = {
    ('MILK', 'DPR'): -0.35, ('MILK', 'CCR'): -0.30, ('MILK', 'SCS'): 0.10,
    ('MILK', 'PL'): -0.15, ('FAT', 'PRO'): 0.65, ('FAT', 'DPR'): -0.25,
    ('PRO', 'DPR'): -0.30, ('DPR', 'CCR'): 0.55, ('DPR', 'PL'): 0.45,
    ('SCS', 'PL'): -0.30, ('SCS', 'MAST'): 0.70,
    ('PTAT', 'UDC'): 0.40, ('PTAT', 'FLC'): 0.30,
    ('NM$', 'TPI'): 0.85, ('NM$', 'CM$'): 0.95,
    ('MILK', 'FAT'): 0.55, ('MILK', 'PRO'): 0.85,
    ('PL', 'LIV'): 0.65, ('DPR', 'FI'): 0.70,
    ('SCE', 'SSB'): 0.60, ('DCE', 'DSB'): 0.55,
}
DOMINANCE_RATIOS = {
    'MILK': 0.12, 'FAT': 0.10, 'PRO': 0.12,
    'SCS': 0.14, 'PL': 0.15, 'DPR': 0.20, 'CCR': 0.15,
    'UDC': 0.11, 'FLC': 0.08, 'LIV': 0.12,
}

TRAIT_MAP = [
    ('TPI',  'GTPI',      'TPI',         'TPI'),
    ('NM$',  'NM$',       'Net Merit',   'NM$'),
    ('CM$',  'CM$',       'CM$',         'CM$'),
    ('MILK', 'MILK',      'PTA Milk',    'PTAM'),
    ('FAT',  'FAT',       'PTA Fat',     'PTAF'),
    ('FAT%', '%F',        '% Fat',       'PTAF%'),
    ('PRO',  'PRO',       'PTA Pro',     'PTAP'),
    ('PRO%', '%P',        '% Pro',       'PTAP%'),
    ('CFP',  'CFP',       'CFP',         'CFP'),
    ('PL',   'PL',        'PL',          'PL'),
    ('SCS',  'SCS',       'SCS',         'SCS'),
    ('DPR',  'DPR',       'PTA DPR',     'DPR'),
    ('LIV',  'LIV',       'PTA LIV',     'LIV'),
    ('FI',   'FI',        'Fertil Index', 'FI'),
    ('HCR',  'HCR',       'HCR',         'HCR'),
    ('CCR',  'CCR',       'CCR',         'CCR'),
    ('MAST', 'MAST',      'Mastitis',    'MAST'),
    ('FSAV', 'FSAV',      'Feed Saved',  'F SAV'),
    ('PTAT', 'PTAT',      'PTA Type',    'PTAT'),
    ('UDC',  'UDC',       'UDC',         'UDC'),
    ('FLC',  'FLC',       'FLC',         'FLC'),
    ('SCE',  'SCE',       'SCE',         'SCE'),
    ('DCE',  'DCE',       'DCE',         'DCE'),
    ('SSB',  'SSB',       'SSB',         'SSB'),
    ('DSB',  'DSB',       'DSB',         'DSB'),
]
ALL_TRAITS = [t[0] for t in TRAIT_MAP]

# Oracle map: best model per trait from shootout
ORACLE_MAP = {
    'TPI':  'GBR',   'NM$':  'RF',    'CM$':  'RF',
    'MILK': 'GBR',   'FAT':  'GBR',   'FAT%': 'GBR',
    'PRO':  'LGBM',  'PRO%': 'GBR',   'CFP':  'GBR',
    'PL':   'LGBM',  'SCS':  'GBR',   'DPR':  'LGBM',
    'LIV':  'GBR',   'FI':   'LGBM',  'HCR':  'GBR',
    'CCR':  'GBR',   'MAST': 'LGBM',  'FSAV': 'LGBM',
    'PTAT': 'XGB',   'UDC':  'GBR',   'FLC':  'LGBM',
    'SCE':  'GBR',   'DCE':  'LGBM',  'SSB':  'LGBM',
    'DSB':  'LGBM',
}

def get_model(name):
    if name == 'GBR':
        return GradientBoostingRegressor(
            n_estimators=500, max_depth=5, learning_rate=0.03,
            min_samples_leaf=3, subsample=0.8, random_state=42)
    elif name == 'LGBM':
        return LGBMRegressor(
            n_estimators=500, max_depth=6, learning_rate=0.03,
            subsample=0.8, colsample_bytree=0.7, min_child_samples=3,
            reg_alpha=0.1, reg_lambda=1.0, random_state=42, verbose=-1)
    elif name == 'RF':
        return RandomForestRegressor(
            n_estimators=500, max_depth=12, min_samples_leaf=3,
            random_state=42, n_jobs=-1)
    elif name == 'XGB':
        return XGBRegressor(
            n_estimators=500, max_depth=6, learning_rate=0.03,
            subsample=0.8, colsample_bytree=0.7, reg_alpha=0.1,
            reg_lambda=1.0, min_child_weight=3, random_state=42, verbosity=0)
    elif name == 'Ridge':
        return Pipeline([('sc', StandardScaler()), ('m', Ridge(alpha=1.0))])
    else:
        raise ValueError(f"Unknown model: {name}")


def build_features(sire_ptas, dam_ptas, trait):
    s = sire_ptas.get(trait)
    d = dam_ptas.get(trait)
    if s is None or d is None:
        return None
    f = {}
    sd = GENETIC_SD.get(trait, 1)
    h2 = HERITABILITY.get(trait, 0.15)
    pa = (s + d) / 2

    f['sire'] = s; f['dam'] = d; f['pa'] = pa
    f['delta_g'] = (s - d) / 2; f['diff'] = s - d
    f['sire_z'] = s / sd if sd else 0; f['dam_z'] = d / sd if sd else 0
    f['sire_sq'] = s * s; f['dam_sq'] = d * d; f['sxd'] = s * d
    f['h2_pa'] = pa * (0.5 + h2)

    ideal = BREED_IDEAL.get(trait)
    if ideal is not None:
        defic = max(0, ((d - ideal) / sd) if trait in LOWER_IS_BETTER else ((ideal - d) / sd))
        f['deficiency'] = defic; f['dg_x_def'] = f['delta_g'] * defic
    else:
        f['deficiency'] = 0; f['dg_x_def'] = 0

    f['ib_effect'] = INBREEDING_DEPRESSION.get(trait, 0) * 0.085
    dom = DOMINANCE_RATIOS.get(trait, 0)
    f['dom_pot'] = dom * abs(s - d) / sd if sd else 0

    for ot in ALL_TRAITS:
        if ot == trait: continue
        sv = sire_ptas.get(ot); dv = dam_ptas.get(ot)
        if sv is not None: f[f's_{ot}'] = sv
        if dv is not None: f[f'd_{ot}'] = dv

    for (t1, t2), corr in GENETIC_CORRELATIONS.items():
        if t1 == trait:
            ov = sire_ptas.get(t2); dov = dam_ptas.get(t2)
            if ov is not None and dov is not None:
                f[f'gc_{t2}'] = ((ov + dov) / 2) * corr
        elif t2 == trait:
            ov = sire_ptas.get(t1); dov = dam_ptas.get(t1)
            if ov is not None and dov is not None:
                f[f'gc_{t1}'] = ((ov + dov) / 2) * corr

    milk_pa = ((sire_ptas.get('MILK', 0) or 0) + (dam_ptas.get('MILK', 0) or 0)) / 2
    fat_pa = ((sire_ptas.get('FAT', 0) or 0) + (dam_ptas.get('FAT', 0) or 0)) / 2
    pro_pa = ((sire_ptas.get('PRO', 0) or 0) + (dam_ptas.get('PRO', 0) or 0)) / 2
    f['prod_press'] = milk_pa / 675 + fat_pa / 29 + pro_pa / 19
    f['sire_nm'] = sire_ptas.get('NM$', 0) or 0
    f['dam_nm'] = dam_ptas.get('NM$', 0) or 0

    return f


# ============================================================
# LOAD DATA
# ============================================================
flush_print("=" * 100)
flush_print("  DSII V8 ORACLE - Best Model Per Trait")
flush_print("=" * 100)

flush_print("\n  Loading data...")
wb = openpyxl.load_workbook(DOWNLOADS / "May 2026.xlsx", read_only=True, data_only=True)
ws = wb.active
may_rows = list(ws.iter_rows(values_only=True))
wb.close()
hdr = [str(c).strip() if c else f'col_{i}' for i, c in enumerate(may_rows[0])]

dams = {}
for i in range(1, 21):
    fp = DOWNLOADS / f"DAM{i}.xlsx"
    if not fp.exists(): continue
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    ws = wb.active
    dr = list(ws.iter_rows(values_only=True))
    wb.close()
    dh = [str(c).strip() if c else '' for j, c in enumerate(dr[0])]
    ni = next(j for j, h in enumerate(dh) if 'naab' in h.lower() and 'code' in h.lower())
    for r in dr[1:]:
        if r[ni]: dams[str(r[ni]).strip()] = {dh[j]: r[j] for j in range(min(len(dh), len(r)))}

bulls_rn, bulls_nm = {}, {}
with open(BULLS_CSV, 'r', encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        rn = row.get('Registration Name', '').strip().upper()
        nm = row.get('Name', '').strip().upper()
        if rn: bulls_rn[rn] = row
        if nm: bulls_nm[nm] = row

records = []
for r in may_rows[1:]:
    rd = {hdr[i]: r[i] for i in range(min(len(hdr), len(r)))}
    sn = str(rd.get('SIRENAME', '')).strip().upper()
    sire = bulls_rn.get(sn) or bulls_nm.get(sn)
    if not sire: continue
    damreg = str(rd.get('DAMREGNUM2', '')).strip()
    dam = dams.get(damreg)
    if not dam: continue

    sire_ptas, dam_ptas = {}, {}
    rec = {'sirename': sn}
    for tn, dcol, mcol, bcol in TRAIT_MAP:
        dv = sf(rd.get(dcol)); sv = sf(sire.get(bcol))
        mv = sf(dam.get(mcol)) if mcol else None
        rec[f'd_{tn}'] = dv
        if sv is not None: sire_ptas[tn] = sv
        if mv is not None: dam_ptas[tn] = mv
    rec['_sp'] = sire_ptas; rec['_dp'] = dam_ptas
    records.append(rec)

df = pd.DataFrame(records)
flush_print(f"  Trios: {len(df)}")

# ============================================================
# V8 ORACLE: CV evaluation + full training + save models
# ============================================================
flush_print(f"\n{'='*100}")
flush_print("  PHASE 1: Cross-Validation Evaluation")
flush_print("=" * 100)

kf = KFold(n_splits=5, shuffle=True, random_state=42)

flush_print(f"\n {'Trait':>6} | {'N':>5} | {'Model':>6} | {'PA R2':>7} | {'V5 R2':>7} | {'V8 R2':>7} | "
            f"{'PA MAE':>8} {'V8 MAE':>8} | {'V8vPA':>7} {'V8vV5':>7}")
flush_print("-" * 105)

# Load V5 results for comparison
v5_path = Path("C:/Users/DiegoGuerra/Documents/Projetos/genetic_prediction/dsii_v5_results/v5_hybrid_results.csv")
v5_data = {}
if v5_path.exists():
    v5_df = pd.read_csv(v5_path)
    for _, row in v5_df.iterrows():
        v5_data[row['Trait']] = row.get('V5_R2', row.get('Best_R2', None))

all_results = []
saved_models = {}

for tn, _, mcol, _ in TRAIT_MAP:
    if mcol is None: continue
    target = f'd_{tn}'
    mask = df[target].notna() & df.apply(lambda r: tn in r['_sp'] and tn in r['_dp'], axis=1)
    sub = df[mask].reset_index(drop=True)
    if len(sub) < 50: continue

    y = sub[target].values

    # Build features
    feat_rows = [build_features(row['_sp'], row['_dp'], tn) for _, row in sub.iterrows()]
    valid = [f is not None for f in feat_rows]
    feat_df = pd.DataFrame([f for f in feat_rows if f is not None])
    y_v = y[valid]

    # Sire consistency (leave-one-out)
    sub_v = sub[valid].reset_index(drop=True)
    sire_groups = defaultdict(list)
    for idx in range(len(sub_v)):
        sire_groups[sub_v.at[idx, 'sirename']].append((idx, y_v[idx]))

    sire_loo = np.full(len(y_v), np.nan)
    for sire, daughters in sire_groups.items():
        if len(daughters) < 2: continue
        for idx, val in daughters:
            others = [v for i, v in daughters if i != idx]
            sire_loo[idx] = np.mean(others)
    feat_df['sire_loo'] = sire_loo

    feat_df = feat_df.dropna(axis=1, thresh=int(len(feat_df) * 0.3))
    feature_cols = list(feat_df.columns)
    for c in feature_cols:
        if feat_df[c].isna().any():
            feat_df[c] = feat_df[c].fillna(feat_df[c].median())

    X = feat_df.values

    # PA baseline
    pa_v = np.array([(sub_v.at[i, '_sp'][tn] + sub_v.at[i, '_dp'][tn]) / 2 for i in range(len(sub_v))])
    pa_r2 = r2_score(y_v, pa_v)
    pa_mae = mean_absolute_error(y_v, pa_v)

    # V8 Oracle: use best model for this trait
    model_name = ORACLE_MAP[tn]
    import copy

    # 5-fold CV
    r2s, maes, preds_all = [], [], np.zeros(len(y_v))
    for tr, te in kf.split(X):
        m = get_model(model_name)
        m.fit(X[tr], y_v[tr])
        pred = m.predict(X[te])
        r2s.append(r2_score(y_v[te], pred))
        maes.append(mean_absolute_error(y_v[te], pred))
        preds_all[te] = pred

    v8_r2 = np.mean(r2s)
    v8_mae = np.mean(maes)
    v5_r2 = v5_data.get(tn, 0)

    # Improvement metrics
    v8_vs_pa_mae = ((pa_mae - v8_mae) / pa_mae * 100) if pa_mae > 0 else 0
    v8_vs_v5_r2 = ((v8_r2 - v5_r2) / abs(v5_r2) * 100) if v5_r2 else 0

    flush_print(f"  {tn:>5} | {len(y_v):>5} | {model_name:>6} | {pa_r2:.4f} | {v5_r2:.4f} | {v8_r2:.4f} | "
                f"{pa_mae:>8.3f} {v8_mae:>8.3f} | {v8_vs_pa_mae:>+6.1f}% {v8_vs_v5_r2:>+6.1f}%")

    all_results.append({
        'Trait': tn, 'N': len(y_v), 'Model': model_name,
        'PA_R2': round(pa_r2, 4), 'V5_R2': round(v5_r2, 4), 'V8_R2': round(v8_r2, 4),
        'PA_MAE': round(pa_mae, 4), 'V8_MAE': round(v8_mae, 4),
        'V8vPA_MAE_pct': round(v8_vs_pa_mae, 1),
    })

    # Train final model on ALL data for production
    final_model = get_model(model_name)
    final_model.fit(X, y_v)
    saved_models[tn] = {
        'model': final_model,
        'model_name': model_name,
        'feature_cols': feature_cols,
        'r2_cv': round(v8_r2, 4),
    }

# ============================================================
# SUMMARY
# ============================================================
flush_print(f"\n{'='*100}")
flush_print("  RESUMO FINAL V8 ORACLE")
flush_print("=" * 100)

pa_r2s = [r['PA_R2'] for r in all_results]
v5_r2s = [r['V5_R2'] for r in all_results]
v8_r2s = [r['V8_R2'] for r in all_results]

flush_print(f"  PA  R2 medio: {np.mean(pa_r2s):.4f}")
flush_print(f"  V5  R2 medio: {np.mean(v5_r2s):.4f}")
flush_print(f"  V8  R2 medio: {np.mean(v8_r2s):.4f}")

v8_beats_v5 = sum(1 for r in all_results if r['V8_R2'] > r['V5_R2'])
v8_beats_pa = sum(1 for r in all_results if r['V8_R2'] > r['PA_R2'])
flush_print(f"\n  V8 vence V5: {v8_beats_v5}/{len(all_results)} traits")
flush_print(f"  V8 vence PA: {v8_beats_pa}/{len(all_results)} traits")

# Traits where V8 > V5
better = [r['Trait'] for r in all_results if r['V8_R2'] > r['V5_R2']]
if better:
    flush_print(f"  V8 > V5 em: {better}")
worse = [r['Trait'] for r in all_results if r['V8_R2'] < r['V5_R2']]
if worse:
    flush_print(f"  V5 > V8 em: {worse}")

# Model usage summary
from collections import Counter
model_counts = Counter(r['Model'] for r in all_results)
flush_print(f"\n  Modelos usados:")
for mn, cnt in model_counts.most_common():
    traits_using = [r['Trait'] for r in all_results if r['Model'] == mn]
    flush_print(f"    {mn:>6}: {cnt} traits - {traits_using}")

# R2 tiers
flush_print(f"\n  Traits por faixa de R2:")
tier_90 = [r for r in all_results if r['V8_R2'] >= 0.90]
tier_80 = [r for r in all_results if 0.80 <= r['V8_R2'] < 0.90]
tier_70 = [r for r in all_results if 0.70 <= r['V8_R2'] < 0.80]
tier_low = [r for r in all_results if r['V8_R2'] < 0.70]

if tier_90:
    flush_print(f"    R2 >= 0.90: {len(tier_90)} traits - {[r['Trait'] for r in tier_90]}")
if tier_80:
    flush_print(f"    R2 0.80-0.89: {len(tier_80)} traits - {[r['Trait'] for r in tier_80]}")
if tier_70:
    flush_print(f"    R2 0.70-0.79: {len(tier_70)} traits - {[r['Trait'] for r in tier_70]}")
if tier_low:
    flush_print(f"    R2 < 0.70: {len(tier_low)} traits - {[r['Trait'] for r in tier_low]}")

# Save results
results_df = pd.DataFrame(all_results)
results_df.to_csv(OUTPUT_DIR / "v8_oracle_results.csv", index=False)
flush_print(f"\n  Resultados: {OUTPUT_DIR / 'v8_oracle_results.csv'}")

# Save trained models
with open(OUTPUT_DIR / "v8_oracle_models.pkl", 'wb') as f:
    pickle.dump(saved_models, f)
flush_print(f"  Modelos salvos: {OUTPUT_DIR / 'v8_oracle_models.pkl'}")
flush_print(f"\n  V8 Oracle pronto para producao!")
