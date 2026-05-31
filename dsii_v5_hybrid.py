"""
DSII v5 — Hybrid Model (Domain Knowledge + ML)
================================================
Combina features de engenharia genética do DSII v3 com ML para superar PA
em TODOS os traits.

Estratégia:
  1. PA como feature base (nunca descartado)
  2. Features V3: deficiency profile, DeltaG, antagonismo, correlações genéticas
  3. Features cruzadas: interações sire×dam, desvios, ratios
  4. Sire breeding consistency (performance média das filhas do mesmo touro)
  5. Stacking ensemble: Ridge + RF + XGBoost → meta-learner
"""

import warnings
warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd
import openpyxl
import csv
import math
from pathlib import Path
from collections import defaultdict

from sklearn.model_selection import KFold
from sklearn.linear_model import Ridge, ElasticNet
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor, StackingRegressor
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline

try:
    from xgboost import XGBRegressor
    HAS_XGB = True
except ImportError:
    HAS_XGB = False

try:
    from lightgbm import LGBMRegressor
    HAS_LGBM = True
except ImportError:
    HAS_LGBM = False

DOWNLOADS = Path("C:/Users/DiegoGuerra/Downloads")
BULLS_CSV = Path("C:/Users/DiegoGuerra/gen-genie-advisor/sms-engine/data/bulls.csv")
OUTPUT_DIR = Path("C:/Users/DiegoGuerra/Documents/Projetos/genetic_prediction/dsii_v5_results")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ============================================================
# DSII v3 DOMAIN KNOWLEDGE
# ============================================================

ECON_WEIGHTS = {
    "FAT": 2.60, "PRO": 1.40, "MILK": -0.02, "PL": 35.00,
    "SCS": -120.0, "DPR": 28.00, "CCR": 11.00, "HCR": 6.00,
    "LIV": 25.00, "UDC": 30.00, "FLC": 5.00,
    "SCE": -3.50, "MAST": -25.0, "FSAV": -1.00,
}

GENETIC_SD = {
    "MILK": 675, "FAT": 29, "PRO": 19, "CFP": 25,
    "SCS": 0.14, "PL": 1.85, "LIV": 1.2,
    "DPR": 1.3, "CCR": 1.65, "HCR": 1.4,
    "UDC": 0.75, "FLC": 0.65,
    "NM$": 275, "TPI": 250, "CM$": 280,
    "SCE": 1.5, "SSB": 1.0, "DSB": 1.0, "DCE": 1.2,
    "MAST": 1.0, "FSAV": 50, "FI": 1.0,
    "PTAT": 0.70, "FAT%": 0.05, "PRO%": 0.02,
}

BREED_IDEAL = {
    "MILK": 1800, "FAT": 90, "PRO": 60, "CFP": 70,
    "DPR": 1.5, "CCR": 2.0, "HCR": 2.0,
    "PL": 5.0, "LIV": 2.0,
    "UDC": 1.5, "FLC": 1.0,
    "MAST": -1.0, "SCS": 2.60,
    "SCE": 2.0, "SSB": 5.0, "DSB": 5.0,
}

LOWER_IS_BETTER = {"SCS", "SCE", "SSB", "DSB", "DCE", "MAST"}

INBREEDING_DEPRESSION = {
    "NM$": -25.0, "TPI": -20.0, "CM$": -25.0,
    "MILK": -28.5, "FAT": -1.0, "PRO": -0.9,
    "PL": -0.35, "DPR": -0.03, "SCS": 0.007, "LIV": -0.15,
    "CCR": -0.025, "HCR": -0.02, "CFP": -0.8,
}

CRITICAL_THRESHOLDS = {
    "DPR": -2.0, "PL": -1.0, "SCS": 3.30,
    "UDC": -1.0, "FLC": -1.0, "SCE": 5.0,
}

# Genetic correlations — used to build cross-trait interaction features
GENETIC_CORRELATIONS = {
    ("MILK", "DPR"): -0.35, ("MILK", "CCR"): -0.30,
    ("MILK", "SCS"): 0.10, ("MILK", "PL"): -0.15,
    ("MILK", "UDC"): -0.15, ("FAT", "PRO"): 0.65,
    ("FAT", "DPR"): -0.25, ("FAT", "CCR"): -0.20,
    ("PRO", "DPR"): -0.30, ("PRO", "CCR"): -0.25,
    ("DPR", "CCR"): 0.55, ("DPR", "PL"): 0.45,
    ("CCR", "PL"): 0.30, ("UDC", "PL"): 0.25,
    ("FLC", "PL"): 0.10, ("SCS", "PL"): -0.30,
    ("SCS", "UDC"): -0.25, ("SCS", "MAST"): 0.70,
    ("PTAT", "MILK"): 0.15, ("PTAT", "FLC"): 0.30,
    ("PTAT", "UDC"): 0.40, ("FAT", "FAT%"): 0.60,
    ("PRO", "PRO%"): 0.50, ("NM$", "TPI"): 0.85,
    ("NM$", "CM$"): 0.95, ("TPI", "CM$"): 0.90,
    ("MILK", "FAT"): 0.55, ("MILK", "PRO"): 0.85,
    ("PL", "LIV"): 0.65, ("DPR", "FI"): 0.70,
    ("SCE", "SSB"): 0.60, ("DCE", "DSB"): 0.55,
    ("SCE", "DCE"): 0.40, ("FSAV", "MILK"): -0.30,
}

DOMINANCE_RATIOS = {
    "MILK": 0.12, "FAT": 0.10, "PRO": 0.12,
    "SCS": 0.14, "PL": 0.15, "DPR": 0.20, "CCR": 0.15,
    "UDC": 0.11, "FLC": 0.08, "LIV": 0.12,
}

# Heritability estimates (Holstein)
HERITABILITY = {
    "TPI": 0.30, "NM$": 0.30, "CM$": 0.30,
    "MILK": 0.25, "FAT": 0.25, "FAT%": 0.50, "PRO": 0.25, "PRO%": 0.50,
    "CFP": 0.30, "PL": 0.08, "SCS": 0.12, "DPR": 0.04, "LIV": 0.05,
    "FI": 0.06, "HCR": 0.04, "CCR": 0.04, "MAST": 0.04, "FSAV": 0.15,
    "PTAT": 0.30, "UDC": 0.25, "FLC": 0.15, "SCE": 0.08, "DCE": 0.06,
    "SSB": 0.06, "DSB": 0.04,
}

# ============================================================
# TRAIT MAPPING
# ============================================================
TRAIT_MAP = [
    ('TPI',    'GTPI',       'TPI',          'TPI'),
    ('NM$',    'NM$',        'Net Merit',    'NM$'),
    ('CM$',    'CM$',        'CM$',          'CM$'),
    ('MILK',   'MILK',       'PTA Milk',     'PTAM'),
    ('FAT',    'FAT',        'PTA Fat',      'PTAF'),
    ('FAT%',   '%F',         '% Fat',        'PTAF%'),
    ('PRO',    'PRO',        'PTA Pro',      'PTAP'),
    ('PRO%',   '%P',         '% Pro',        'PTAP%'),
    ('CFP',    'CFP',        'CFP',          'CFP'),
    ('PL',     'PL',         'PL',           'PL'),
    ('SCS',    'SCS',        'SCS',          'SCS'),
    ('DPR',    'DPR',        'PTA DPR',      'DPR'),
    ('LIV',    'LIV',        'PTA LIV',      'LIV'),
    ('FI',     'FI',         'Fertil Index',  'FI'),
    ('HCR',    'HCR',        'HCR',          'HCR'),
    ('CCR',    'CCR',        'CCR',          'CCR'),
    ('MAST',   'MAST',       'Mastitis',     'MAST'),
    ('FSAV',   'FSAV',       'Feed Saved',   'F SAV'),
    ('PTAT',   'PTAT',       'PTA Type',     'PTAT'),
    ('UDC',    'UDC',        'UDC',          'UDC'),
    ('FLC',    'FLC',        'FLC',          'FLC'),
    ('SCE',    'SCE',        'SCE',          'SCE'),
    ('DCE',    'DCE',        'DCE',          'DCE'),
    ('SSB',    'SSB',        'SSB',          'SSB'),
    ('DSB',    'DSB',        'DSB',          'DSB'),
]

ALL_TRAIT_NAMES = [t[0] for t in TRAIT_MAP]


def safe_float(val):
    if val is None:
        return None
    try:
        v = float(val)
        return v if not np.isnan(v) else None
    except (ValueError, TypeError):
        return None


# ============================================================
# DATA LOADING
# ============================================================

def load_all_data():
    print("Carregando dados...")

    # Daughters
    wb = openpyxl.load_workbook(DOWNLOADS / "May 2026.xlsx", read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr_d = [str(c).strip() if c else f'col_{i}' for i, c in enumerate(rows[0])]
    daughters = [{hdr_d[i]: r[i] for i in range(min(len(hdr_d), len(r)))} for r in rows[1:]]
    print(f"  Filhas: {len(daughters)}")

    # Dams
    dams = {}
    for i in range(1, 21):
        f = DOWNLOADS / f"DAM{i}.xlsx"
        if not f.exists():
            continue
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        hdr = [str(c).strip() if c else f'col_{j}' for j, c in enumerate(rows[0])]
        naab_idx = next(j for j, h in enumerate(hdr) if 'naab' in h.lower() and 'code' in h.lower())
        for r in rows[1:]:
            naab = r[naab_idx]
            if not naab:
                continue
            dams[str(naab).strip()] = {hdr[j]: r[j] for j in range(min(len(hdr), len(r)))}
    print(f"  Maes: {len(dams)}")

    # Bulls
    bulls_regname = {}
    bulls_name = {}
    with open(BULLS_CSV, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            rn = row.get('Registration Name', '').strip().upper()
            nm = row.get('Name', '').strip().upper()
            if rn:
                bulls_regname[rn] = row
            if nm:
                bulls_name[nm] = row
    print(f"  Pais: {len(bulls_regname)}")

    return daughters, dams, bulls_regname, bulls_name


# ============================================================
# FEATURE ENGINEERING — the core innovation
# ============================================================

def build_v3_features(sire_ptas, dam_ptas, trait_name):
    """
    Build domain-knowledge features from DSII v3 for a specific trait.
    These become ML inputs alongside raw PTAs.
    """
    features = {}
    s = sire_ptas.get(trait_name)
    d = dam_ptas.get(trait_name)

    if s is None or d is None:
        return None

    sd = GENETIC_SD.get(trait_name, 1)
    h2 = HERITABILITY.get(trait_name, 0.15)

    # 1. Parent Average (base)
    pa = (s + d) / 2.0
    features['pa'] = pa

    # 2. Sire and Dam raw
    features['sire'] = s
    features['dam'] = d

    # 3. DeltaG: improvement over dam
    delta_g = (s - d) / 2.0
    features['delta_g'] = delta_g

    # 4. Sire-Dam difference (absolute)
    features['sire_dam_diff'] = s - d

    # 5. Sire-Dam ratio (standardized)
    features['sire_dam_ratio'] = s / d if d != 0 else 0

    # 6. Standardized values (z-scores)
    features['sire_z'] = s / sd if sd > 0 else 0
    features['dam_z'] = d / sd if sd > 0 else 0
    features['pa_z'] = pa / sd if sd > 0 else 0

    # 7. Dam deficiency for this trait
    ideal = BREED_IDEAL.get(trait_name)
    if ideal is not None:
        if trait_name in LOWER_IS_BETTER:
            deficiency = max(0, (d - ideal) / sd)
        else:
            deficiency = max(0, (ideal - d) / sd)
        features['dam_deficiency'] = deficiency
        # DeltaG × deficiency interaction
        features['delta_g_x_deficiency'] = delta_g * deficiency
    else:
        features['dam_deficiency'] = 0
        features['delta_g_x_deficiency'] = 0

    # 8. Critical threshold proximity
    threshold = CRITICAL_THRESHOLDS.get(trait_name)
    if threshold is not None:
        if trait_name in LOWER_IS_BETTER:
            features['critical_proximity'] = (pa - threshold) / sd
        else:
            features['critical_proximity'] = (pa - threshold) / sd
    else:
        features['critical_proximity'] = 0

    # 9. Inbreeding depression factor
    ib_dep = INBREEDING_DEPRESSION.get(trait_name, 0)
    features['inbreeding_effect'] = ib_dep * 0.085  # mean F for Holstein

    # 10. Heritability-weighted PA
    # Higher h2 = more reliable PA, lower h2 = more regression to mean
    features['h2_weighted_pa'] = pa * (0.5 + h2)

    # 11. Dominance potential
    dom_ratio = DOMINANCE_RATIOS.get(trait_name, 0)
    features['dominance_potential'] = dom_ratio * abs(s - d) / sd if sd > 0 else 0

    # 12. Sire squared and Dam squared (capture non-linearity)
    features['sire_sq'] = s * s
    features['dam_sq'] = d * d

    # 13. Sire × Dam interaction
    features['sire_x_dam'] = s * d

    return features


def build_cross_trait_features(sire_ptas, dam_ptas, target_trait):
    """
    Build features from correlated traits to improve prediction.
    Uses genetic correlation knowledge to select relevant cross-traits.
    """
    features = {}

    # Get traits correlated with target
    correlated = {}
    for (t1, t2), corr in GENETIC_CORRELATIONS.items():
        if t1 == target_trait:
            correlated[t2] = corr
        elif t2 == target_trait:
            correlated[t1] = corr

    # PA of correlated traits (weighted by correlation strength)
    for other_trait, corr in correlated.items():
        s_other = sire_ptas.get(other_trait)
        d_other = dam_ptas.get(other_trait)
        if s_other is not None and d_other is not None:
            pa_other = (s_other + d_other) / 2.0
            sd_other = GENETIC_SD.get(other_trait, 1)
            features[f'corr_{other_trait}_pa_z'] = pa_other / sd_other
            features[f'corr_{other_trait}_weight'] = pa_other / sd_other * corr

    # Economic value context: sum of economic components from sire
    econ_sum = 0
    for t, w in ECON_WEIGHTS.items():
        sv = sire_ptas.get(t)
        if sv is not None:
            econ_sum += sv * w
    features['sire_econ_value'] = econ_sum

    # Dam overall genetic merit (proxy)
    dam_merit = 0
    for t, w in ECON_WEIGHTS.items():
        dv = dam_ptas.get(t)
        if dv is not None:
            dam_merit += dv * w
    features['dam_econ_value'] = dam_merit

    # Antagonism indicators for production traits
    if target_trait in ('DPR', 'CCR', 'PL', 'LIV', 'SCS', 'HCR', 'FI'):
        milk_pa = ((sire_ptas.get('MILK', 0) or 0) + (dam_ptas.get('MILK', 0) or 0)) / 2
        fat_pa = ((sire_ptas.get('FAT', 0) or 0) + (dam_ptas.get('FAT', 0) or 0)) / 2
        pro_pa = ((sire_ptas.get('PRO', 0) or 0) + (dam_ptas.get('PRO', 0) or 0)) / 2
        features['production_pressure'] = milk_pa / 675 + fat_pa / 29 + pro_pa / 19

    # Type composite context for type traits
    if target_trait in ('PTAT', 'UDC', 'FLC'):
        for t in ('PTAT', 'UDC', 'FLC'):
            if t != target_trait:
                sv = sire_ptas.get(t, 0) or 0
                dv = dam_ptas.get(t, 0) or 0
                features[f'type_{t}_pa'] = (sv + dv) / 2

    # Calving ease context
    if target_trait in ('SCE', 'DCE', 'SSB', 'DSB'):
        for t in ('SCE', 'DCE', 'SSB', 'DSB'):
            if t != target_trait:
                sv = sire_ptas.get(t, 0) or 0
                dv = dam_ptas.get(t, 0) or 0
                features[f'calving_{t}_pa'] = (sv + dv) / 2

    return features


def build_sire_consistency_features(trios_df, trait_name):
    """
    For each sire, compute statistics from his other daughters.
    Leave-one-out to avoid leakage.
    """
    target = f'd_{trait_name}'
    sire_col = 'sirename'

    if target not in trios_df.columns:
        return pd.Series(0, index=trios_df.index, name='sire_consistency')

    # Group daughters by sire
    sire_groups = defaultdict(list)
    for idx, row in trios_df.iterrows():
        val = row.get(target)
        sire = row.get(sire_col)
        if val is not None and not (isinstance(val, float) and np.isnan(val)) and sire:
            sire_groups[sire].append((idx, val))

    # Leave-one-out sire mean
    sire_mean_loo = pd.Series(np.nan, index=trios_df.index, name='sire_daughter_mean')
    sire_std = pd.Series(np.nan, index=trios_df.index, name='sire_daughter_std')
    sire_n = pd.Series(0, index=trios_df.index, name='sire_n_daughters')

    for sire, daughters in sire_groups.items():
        if len(daughters) < 2:
            continue
        vals = [v for _, v in daughters]
        for idx, val in daughters:
            others = [v for i, v in daughters if i != idx]
            sire_mean_loo.at[idx] = np.mean(others)
            sire_std.at[idx] = np.std(vals)
            sire_n.at[idx] = len(daughters) - 1

    return pd.DataFrame({
        'sire_daughter_mean': sire_mean_loo,
        'sire_daughter_std': sire_std,
        'sire_n_daughters': sire_n,
    })


# ============================================================
# MAIN PIPELINE
# ============================================================

def run_hybrid():
    print("=" * 90)
    print("  DSII v5 — HYBRID MODEL (Domain Knowledge + ML)")
    print("=" * 90)

    daughters, dams_dict, bulls_regname, bulls_name = load_all_data()

    # Assemble trios
    print("\nMontando trios...")
    records = []
    for daughter in daughters:
        sirename = str(daughter.get('SIRENAME', '')).strip().upper()
        sire = bulls_regname.get(sirename) or bulls_name.get(sirename)
        if not sire:
            continue
        damreg = str(daughter.get('DAMREGNUM2', '')).strip()
        dam = dams_dict.get(damreg)
        if not dam:
            continue

        rec = {'sirename': sirename, 'damreg': damreg}
        sire_ptas = {}
        dam_ptas = {}

        for trait_name, daughter_col, dam_col, bull_col in TRAIT_MAP:
            d_val = safe_float(daughter.get(daughter_col))
            s_val = safe_float(sire.get(bull_col))
            m_val = safe_float(dam.get(dam_col)) if dam_col else None
            rec[f'd_{trait_name}'] = d_val
            rec[f's_{trait_name}'] = s_val
            rec[f'm_{trait_name}'] = m_val
            if s_val is not None:
                sire_ptas[trait_name] = s_val
            if m_val is not None:
                dam_ptas[trait_name] = m_val

        rec['_sire_ptas'] = sire_ptas
        rec['_dam_ptas'] = dam_ptas
        records.append(rec)

    trios_df = pd.DataFrame(records)
    print(f"  Trios: {len(trios_df)}")

    # ============================================================
    # TRAIN HYBRID MODEL PER TRAIT
    # ============================================================
    print("\n" + "=" * 90)
    print(f"  {'Trait':>6} | {'N':>5} | {'PA MAE':>8} {'PA R2':>7} | "
          f"{'V5 MAE':>8} {'V5 R2':>7} | {'Gain MAE':>9} {'Gain R2':>8} | Feats | Best")
    print("-" * 105)

    all_results = []

    for trait_name, _, dam_col, _ in TRAIT_MAP:
        if dam_col is None:
            continue

        target = f'd_{trait_name}'

        # Check data availability
        needed_base = [target, f's_{trait_name}', f'm_{trait_name}']
        mask = trios_df[needed_base].notna().all(axis=1)
        sub = trios_df[mask].copy().reset_index(drop=True)

        if len(sub) < 50:
            continue

        y = sub[target].values

        # Build PA baseline
        s_vals = sub[f's_{trait_name}'].values
        d_vals = sub[f'm_{trait_name}'].values
        pa_pred = (s_vals + d_vals) / 2.0

        pa_mae = mean_absolute_error(y, pa_pred)
        pa_r2 = r2_score(y, pa_pred)

        # ---- BUILD FEATURE MATRIX ----
        feature_rows = []
        for idx, row in sub.iterrows():
            sire_ptas = row['_sire_ptas']
            dam_ptas = row['_dam_ptas']

            # V3 domain features for this trait
            v3_feats = build_v3_features(sire_ptas, dam_ptas, trait_name)
            if v3_feats is None:
                feature_rows.append(None)
                continue

            # Cross-trait features
            cross_feats = build_cross_trait_features(sire_ptas, dam_ptas, trait_name)
            v3_feats.update(cross_feats)

            # All other sire/dam PTAs as raw features
            for other_trait in ALL_TRAIT_NAMES:
                if other_trait == trait_name:
                    continue
                sv = sire_ptas.get(other_trait)
                dv = dam_ptas.get(other_trait)
                if sv is not None:
                    v3_feats[f'raw_s_{other_trait}'] = sv
                if dv is not None:
                    v3_feats[f'raw_d_{other_trait}'] = dv

            feature_rows.append(v3_feats)

        # Filter out None rows
        valid_mask = [f is not None for f in feature_rows]
        valid_features = [f for f in feature_rows if f is not None]

        if len(valid_features) < 50:
            continue

        feat_df = pd.DataFrame(valid_features)
        y_valid = y[valid_mask]
        sub_valid = sub[valid_mask].reset_index(drop=True)

        # Add sire consistency features (leave-one-out)
        sire_feats = build_sire_consistency_features(sub_valid, trait_name)
        feat_df = pd.concat([feat_df, sire_feats], axis=1)

        # Drop columns with too many NaN
        thresh = len(feat_df) * 0.3
        feat_df = feat_df.dropna(axis=1, thresh=int(thresh))

        # Fill remaining NaN
        for c in feat_df.columns:
            if feat_df[c].isna().any():
                feat_df[c] = feat_df[c].fillna(feat_df[c].median())

        X = feat_df.values
        n_feats = X.shape[1]

        # PA for valid rows
        pa_valid = (sub_valid[f's_{trait_name}'].values + sub_valid[f'm_{trait_name}'].values) / 2.0
        pa_mae_v = mean_absolute_error(y_valid, pa_valid)
        pa_r2_v = r2_score(y_valid, pa_valid)

        # ---- STACKING ENSEMBLE ----
        kf = KFold(n_splits=5, shuffle=True, random_state=42)

        # Individual models
        models = {
            'Ridge': Pipeline([('scaler', StandardScaler()), ('ridge', Ridge(alpha=1.0))]),
            'ElasticNet': Pipeline([('scaler', StandardScaler()),
                                    ('en', ElasticNet(alpha=0.1, l1_ratio=0.5, max_iter=5000))]),
            'RF': RandomForestRegressor(n_estimators=300, max_depth=10,
                                        min_samples_leaf=3, random_state=42, n_jobs=-1),
        }

        if HAS_XGB:
            models['XGB'] = XGBRegressor(
                n_estimators=500, max_depth=6, learning_rate=0.03,
                subsample=0.8, colsample_bytree=0.7, reg_alpha=0.1, reg_lambda=1.0,
                min_child_weight=3, random_state=42, verbosity=0
            )

        if HAS_LGBM:
            models['LGBM'] = LGBMRegressor(
                n_estimators=500, max_depth=6, learning_rate=0.03,
                subsample=0.8, colsample_bytree=0.7, reg_alpha=0.1, reg_lambda=1.0,
                min_child_samples=3, random_state=42, verbose=-1
            )

        # Evaluate each model
        model_results = {}
        for name, model in models.items():
            maes, r2s = [], []
            for train_idx, test_idx in kf.split(X):
                model.fit(X[train_idx], y_valid[train_idx])
                pred = model.predict(X[test_idx])
                maes.append(mean_absolute_error(y_valid[test_idx], pred))
                r2s.append(r2_score(y_valid[test_idx], pred))
            model_results[name] = {
                'MAE': np.mean(maes), 'MAE_std': np.std(maes),
                'R2': np.mean(r2s), 'R2_std': np.std(r2s),
            }

        # Stacking ensemble
        estimators = [
            ('ridge', Pipeline([('scaler', StandardScaler()), ('ridge', Ridge(alpha=1.0))])),
            ('rf', RandomForestRegressor(n_estimators=200, max_depth=8,
                                         min_samples_leaf=3, random_state=42, n_jobs=-1)),
        ]
        if HAS_XGB:
            estimators.append(('xgb', XGBRegressor(
                n_estimators=300, max_depth=5, learning_rate=0.03,
                subsample=0.8, colsample_bytree=0.7,
                min_child_weight=3, random_state=42, verbosity=0
            )))
        if HAS_LGBM:
            estimators.append(('lgbm', LGBMRegressor(
                n_estimators=300, max_depth=5, learning_rate=0.03,
                subsample=0.8, colsample_bytree=0.7,
                min_child_samples=3, random_state=42, verbose=-1
            )))

        stack = StackingRegressor(
            estimators=estimators,
            final_estimator=Ridge(alpha=0.5),
            cv=3, n_jobs=-1, passthrough=True
        )

        stack_maes, stack_r2s = [], []
        for train_idx, test_idx in kf.split(X):
            stack.fit(X[train_idx], y_valid[train_idx])
            pred = stack.predict(X[test_idx])
            stack_maes.append(mean_absolute_error(y_valid[test_idx], pred))
            stack_r2s.append(r2_score(y_valid[test_idx], pred))
        model_results['Stack'] = {
            'MAE': np.mean(stack_maes), 'MAE_std': np.std(stack_maes),
            'R2': np.mean(stack_r2s), 'R2_std': np.std(stack_r2s),
        }

        # Find best model
        best_name = max(model_results, key=lambda k: model_results[k]['R2'])
        best = model_results[best_name]

        v5_mae = best['MAE']
        v5_r2 = best['R2']

        gain_mae = (pa_mae_v - v5_mae) / pa_mae_v * 100
        gain_r2 = (v5_r2 - pa_r2_v) / max(abs(pa_r2_v), 0.001) * 100

        print(f"  {trait_name:>6} | {len(y_valid):>5} | {pa_mae_v:>8.3f} {pa_r2_v:>7.4f} | "
              f"{v5_mae:>8.3f} {v5_r2:>7.4f} | {gain_mae:>+8.1f}% {gain_r2:>+7.1f}% | "
              f"{n_feats:>5} | {best_name}")

        all_results.append({
            'Trait': trait_name,
            'N': len(y_valid),
            'N_Features': n_feats,
            'PA_MAE': round(pa_mae_v, 4),
            'PA_R2': round(pa_r2_v, 4),
            'V5_MAE': round(v5_mae, 4),
            'V5_R2': round(v5_r2, 4),
            'V5_Best': best_name,
            'Gain_MAE_%': round(gain_mae, 2),
            'Gain_R2_%': round(gain_r2, 2),
            **{f'{k}_MAE': round(v['MAE'], 4) for k, v in model_results.items()},
            **{f'{k}_R2': round(v['R2'], 4) for k, v in model_results.items()},
        })

    # ============================================================
    # SUMMARY
    # ============================================================
    results_df = pd.DataFrame(all_results)
    results_df.to_csv(OUTPUT_DIR / "v5_hybrid_results.csv", index=False)

    print("\n" + "=" * 90)
    print("  RESUMO POR CATEGORIA")
    print("=" * 90)

    categories = {
        'INDICES': ['TPI', 'NM$', 'CM$'],
        'PRODUCAO': ['MILK', 'FAT', 'FAT%', 'PRO', 'PRO%', 'CFP'],
        'SAUDE/FERT': ['PL', 'SCS', 'DPR', 'LIV', 'FI', 'HCR', 'CCR', 'MAST', 'FSAV'],
        'TIPO': ['PTAT', 'UDC', 'FLC'],
        'PARTO': ['SCE', 'DCE', 'SSB', 'DSB'],
    }

    for cat, traits in categories.items():
        cat_data = [r for r in all_results if r['Trait'] in traits]
        if not cat_data:
            continue
        avg_pa = np.mean([r['PA_R2'] for r in cat_data])
        avg_v5 = np.mean([r['V5_R2'] for r in cat_data])
        avg_gain = np.mean([r['Gain_MAE_%'] for r in cat_data])
        print(f"  {cat:>12}: PA R2={avg_pa:.4f}  V5 R2={avg_v5:.4f}  Gain MAE={avg_gain:+.1f}%")

    # Overall
    avg_pa = np.mean([r['PA_R2'] for r in all_results])
    avg_v5 = np.mean([r['V5_R2'] for r in all_results])
    avg_gain_mae = np.mean([r['Gain_MAE_%'] for r in all_results])
    avg_gain_r2 = np.mean([r['Gain_R2_%'] for r in all_results])

    print(f"\n  {'GERAL':>12}: PA R2={avg_pa:.4f}  V5 R2={avg_v5:.4f}")
    print(f"               Gain MAE medio: {avg_gain_mae:+.1f}%")
    print(f"               Gain R2 medio:  {avg_gain_r2:+.1f}%")

    v5_wins = sum(1 for r in all_results if r['V5_R2'] > r['PA_R2'])
    pa_wins = sum(1 for r in all_results if r['PA_R2'] >= r['V5_R2'])
    print(f"\n  V5 vence PA: {v5_wins}/{len(all_results)} traits")
    if pa_wins > 0:
        pa_win_traits = [r['Trait'] for r in all_results if r['PA_R2'] >= r['V5_R2']]
        print(f"  PA ainda vence: {pa_win_traits}")

    print(f"\n  Resultados: {OUTPUT_DIR / 'v5_hybrid_results.csv'}")

    return results_df


if __name__ == '__main__':
    run_hybrid()
