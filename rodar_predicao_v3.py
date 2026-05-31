"""
================================================================================
PREDICAO GENOMICA v3 -- RACA HOLANDESA
Motor: DSII (Dam-Specific Improvement Index) + Multi-Trait Completo
================================================================================

DIFERENCIAIS REAIS sobre PA simples:

1. DSII (Dam-Specific Improvement Index):
   - Cada femea tem perfil unico de deficiencias
   - Pesos economicos AJUSTADOS por deficiencia: traits fracos da femea
     recebem peso MAIOR ao avaliar o touro
   - Rankings de touros MUDAM conforme a femea (touro A pode ser melhor
     para femea X mas pior para femea Y)

2. Perfil de Melhoria Genetica (DeltaG):
   - Calcula a MELHORIA esperada sobre a mae: DeltaG = PA - PTA_mae
   - Diferente de PA puro: mostra o GANHO real por geracao

3. Multi-Trait Completo (30+ traits):
   - Producao: MILK, FAT, PROT, CFP
   - Fertilidade: DPR, CCR, HCR, EFC
   - Saude: SCS, PL, LIV, HLIV, Mastite, Cetose, DA, MilkFever, Metrite, RP
   - Tipo: UDC, FLC, BWC + 18 lineares (STA, STR, BD, DF, RA, TW, etc.)
   - Eficiencia: RFI, FeedSaved
   - Parto: SCE, SSB, DSB

4. Penalidade de Limite Critico:
   - Se QUALQUER trait da cria fica abaixo de limite critico, penaliza
   - "Corrente e tao forte quanto o elo mais fraco"

5. Antagonismo Producao x Funcionalidade:
   - Penaliza combos onde producao alta puxa fertilidade/saude para baixo
   - Usa correlacoes geneticas conhecidas

6. Diversificacao de Pai-da-Vaca:
   - Identifica grupos de meio-irmas (mesmo pai)
   - Recomenda variar touros entre grupos para evitar F futuro

7. Distancia ao Perfil Ideal:
   - Mede quao perto a cria esperada esta do "animal ideal"
   - Diferente por combinacao

Refs: VanRaden(2008), Cole/VanRaden(2011), Pryce(2012), Kinghorn(2011),
      Sun et al(2013), Aliloo et al(2016), CDCB NM$2025
================================================================================
"""

import os, sys, math, csv
from dataclasses import dataclass, field
from datetime import datetime
import openpyxl

# =============================================================================
# CONFIGURACAO
# =============================================================================
DOWNLOADS = os.path.expanduser("~/Downloads")
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
# Pesos economicos NM$ 2024 (CDCB, VanRaden & Cole, NM$8 revision)
# Enfase (%) e $/unidade de PTA
# Fonte: VanRaden PM, Cole JB, Parker Gaddis KL. AIPL Research Report NM$8
# ---------------------------------------------------------------------------
ECON_WEIGHTS = {
    "FAT":   2.60,   # $/lb fat (~27% enfase NM$)
    "PROT":  1.40,   # $/lb protein (~17% enfase)
    "MILK": -0.02,   # $/lb milk (~-1% enfase, penaliza excesso volume)
    "PL":   35.00,   # $/month productive life (~13% enfase)
    "SCS": -120.0,   # $/unit SCS (~7% enfase, lower = better)
    "DPR":  28.00,   # $/% daughter pregnancy rate (~7% enfase)
    "CCR":  11.00,   # $/% cow conception rate (~2% enfase)
    "HCR":   6.00,   # $/% heifer conception rate (~1% enfase)
    "LIV":  25.00,   # $/% livability (~7% enfase)
    "HLIV":  6.00,   # $/% heifer livability
    "EFC":   2.00,   # $/unit early first calving
    "UDC":  30.00,   # $/point udder composite (~4% enfase)
    "FLC":   5.00,   # $/point feet & legs composite (~3% enfase)
    "BWC":  -6.00,   # $/point body weight (~-5% enfase, menor = melhor)
    "SCE":  -3.50,   # $/% sire calving ease (~2% enfase)
    "RFI":  -1.00,   # $/lb RFI (~3% enfase, lower = better)
    "MF":  -18.0,    # $/unit milk fever (~2% saude)
    "DA":  -18.0,    # $/unit displaced abomasum
    "KET": -18.0,    # $/unit ketosis
    "MAST":-25.0,    # $/unit mastitis
    "MET": -12.0,    # $/unit metritis
    "RP":  -12.0,    # $/unit retained placenta
}

# ---------------------------------------------------------------------------
# Desvios-padrao geneticos aditivos (sigma_a) -- Holstein
# ---------------------------------------------------------------------------
# Fonte: Cole & VanRaden (2018) J. Dairy Sci. 101:5227-5236, CDCB PTA SDs
GENETIC_SD = {
    "MILK": 675, "FAT": 29, "PROT": 19, "CFP": 25,
    "SCS": 0.14, "PL": 1.85, "LIV": 1.2, "HLIV": 0.8,
    "DPR": 1.3, "CCR": 1.65, "HCR": 1.4, "EFC": 5.0,
    "UDC": 0.75, "FLC": 0.65, "BWC": 1.0,
    "NM$": 275, "TPI": 250,
    "SCE": 1.5, "SSB": 1.0, "DSB": 1.0,
    "RFI": 50,
    "MF": 0.5, "DA": 0.5, "KET": 0.5, "MAST": 1.0, "MET": 0.5, "RP": 0.5,
    "STA": 1.0, "STR": 1.0, "BD": 1.0, "DF": 1.0, "RA": 1.0, "TW": 1.0,
    "RLS": 1.0, "RLR": 1.0, "FA": 1.0, "FLS": 1.0,
    "FUA": 1.0, "RUH": 1.0, "RUW": 1.0, "UC": 1.0, "UD": 1.0,
    "FTP": 1.0, "RTP": 1.0, "TL": 1.0,
}

# ---------------------------------------------------------------------------
# Depressao endogamica por 1% de F (CDCB / literatura)
# ---------------------------------------------------------------------------
# Fonte: VanRaden(2005), Bjelland et al.(2013), Pryce et al.(2014), CDCB
INBREEDING_DEPRESSION = {
    "NM$": -25.0, "MILK": -28.5, "FAT": -1.0, "PROT": -0.9,
    "PL": -0.35, "DPR": -0.03, "SCS": +0.007, "LIV": -0.15,
    "CCR": -0.025, "HCR": -0.02,
}

# ---------------------------------------------------------------------------
# Alvos ideais por trait (Holstein elite, metas de selecao)
# Traits onde MAIOR e melhor
# ---------------------------------------------------------------------------
BREED_IDEAL = {
    "MILK": 1800, "FAT": 90, "PROT": 60, "CFP": 70,
    "DPR": 1.5, "CCR": 2.0, "HCR": 2.0, "EFC": 2.0,
    "PL": 5.0, "LIV": 2.0, "HLIV": 1.0,
    "UDC": 1.5, "FLC": 1.0,
    "RFI": -50, "MF": -0.5, "DA": -0.5, "KET": -0.5,
    "MAST": -1.0, "MET": -0.5, "RP": -0.5,
}

# Traits onde MENOR e melhor (ou abs menor)
LOWER_IS_BETTER = {"SCS", "BWC", "SCE", "SSB", "DSB", "RFI",
                    "MF", "DA", "KET", "MAST", "MET", "RP"}

# ---------------------------------------------------------------------------
# Limites criticos: abaixo disso, penaliza pesadamente
# ---------------------------------------------------------------------------
CRITICAL_THRESHOLDS = {
    "DPR": -2.0,   # fertilidade critica
    "PL":  -1.0,   # longevidade critica
    "SCS":  3.30,  # mastite cronica (acima = ruim)
    "UDC": -1.0,   # ubere muito ruim
    "FLC": -1.0,   # pernas muito ruins
    "SCE":  5.0,   # parto dificil (acima = ruim)
}

# ---------------------------------------------------------------------------
# Correlacoes geneticas entre traits chave (Holstein, aproximado)
# Usadas para estimar antagonismo e risco correlacionado
# ---------------------------------------------------------------------------
# Fonte: VanRaden et al. (2004, 2014, 2018), CDCB multi-trait model
GENETIC_CORRELATIONS = {
    ("MILK", "DPR"):  -0.35,
    ("MILK", "CCR"):  -0.30,
    ("MILK", "SCS"):   0.10,
    ("MILK", "PL"):   -0.15,
    ("MILK", "UDC"):  -0.15,
    ("FAT", "PROT"):   0.65,
    ("FAT", "DPR"):   -0.25,
    ("FAT", "CCR"):   -0.20,
    ("PROT", "DPR"):  -0.30,
    ("PROT", "CCR"):  -0.25,
    ("DPR", "CCR"):    0.55,
    ("DPR", "PL"):     0.45,
    ("CCR", "PL"):     0.30,
    ("UDC", "PL"):     0.25,
    ("FLC", "PL"):     0.10,
    ("SCS", "PL"):    -0.30,
    ("SCS", "UDC"):   -0.25,
    ("SCS", "MAST"):   0.70,
}

# Razoes dominancia/aditivo (d²/σ²a) por trait (Holstein)
# Fonte: Sun et al. (2014) J. Dairy Sci. 97:3852-3861, VanRaden et al. (2014)
DOMINANCE_RATIOS = {
    "MILK": 0.12, "FAT": 0.10, "PROT": 0.12,
    "SCS": 0.14, "PL": 0.15, "DPR": 0.20, "CCR": 0.15,
    "UDC": 0.11, "FLC": 0.08, "LIV": 0.12,
}


# =============================================================================
# ESTRUTURAS
# =============================================================================

@dataclass
class Animal:
    id: str
    name: str
    sex: str
    num: int = 0
    reg: str = ""
    naab: str = ""
    tpi: float = 0.0
    nm_dollar: float = 0.0
    ptas: dict = field(default_factory=dict)
    inbreeding: float = 0.085
    sire_name: str = ""
    mgs_name: str = ""
    mggs_name: str = ""
    beta_casein: str = ""
    kappa_casein: str = ""
    haplotypes: str = ""


@dataclass
class Prediction:
    rank: int = 0
    sire: Animal = None
    dam: Animal = None

    # Parent Average (referencia)
    nm_pa: float = 0.0
    tpi_pa: float = 0.0

    # Endogamia
    expected_f: float = 0.0
    endo_depression: float = 0.0
    nm_corrected: float = 0.0
    tpi_corrected: float = 0.0
    tpi_ic_lower: float = 0.0
    tpi_ic_upper: float = 0.0
    delta_tpi: float = 0.0

    # DeltaG: melhoria sobre a mae
    delta_g: dict = field(default_factory=dict)  # trait -> melhoria
    delta_nm: float = 0.0  # melhoria em NM$ sobre a mae

    # DSII: Dam-Specific Improvement Index
    dsii: float = 0.0

    # Trait predictions (PA corrigido)
    trait_pred: dict = field(default_factory=dict)

    # Mendelian sampling
    nm_sd: float = 0.0
    nm_ic_lower: float = 0.0
    nm_ic_upper: float = 0.0

    # Probabilidades
    prob_above_avg: float = 0.0
    prob_top25: float = 0.0
    prob_top10: float = 0.0

    # Penalidades e bonus
    critical_penalty: float = 0.0      # traits abaixo do critico
    antagonism_penalty: float = 0.0    # antagonismo prod x func
    dominance_bonus: float = 0.0       # desvio de dominancia
    distance_to_ideal: float = 0.0     # distancia ao perfil ideal

    # Caseinas e haplotipos
    haplo_conflicts: list = field(default_factory=list)
    beta_info: str = ""
    kappa_info: str = ""
    casein_score: int = 0

    # Score final v3
    final_score: float = 0.0

    # Para comparacao: rank que teria no PA simples
    rank_simple_pa: int = 0


# =============================================================================
# FUNCOES AUXILIARES
# =============================================================================

def sf(val, default=0.0):
    if val is None: return default
    try: return float(val)
    except: return default


def normal_cdf(z):
    if z > 6: return 1.0
    if z < -6: return 0.0
    a = abs(z)
    t = 1.0 / (1.0 + 0.2316419 * a)
    pdf = math.exp(-0.5 * a * a) / math.sqrt(2 * math.pi)
    cdf = 1.0 - pdf * t * (0.319381530 + t * (-0.356563782 + t * (
        1.781477937 + t * (-1.821255978 + t * 1.330274429))))
    return cdf if z >= 0 else 1.0 - cdf


# =============================================================================
# CARGA DE DADOS (TODOS OS TRAITS)
# =============================================================================

def load_sires(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    animals = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        naab = row[1].value
        if not naab: continue
        a = Animal(
            id=str(naab), name=str(row[2].value or "?"), sex="M",
            naab=str(naab), reg=str(row[3].value or ""),
            tpi=sf(row[13].value), nm_dollar=sf(row[14].value),
            ptas={
                "MILK": sf(row[18].value), "FAT": sf(row[19].value),
                "PROT": sf(row[21].value), "CFP": sf(row[23].value),
                "SCS": sf(row[24].value),
                "DPR": sf(row[25].value), "CCR": sf(row[26].value),
                "PL": sf(row[27].value), "LIV": sf(row[28].value),
                "MF": sf(row[30].value), "DA": sf(row[31].value),
                "KET": sf(row[32].value), "MAST": sf(row[33].value),
                "MET": sf(row[34].value), "RP": sf(row[35].value),
                "HCR": sf(row[36].value), "EFC": sf(row[39].value),
                "HLIV": sf(row[40].value),
                "RFI": sf(row[41].value),
                "UDC": sf(row[44].value), "FLC": sf(row[45].value),
                "BWC": sf(row[46].value),
                "STA": sf(row[47].value), "STR": sf(row[48].value),
                "BD": sf(row[49].value), "DF": sf(row[50].value),
                "RA": sf(row[51].value), "TW": sf(row[52].value),
                "RLS": sf(row[53].value), "RLR": sf(row[54].value),
                "FA": sf(row[55].value), "FLS": sf(row[56].value),
                "FUA": sf(row[57].value), "RUH": sf(row[58].value),
                "RUW": sf(row[59].value), "UC": sf(row[60].value),
                "UD": sf(row[61].value),
                "FTP": sf(row[62].value), "RTP": sf(row[63].value),
                "TL": sf(row[64].value),
                "SCE": sf(row[65].value), "SSB": sf(row[66].value),
                "DSB": sf(row[67].value),
            },
            sire_name=str(row[68].value or ""),
            mgs_name=str(row[69].value or ""),
            mggs_name=str(row[70].value or "") if len(row) > 70 else "",
            beta_casein=str(row[5].value or ""),
            kappa_casein=str(row[6].value or ""),
            haplotypes=str(row[4].value or ""),
        )
        animals.append(a)
    return animals


def load_dams(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    animals = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        naab = row[2].value
        if not naab: continue
        a = Animal(
            id=str(naab),
            name=f"#{row[0].value} Reg.{row[3].value or '?'}",
            sex="F", num=int(row[0].value or 0),
            naab=str(naab), reg=str(row[3].value or ""),
            tpi=sf(row[14].value), nm_dollar=sf(row[15].value),
            ptas={
                "MILK": sf(row[17].value), "FAT": sf(row[18].value),
                "PROT": sf(row[20].value), "CFP": sf(row[22].value),
                "SCS": sf(row[23].value),
                "DPR": sf(row[24].value), "CCR": sf(row[25].value),
                "PL": sf(row[26].value), "LIV": sf(row[27].value),
                "MF": sf(row[28].value), "DA": sf(row[29].value),
                "KET": sf(row[30].value), "MAST": sf(row[31].value),
                "MET": sf(row[32].value), "RP": sf(row[33].value),
                "HCR": sf(row[34].value), "EFC": sf(row[37].value),
                "HLIV": sf(row[38].value),
                "RFI": sf(row[39].value),
                "UDC": sf(row[42].value), "FLC": sf(row[43].value),
                "BWC": sf(row[44].value),
                "STA": sf(row[45].value), "STR": sf(row[46].value),
                "BD": sf(row[47].value), "DF": sf(row[48].value),
                "RA": sf(row[49].value), "TW": sf(row[50].value),
                "RLS": sf(row[51].value), "RLR": sf(row[52].value),
                "FA": sf(row[53].value), "FLS": sf(row[54].value),
                "FUA": sf(row[55].value), "RUH": sf(row[56].value),
                "RUW": sf(row[57].value), "UC": sf(row[58].value),
                "UD": sf(row[59].value),
                "FTP": sf(row[60].value), "RTP": sf(row[61].value),
                "TL": sf(row[62].value),
                "SCE": sf(row[63].value), "SSB": sf(row[64].value),
                "DSB": sf(row[65].value),
            },
            sire_name=str(row[66].value or ""),
            mgs_name=str(row[67].value or ""),
            mggs_name=str(row[68].value or "") if len(row) > 68 else "",
            beta_casein=str(row[5].value or ""),
            kappa_casein=str(row[6].value or ""),
            haplotypes=str(row[4].value or ""),
        )
        animals.append(a)
    return animals


# =============================================================================
# PERFIL DE DEFICIENCIAS DA FEMEA
# =============================================================================

def compute_dam_deficiency_profile(dam):
    """
    Calcula o perfil de deficiencias da femea: quais traits estao abaixo
    do ideal da raca e quanto.

    Retorna dict: trait -> deficiency_factor (0 = nao deficiente, >0 = deficiente)
    Quanto maior, mais a femea precisa de melhoria nesse trait.
    """
    profile = {}
    for trait, ideal in BREED_IDEAL.items():
        val = dam.ptas.get(trait, 0)
        sd = GENETIC_SD.get(trait, 1)

        if trait in LOWER_IS_BETTER:
            # Para traits "menor melhor", deficiencia = val > ideal
            deficit = (val - ideal) / sd  # positivo = deficiente
        else:
            # Para traits "maior melhor", deficiencia = val < ideal
            deficit = (ideal - val) / sd  # positivo = deficiente

        profile[trait] = max(0, deficit)

    return profile


# =============================================================================
# DSII: DAM-SPECIFIC IMPROVEMENT INDEX
# =============================================================================

def compute_dsii(sire, dam, deficiency_profile):
    """
    Indice de Melhoria Especifico por Femea.

    Para cada trait:
    1. Calcula DeltaG = (PTA_touro - PTA_femea) / 2  (melhoria sobre a mae)
    2. Multiplica pelo peso economico
    3. AMPLIFICA o peso se a femea e deficiente nesse trait

    Resultado: score que reflete quanto ESTE touro melhora ESTA femea especifica.
    O mesmo touro pode ter DSII diferente para femeas diferentes!
    """
    score = 0.0

    for trait, econ_w in ECON_WEIGHTS.items():
        s_val = sire.ptas.get(trait, 0)
        d_val = dam.ptas.get(trait, 0)

        # DeltaG: melhoria esperada sobre a mae
        delta_g = (s_val - d_val) / 2.0

        # Fator de amplificacao baseado na deficiencia da femea
        deficiency = deficiency_profile.get(trait, 0)

        # Peso ajustado: base economico * (1 + amplificacao por deficiencia)
        # Se femea e deficiente nesse trait, peso sobe ate 3x
        amplification = 1.0 + min(2.0, deficiency * 0.8)

        # Para traits "lower is better", DeltaG negativo = melhoria
        if trait in LOWER_IS_BETTER:
            contribution = -delta_g * abs(econ_w) * amplification
        else:
            contribution = delta_g * econ_w * amplification

        score += contribution

    return score


# =============================================================================
# PENALIDADE DE LIMITES CRITICOS
# =============================================================================

def compute_critical_penalty(trait_pred):
    """
    Penaliza se qualquer trait predito fica abaixo de limite critico.
    "Corrente tao forte quanto o elo mais fraco."
    """
    penalty = 0.0
    for trait, threshold in CRITICAL_THRESHOLDS.items():
        pred = trait_pred.get(trait, 0)
        if trait in LOWER_IS_BETTER:
            # threshold e o maximo aceitavel (ex: SCS < 3.30)
            if pred > threshold:
                penalty += (pred - threshold) * 20
        else:
            # threshold e o minimo aceitavel (ex: DPR > -2.0)
            if pred < threshold:
                penalty += (threshold - pred) * 20
    return penalty


# =============================================================================
# ANTAGONISMO PRODUCAO x FUNCIONALIDADE
# =============================================================================

def compute_antagonism(sire, dam):
    """
    Penaliza quando a combinacao tem producao muito alta mas
    funcionalidade fraca, considerando correlacoes geneticas antagonicas.
    """
    penalty = 0.0

    for (t1, t2), corr in GENETIC_CORRELATIONS.items():
        if corr >= 0:
            continue  # so interessa antagonismo (corr negativa)

        s1 = sire.ptas.get(t1, 0)
        d1 = dam.ptas.get(t1, 0)
        s2 = sire.ptas.get(t2, 0)
        d2 = dam.ptas.get(t2, 0)

        sd1 = GENETIC_SD.get(t1, 1)
        sd2 = GENETIC_SD.get(t2, 1)

        pa1 = (s1 + d1) / 2.0
        pa2 = (s2 + d2) / 2.0

        z1 = pa1 / sd1
        z2 = pa2 / sd2

        # Se trait1 e alto mas trait2 e baixo (antagonismo realizado)
        if z1 > 1.0 and z2 < -0.5:
            severity = abs(corr) * (z1 - z2) * 3
            penalty += severity

    return penalty


# =============================================================================
# DOMINANCIA (efeito nao-aditivo)
# =============================================================================

def compute_dominance_bonus(expected_f):
    """
    Estima o bonus de dominancia (heterose intra-raca).
    Pais menos aparentados -> mais heterozigosidade -> mais dominancia positiva.
    """
    total = 0.0
    heterozygosity = 1.0 - expected_f

    for trait, d_ratio in DOMINANCE_RATIOS.items():
        econ_w = ECON_WEIGHTS.get(trait, 0)
        sd = GENETIC_SD.get(trait, 1)
        # Desvio de dominancia esperado (proporcional a heterozigosidade)
        # d = heterozygosity * sqrt(d²) - baseline
        dom_sd = math.sqrt(d_ratio) * sd
        # Bonus = partial recovery through heterozygosity
        bonus = heterozygosity * dom_sd * 0.15  # conservative estimate
        total += bonus * abs(econ_w) / 100  # normalizado

    return total


# =============================================================================
# DISTANCIA AO PERFIL IDEAL
# =============================================================================

def compute_distance_to_ideal(trait_pred):
    """
    Distancia euclidiana ponderada do perfil predito ao perfil ideal.
    Menor distancia = melhor.
    """
    dist_sq = 0.0
    for trait, ideal in BREED_IDEAL.items():
        pred = trait_pred.get(trait, 0)
        sd = GENETIC_SD.get(trait, 1)
        w = abs(ECON_WEIGHTS.get(trait, 1))

        if trait in LOWER_IS_BETTER:
            diff = max(0, pred - ideal) / sd
        else:
            diff = max(0, ideal - pred) / sd

        dist_sq += (diff ** 2) * (w / 30)  # normalizado

    return math.sqrt(dist_sq)


# =============================================================================
# ENDOGAMIA
# =============================================================================

def estimate_inbreeding(sire, dam, all_sire_names):
    """
    Estima F da cria.
    Usa pedigree + conhecimento de ancestrais comuns na raca.
    """
    f = 0.03125  # background Holstein (~3.125% base)

    sn = sire.sire_name
    dn = dam.sire_name
    mn_s = sire.mgs_name
    mn_d = dam.mgs_name
    mg_s = sire.mggs_name

    # Mesmo pai -> meio-irmaos -> F=12.5%
    if sn and dn and sn == dn and sn not in ("", "None"):
        f = max(f, 0.125)

    # Pai do touro = avo materno da vaca
    if sn and mn_d and sn == mn_d and sn not in ("", "None"):
        f = max(f, 0.0625)

    # MGS do touro = pai da vaca
    if mn_s and dn and mn_s == dn and mn_s not in ("", "None"):
        f = max(f, 0.0625)

    # MGGS do touro = pai da vaca
    if mg_s and dn and mg_s == dn and mg_s not in ("", "None"):
        f = max(f, 0.03125)

    # Pai do touro = pai de outro touro na lista (relacao indireta)
    # Touros compartilhando pai = potencial acumulo futuro de F
    if sn in all_sire_names and sn not in ("", "None"):
        f += 0.005  # pequeno incremento por concentracao genetica

    # Ajuste pelo nivel de endogamia parental
    f += (sire.inbreeding + dam.inbreeding) * 0.03

    return min(f, 0.25)


# =============================================================================
# CASEINAS E HAPLOTIPOS
# =============================================================================

LETHAL_HAPLOTYPES = {"HH1","HH2","HH3","HH4","HH5","HH6","HCD","HHR","HMW3","HMW4","HHP"}

def check_haplotypes(sire, dam):
    s_hap = set(str(sire.haplotypes or "").replace(",", " ").split()) & LETHAL_HAPLOTYPES
    d_hap = set(str(dam.haplotypes or "").replace(",", " ").split()) & LETHAL_HAPLOTYPES
    return sorted(s_hap & d_hap)


def check_caseins(sire, dam):
    sb = sire.beta_casein or ""
    db = dam.beta_casein or ""
    sk = sire.kappa_casein or ""
    dk = dam.kappa_casein or ""
    score = 0
    beta_info = f"{sb}x{db}" if sb or db else "N/A"
    kappa_info = f"{sk}x{dk}" if sk or dk else "N/A"

    if "A2A2" in sb and "A2A2" in db:
        score += 3; beta_info += "=100%A2"
    elif ("A2A2" in sb and "A1A2" in db) or ("A1A2" in sb and "A2A2" in db):
        score += 1; beta_info += "=50%A2"
    elif "A1A2" in sb and "A1A2" in db:
        score += 0; beta_info += "=25%A2"
    elif "A1A1" in sb or "A1A1" in db:
        score -= 2

    if "BB" in sk and "BB" in dk:
        score += 2; kappa_info += "=100%BB"
    elif ("BB" in sk and "AB" in dk) or ("AB" in sk and "BB" in dk):
        score += 1; kappa_info += "=50%BB"
    elif ("BB" in sk and "AA" in dk) or ("AA" in sk and "BB" in dk):
        score += 1; kappa_info += "=100%AB"

    return beta_info, kappa_info, score


# =============================================================================
# MOTOR DE PREDICAO v3
# =============================================================================

def predict_mating_v3(sire, dam, deficiency_profile, pop_avg_nm, pop_sd_nm, all_sire_names):
    p = Prediction(sire=sire, dam=dam)

    # 1. Parent Average (referencia - igual ao PA simples)
    p.nm_pa = (sire.nm_dollar + dam.nm_dollar) / 2.0
    p.tpi_pa = (sire.tpi + dam.tpi) / 2.0

    # 2. Endogamia (variavel por combinacao)
    p.expected_f = estimate_inbreeding(sire, dam, all_sire_names)
    f_pct = p.expected_f * 100.0
    p.endo_depression = INBREEDING_DEPRESSION["NM$"] * f_pct
    p.nm_corrected = p.nm_pa + p.endo_depression

    # 2b. TPI corrigido (TPI nao tem depressao endogamica direta, mas
    #     como e composto de traits que sofrem depressao, aplicamos
    #     proporcional: ~60% do TPI sao traits economicos afetados por F)
    tpi_endo = -20.0 * f_pct  # ~$20/1%F de impacto nos componentes do TPI
    p.tpi_corrected = p.tpi_pa + tpi_endo
    p.delta_tpi = p.tpi_corrected - dam.tpi

    # 2c. IC do TPI (Mendelian sampling)
    sigma2_tpi = GENETIC_SD["TPI"] ** 2
    var_tpi = 0.5 * (1.0 - p.expected_f) * sigma2_tpi + 0.20 * sigma2_tpi
    tpi_sd = math.sqrt(var_tpi)
    p.tpi_ic_lower = p.tpi_corrected - 1.96 * tpi_sd
    p.tpi_ic_upper = p.tpi_corrected + 1.96 * tpi_sd

    # 3. Trait predictions (PA + correcao endogamia por trait)
    for trait in sire.ptas:
        if trait not in dam.ptas:
            continue
        pa = (sire.ptas[trait] + dam.ptas.get(trait, 0)) / 2.0
        dep = INBREEDING_DEPRESSION.get(trait, 0) * f_pct
        p.trait_pred[trait] = pa + dep

    # 4. DeltaG: melhoria sobre a mae (DIFERENCIAL do v3)
    for trait in sire.ptas:
        d_val = dam.ptas.get(trait, 0)
        pa = p.trait_pred.get(trait, (sire.ptas[trait] + d_val) / 2.0)
        p.delta_g[trait] = pa - d_val

    p.delta_nm = p.nm_corrected - dam.nm_dollar

    # 5. DSII: Dam-Specific Improvement Index (DIFERENCIAL PRINCIPAL)
    p.dsii = compute_dsii(sire, dam, deficiency_profile)

    # 6. Mendelian Sampling variance
    sigma2_nm = GENETIC_SD["NM$"] ** 2
    var_mendelian = 0.5 * (1.0 - p.expected_f) * sigma2_nm
    var_uncertainty = 0.20 * sigma2_nm  # incerteza media
    total_var = var_mendelian + var_uncertainty
    p.nm_sd = math.sqrt(total_var)
    p.nm_ic_lower = p.nm_corrected - 1.96 * p.nm_sd
    p.nm_ic_upper = p.nm_corrected + 1.96 * p.nm_sd

    # 7. Probabilidades
    if p.nm_sd > 0:
        z_avg = (p.nm_corrected - pop_avg_nm) / p.nm_sd
        z_25 = (pop_avg_nm + 0.674 * pop_sd_nm - p.nm_corrected) / p.nm_sd
        z_10 = (pop_avg_nm + 1.282 * pop_sd_nm - p.nm_corrected) / p.nm_sd
        p.prob_above_avg = normal_cdf(z_avg)
        p.prob_top25 = 1.0 - normal_cdf(z_25)
        p.prob_top10 = 1.0 - normal_cdf(z_10)

    # 8. Penalidade de limites criticos
    p.critical_penalty = compute_critical_penalty(p.trait_pred)

    # 9. Antagonismo producao x funcionalidade
    p.antagonism_penalty = compute_antagonism(sire, dam)

    # 10. Dominancia
    p.dominance_bonus = compute_dominance_bonus(p.expected_f)

    # 11. Distancia ao perfil ideal
    p.distance_to_ideal = compute_distance_to_ideal(p.trait_pred)

    # 12. Haplotipos e caseinas
    p.haplo_conflicts = check_haplotypes(sire, dam)
    p.beta_info, p.kappa_info, p.casein_score = check_caseins(sire, dam)

    # 13. SCORE FINAL v3
    # Normaliza DSII para escala 0-100
    # (sera normalizado globalmente depois)
    p.final_score = p.dsii  # temporario, sera recalculado

    return p


def compute_final_scores(predictions, pop_avg_nm, pop_sd_nm):
    """
    Calcula score final apos todas as predicoes serem geradas.
    Normaliza DSII e compoe com outros fatores.
    """
    # Normaliza DSII globalmente
    all_dsii = [p.dsii for p in predictions]
    dsii_min = min(all_dsii)
    dsii_max = max(all_dsii)
    dsii_range = dsii_max - dsii_min if dsii_max != dsii_min else 1

    # Normaliza distancia ao ideal
    all_dist = [p.distance_to_ideal for p in predictions]
    dist_max = max(all_dist) if max(all_dist) > 0 else 1

    # Normaliza NM$ corrigido
    all_nm = [p.nm_corrected for p in predictions]
    nm_min = min(all_nm)
    nm_max = max(all_nm)
    nm_range = nm_max - nm_min if nm_max != nm_min else 1

    # Normaliza antagonismo
    all_ant = [p.antagonism_penalty for p in predictions]
    ant_max = max(all_ant) if max(all_ant) > 0 else 1

    for p in predictions:
        # Componentes normalizados (0-100)
        dsii_norm = ((p.dsii - dsii_min) / dsii_range) * 100
        nm_norm = ((p.nm_corrected - nm_min) / nm_range) * 100
        dist_norm = (1.0 - p.distance_to_ideal / dist_max) * 100
        ant_norm = (1.0 - p.antagonism_penalty / ant_max) * 100 if ant_max > 0 else 100

        # Score final composto
        # DSII tem peso MAIOR que NM$ puro: 35% vs 20%
        # Isso e o que gera rankings diferentes do PA simples
        p.final_score = (
            0.35 * dsii_norm +                     # Melhoria especifica da femea
            0.20 * nm_norm +                        # Merito economico bruto
            0.15 * dist_norm +                      # Proximidade ao ideal
            0.10 * ant_norm +                       # Sem antagonismo
            0.10 * max(0, 100 - p.critical_penalty) + # Sem traits criticos
            0.05 * max(0, min(100, 100 - p.expected_f * 800)) + # Endogamia
            0.05 * min(100, p.prob_top25 * 100)     # Probabilidade top25
            - len(p.haplo_conflicts) * 15            # Haplotipos letais
            + p.casein_score * 1.5                   # Caseinas
            + p.dominance_bonus                      # Dominancia
        )
        p.final_score = max(0, min(100, p.final_score))


# =============================================================================
# MAIN
# =============================================================================

def main():
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    print("=" * 130)
    print("  PREDICAO GENOMICA v3 -- Motor DSII (Dam-Specific Improvement Index)")
    print("  Diferenciais: DSII + Multi-Trait (30+) + Limites Criticos + Antagonismo + Dominancia")
    print("  Data:", datetime.now().strftime("%Y-%m-%d %H:%M"))
    print("=" * 130)

    # Carrega dados
    sires = load_sires(os.path.join(DOWNLOADS, "Lista de touros.xlsx"))
    dams = load_dams(os.path.join(DOWNLOADS, "Lista de f\u00eameas.xlsx"))

    # Coleta nomes de pais dos touros para deteccao de concentracao
    all_sire_names = set()
    for s in sires:
        if s.sire_name and s.sire_name != "None":
            all_sire_names.add(s.sire_name)

    print(f"\nTouros: {len(sires)}")
    for s in sires:
        print(f"  {s.name:<15} TPI={s.tpi:>5.0f}  NM$={s.nm_dollar:>5.0f}  "
              f"Milk={s.ptas['MILK']:>5.0f}  Fat={s.ptas['FAT']:>4.0f}  "
              f"DPR={s.ptas['DPR']:>+5.1f}  PL={s.ptas['PL']:>4.1f}  "
              f"UDC={s.ptas['UDC']:>+5.2f}  FLC={s.ptas['FLC']:>+5.2f}  "
              f"SCS={s.ptas['SCS']:.2f}  "
              f"Beta={s.beta_casein}  Kappa={s.kappa_casein}  "
              f"Pai={s.sire_name}")

    print(f"\nFemeas: {len(dams)}")
    print(f"  TPI range: {min(d.tpi for d in dams):.0f} - {max(d.tpi for d in dams):.0f}")
    print(f"  NM$ range: {min(d.nm_dollar for d in dams):.0f} - {max(d.nm_dollar for d in dams):.0f}")

    # Grupos de pai-da-vaca (diversificacao)
    dam_sire_groups = {}
    for d in dams:
        sn = d.sire_name if d.sire_name and d.sire_name != "None" else "Desconhecido"
        dam_sire_groups.setdefault(sn, []).append(d)

    print(f"\n  Grupos de pai-da-vaca:")
    for sn, ds in sorted(dam_sire_groups.items(), key=lambda x: -len(x[1])):
        print(f"    {sn}: {len(ds)} filhas")

    # Populacao de referencia
    all_nm = [d.nm_dollar for d in dams]
    pop_avg = sum(all_nm) / len(all_nm)
    pop_sd = max((sum((x - pop_avg)**2 for x in all_nm) / (len(all_nm)-1))**0.5, 100)
    print(f"\n  Media NM$ femeas: ${pop_avg:.0f}  |  SD: ${pop_sd:.0f}")

    # =========================================================================
    # ANALISE DE DEFICIENCIAS POR FEMEA
    # =========================================================================
    print(f"\n{'='*130}")
    print("  PERFIL DE DEFICIENCIAS POR FEMEA (traits abaixo do ideal)")
    print(f"{'='*130}")

    dam_profiles = {}
    for d in dams:
        profile = compute_dam_deficiency_profile(d)
        dam_profiles[d.id] = profile

    # Mostra top deficiencias por femea
    print(f"\n  {'Femea':<22} | {'Deficiencias Principais (trait: severidade)'}")
    print("  " + "-" * 100)
    for d in sorted(dams, key=lambda x: -x.tpi)[:15]:
        prof = dam_profiles[d.id]
        top_def = sorted(prof.items(), key=lambda x: -x[1])[:5]
        def_str = "  ".join(f"{t}:{v:.1f}" for t, v in top_def if v > 0.3)
        print(f"  {d.name:<22} | {def_str or 'Nenhuma deficiencia critica'}")

    # =========================================================================
    # PREDICOES
    # =========================================================================
    print(f"\n  Rodando {len(sires) * len(dams)} predicoes...")

    predictions = []
    for s in sires:
        for d in dams:
            prof = dam_profiles[d.id]
            p = predict_mating_v3(s, d, prof, pop_avg, pop_sd, all_sire_names)
            predictions.append(p)

    # Calcula scores finais (normalizacao global)
    compute_final_scores(predictions, pop_avg, pop_sd)

    # Ranking por score v3
    predictions.sort(key=lambda x: x.final_score, reverse=True)
    for i, p in enumerate(predictions):
        p.rank = i + 1

    # Ranking por PA simples (para comparacao)
    by_pa = sorted(predictions, key=lambda x: x.nm_pa, reverse=True)
    for i, p in enumerate(by_pa):
        p.rank_simple_pa = i + 1

    # =========================================================================
    # COMPARACAO: v3 vs PA SIMPLES
    # =========================================================================
    print(f"\n{'='*130}")
    print("  COMPARACAO: RANKING v3 (DSII) vs RANKING PA SIMPLES")
    print("  >> Diferenca entre os dois rankings mostra o VALOR ADICIONADO do v3 <<")
    print(f"{'='*130}")

    print(f"\n  {'Rk_v3':>5} | {'Rk_PA':>5} | {'Diff':>5} | {'Touro':<12} | {'Femea':<22} | "
          f"{'NM$PA':>6} | {'NM$Cor':>6} | {'DSII':>7} | {'Score':>5} | "
          f"{'Antag':>5} | {'Crit':>5} | {'DistIdeal':>9}")
    print("  " + "-" * 125)

    for p in predictions[:50]:
        diff = p.rank_simple_pa - p.rank
        diff_str = f"{diff:>+5d}" if diff != 0 else "    ="
        print(
            f"  {p.rank:>5} | {p.rank_simple_pa:>5} | {diff_str} | "
            f"{p.sire.name:<12.12} | {p.dam.name:<22.22} | "
            f"${p.nm_pa:>5.0f} | ${p.nm_corrected:>5.0f} | "
            f"{p.dsii:>7.0f} | {p.final_score:>5.1f} | "
            f"{p.antagonism_penalty:>5.1f} | {p.critical_penalty:>5.1f} | "
            f"{p.distance_to_ideal:>9.2f}"
        )

    # =========================================================================
    # MELHOR TOURO POR FEMEA (RANKING DIFERENCIADO)
    # =========================================================================
    print(f"\n{'='*130}")
    print("  MELHOR TOURO POR FEMEA -- Ranking DSII (note: ranking pode DIFERIR do PA)")
    print(f"{'='*130}")

    best_per_dam = {}
    all_per_dam = {}
    for p in predictions:
        did = p.dam.id
        all_per_dam.setdefault(did, []).append(p)

    # Ordena dentro de cada femea
    for did in all_per_dam:
        all_per_dam[did].sort(key=lambda x: x.final_score, reverse=True)

    print(f"\n  {'Femea':<22} | {'TPI':>4} | {'NM$F':>5} | "
          f"{'1o(v3)':<12} | {'Scr1':>5} | {'DeltaNM':>7} | "
          f"{'2o(v3)':<12} | {'Scr2':>5} | "
          f"{'3o(v3)':<12} | {'Scr3':>5} | "
          f"{'1o(PA)':<12} | {'Defic.Principal'}")
    print("  " + "-" * 130)

    for d in sorted(dams, key=lambda x: -x.tpi):
        ranked = all_per_dam[d.id]
        p1, p2, p3 = ranked[0], ranked[1], ranked[2]

        # Melhor por PA simples (para comparacao)
        by_pa_dam = sorted(all_per_dam[d.id], key=lambda x: -x.nm_pa)
        pa_best = by_pa_dam[0]

        # Deficiencia principal
        prof = dam_profiles[d.id]
        top_def = sorted(prof.items(), key=lambda x: -x[1])
        def_main = top_def[0][0] if top_def[0][1] > 0.3 else "-"

        # Marca se ranking v3 difere do PA
        marker = " *" if p1.sire.name != pa_best.sire.name else ""

        print(
            f"  {d.name:<22} | {d.tpi:>4.0f} | ${d.nm_dollar:>4.0f} | "
            f"{p1.sire.name:<12.12} | {p1.final_score:>5.1f} | ${p1.delta_nm:>6.0f} | "
            f"{p2.sire.name:<12.12} | {p2.final_score:>5.1f} | "
            f"{p3.sire.name:<12.12} | {p3.final_score:>5.1f} | "
            f"{pa_best.sire.name:<12.12} | {def_main}{marker}"
        )

    # Conta quantas femeas tiveram ranking diferente
    diff_count = 0
    for d in dams:
        ranked = all_per_dam[d.id]
        by_pa_dam = sorted(all_per_dam[d.id], key=lambda x: -x.nm_pa)
        if ranked[0].sire.name != by_pa_dam[0].sire.name:
            diff_count += 1

    print(f"\n  * = Ranking v3 DIFERE do PA simples")
    print(f"  Femeas com ranking diferente: {diff_count}/{len(dams)} "
          f"({diff_count/len(dams)*100:.0f}%)")

    # =========================================================================
    # RELATORIO POR TOURO COM DELTAS
    # =========================================================================
    print(f"\n{'='*130}")
    print("  RELATORIO POR TOURO -- Mostrando DeltaG (melhoria sobre a mae)")
    print(f"{'='*130}")

    for s in sires:
        preds_s = [p for p in predictions if p.sire.id == s.id]
        preds_s.sort(key=lambda x: x.final_score, reverse=True)

        print(f"\n{'='*130}")
        print(f"  TOURO: {s.name} ({s.naab})  |  TPI={s.tpi:.0f}  NM$={s.nm_dollar:.0f}  |  "
              f"Milk={s.ptas['MILK']:.0f}  Fat={s.ptas['FAT']:.0f}  Prot={s.ptas['PROT']:.0f}  |  "
              f"DPR={s.ptas['DPR']:+.1f}  CCR={s.ptas['CCR']:+.1f}  PL={s.ptas['PL']:.1f}  |  "
              f"UDC={s.ptas['UDC']:+.2f}  FLC={s.ptas['FLC']:+.2f}  |  "
              f"Saude: SCS={s.ptas['SCS']:.2f}  Mast={s.ptas.get('MAST',0):+.1f}  |  "
              f"Pai={s.sire_name}")
        print(f"{'='*130}")

        hdr = (
            f"{'Rk':>3} | {'Femea':<22} | {'NM$F':>5} | {'NM$PA':>6} | {'NM$C':>5} | "
            f"{'DSII':>6} | {'Score':>5} | {'RkPA':>4} | "
            f"{'dMilk':>6} | {'dFat':>5} | {'dPro':>4} | "
            f"{'dDPR':>5} | {'dUDC':>5} | {'dSCS':>5} | "
            f"{'Antag':>5} | {'Crit':>5} | {'DstId':>5} | "
            f"{'Caseina':>13}"
        )
        print(hdr)
        print("-" * 130)

        for p in preds_s:
            diff_rank = p.rank_simple_pa - p.rank
            print(
                f"{preds_s.index(p)+1:>3} | "
                f"{p.dam.name:<22.22} | "
                f"${p.dam.nm_dollar:>4.0f} | "
                f"${p.nm_pa:>5.0f} | "
                f"${p.nm_corrected:>4.0f} | "
                f"{p.dsii:>6.0f} | "
                f"{p.final_score:>5.1f} | "
                f"{diff_rank:>+4d} | "
                f"{p.delta_g.get('MILK',0):>+6.0f} | "
                f"{p.delta_g.get('FAT',0):>+5.1f} | "
                f"{p.delta_g.get('PROT',0):>+4.0f} | "
                f"{p.delta_g.get('DPR',0):>+5.1f} | "
                f"{p.delta_g.get('UDC',0):>+5.2f} | "
                f"{p.delta_g.get('SCS',0):>+5.2f} | "
                f"{p.antagonism_penalty:>5.1f} | "
                f"{p.critical_penalty:>5.1f} | "
                f"{p.distance_to_ideal:>5.2f} | "
                f"{p.beta_info[:13]:>13}"
            )

        scores = [p.final_score for p in preds_s]
        nms = [p.nm_corrected for p in preds_s]
        dsiis = [p.dsii for p in preds_s]
        print("-" * 130)
        print(f"  Resumo {s.name}: Score min={min(scores):.1f} max={max(scores):.1f} "
              f"med={sum(scores)/len(scores):.1f}  |  "
              f"NM$ min=${min(nms):.0f} max=${max(nms):.0f} med=${sum(nms)/len(nms):.0f}  |  "
              f"DSII min={min(dsiis):.0f} max={max(dsiis):.0f}")

    # =========================================================================
    # ANALISE DE DIVERSIFICACAO
    # =========================================================================
    print(f"\n{'='*130}")
    print("  ANALISE DE DIVERSIFICACAO (evitar concentracao genetica futura)")
    print(f"{'='*130}")

    for sire_name, dam_group in sorted(dam_sire_groups.items(), key=lambda x: -len(x[1])):
        if len(dam_group) < 3:
            continue
        print(f"\n  Grupo {sire_name} ({len(dam_group)} filhas):")

        # Conta qual touro foi escolhido para cada filha
        bull_usage = {}
        for d in dam_group:
            ranked = all_per_dam[d.id]
            best = ranked[0].sire.name
            bull_usage[best] = bull_usage.get(best, 0) + 1

        for bull, count in sorted(bull_usage.items(), key=lambda x: -x[1]):
            pct = count / len(dam_group) * 100
            warn = " << CONCENTRACAO!" if pct > 60 else ""
            print(f"    {bull}: {count}/{len(dam_group)} filhas ({pct:.0f}%){warn}")

        if len(bull_usage) < 3:
            print(f"    ALERTA: Pouca diversidade de touros neste grupo!")
            print(f"    Recomendacao: Variar touros para evitar F elevado na proxima geracao")

    # =========================================================================
    # EXPORT CSV
    # =========================================================================
    csv_path = os.path.join(OUTPUT_DIR, "predicao_v3_dsii.csv")
    fields = [
        "Rank_v3", "Rank_PA_Simples", "Diferenca_Ranking",
        "Touro_NAAB", "Touro_Nome", "Touro_TPI", "Touro_NM$",
        "Femea_NAAB", "Femea_Num", "Femea_Reg", "Femea_TPI", "Femea_NM$",
        "NM$_PA", "NM$_Corrigido", "NM$_IC_Lower", "NM$_IC_Upper",
        "Delta_NM$_sobre_Mae",
        "TPI_PA", "TPI_Corrigido", "TPI_IC_Lower", "TPI_IC_Upper",
        "Delta_TPI_sobre_Mae",
        "DSII", "Score_Final_v3",
        "F_Esperada_%", "Depressao_$",
        "Penalidade_Critica", "Penalidade_Antagonismo",
        "Bonus_Dominancia", "Distancia_Ideal",
        "P_Acima_Media_%", "P_Top25_%", "P_Top10_%",
        "Haplo_Conflitos", "Beta_Caseina", "Kappa_Caseina",
        "Pai_da_Vaca",
    ]

    all_traits = ["MILK","FAT","PROT","SCS","DPR","CCR","PL","LIV",
                   "HCR","EFC","HLIV","UDC","FLC","BWC",
                   "MF","DA","KET","MAST","MET","RP","SCE",
                   "STA","STR","BD","DF","RA","TW","RLS","RLR","FA","FLS",
                   "FUA","RUH","RUW","UC","UD","FTP","RTP","TL"]
    for t in all_traits:
        fields.extend([f"PA_{t}", f"DeltaG_{t}"])

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fields, delimiter=";")
        w.writeheader()
        for p in predictions:
            row = {
                "Rank_v3": p.rank,
                "Rank_PA_Simples": p.rank_simple_pa,
                "Diferenca_Ranking": p.rank_simple_pa - p.rank,
                "Touro_NAAB": p.sire.naab, "Touro_Nome": p.sire.name,
                "Touro_TPI": f"{p.sire.tpi:.0f}", "Touro_NM$": f"{p.sire.nm_dollar:.0f}",
                "Femea_NAAB": p.dam.naab, "Femea_Num": p.dam.num,
                "Femea_Reg": p.dam.reg, "Femea_TPI": f"{p.dam.tpi:.0f}",
                "Femea_NM$": f"{p.dam.nm_dollar:.0f}",
                "NM$_PA": f"{p.nm_pa:.0f}",
                "NM$_Corrigido": f"{p.nm_corrected:.0f}",
                "NM$_IC_Lower": f"{p.nm_ic_lower:.0f}",
                "NM$_IC_Upper": f"{p.nm_ic_upper:.0f}",
                "Delta_NM$_sobre_Mae": f"{p.delta_nm:.0f}",
                "TPI_PA": f"{p.tpi_pa:.0f}",
                "TPI_Corrigido": f"{p.tpi_corrected:.0f}",
                "TPI_IC_Lower": f"{p.tpi_ic_lower:.0f}",
                "TPI_IC_Upper": f"{p.tpi_ic_upper:.0f}",
                "Delta_TPI_sobre_Mae": f"{p.delta_tpi:.0f}",
                "DSII": f"{p.dsii:.1f}",
                "Score_Final_v3": f"{p.final_score:.1f}",
                "F_Esperada_%": f"{p.expected_f*100:.2f}",
                "Depressao_$": f"{p.endo_depression:.0f}",
                "Penalidade_Critica": f"{p.critical_penalty:.1f}",
                "Penalidade_Antagonismo": f"{p.antagonism_penalty:.1f}",
                "Bonus_Dominancia": f"{p.dominance_bonus:.2f}",
                "Distancia_Ideal": f"{p.distance_to_ideal:.2f}",
                "P_Acima_Media_%": f"{p.prob_above_avg*100:.1f}",
                "P_Top25_%": f"{p.prob_top25*100:.1f}",
                "P_Top10_%": f"{p.prob_top10*100:.1f}",
                "Haplo_Conflitos": ",".join(p.haplo_conflicts) if p.haplo_conflicts else "",
                "Beta_Caseina": p.beta_info,
                "Kappa_Caseina": p.kappa_info,
                "Pai_da_Vaca": p.dam.sire_name,
            }
            for t in all_traits:
                row[f"PA_{t}"] = f"{p.trait_pred.get(t,0):.2f}"
                row[f"DeltaG_{t}"] = f"{p.delta_g.get(t,0):.2f}"
            w.writerow(row)

    print(f"\n[OK] CSV v3 exportado: {csv_path}")

    # =========================================================================
    # RESUMO FINAL
    # =========================================================================
    print(f"\n{'='*130}")
    print("  RESUMO v3")
    print(f"{'='*130}")
    scores = [p.final_score for p in predictions]
    nms = [p.nm_corrected for p in predictions]

    # Quantos rankings mudaram
    rank_changes = sum(1 for p in predictions if abs(p.rank - p.rank_simple_pa) >= 5)
    big_changes = sum(1 for p in predictions if abs(p.rank - p.rank_simple_pa) >= 20)

    print(f"  Combinacoes: {len(predictions)}")
    print(f"  Score v3: min={min(scores):.1f}  max={max(scores):.1f}  media={sum(scores)/len(scores):.1f}")
    print(f"  NM$ corrigido: min=${min(nms):.0f}  max=${max(nms):.0f}  media=${sum(nms)/len(nms):.0f}")
    print(f"\n  Rankings que mudaram >= 5 posicoes: {rank_changes} ({rank_changes/len(predictions)*100:.0f}%)")
    print(f"  Rankings que mudaram >= 20 posicoes: {big_changes} ({big_changes/len(predictions)*100:.0f}%)")

    print(f"\n  Metodologia v3:")
    print(f"    - DSII: Indice de Melhoria Especifico por Femea (pesos adaptativos)")
    print(f"    - Multi-trait: 30+ traits (producao + saude + fertilidade + tipo + eficiencia)")
    print(f"    - Penalidade de limites criticos (DPR, PL, SCS, UDC, FLC, SCE)")
    print(f"    - Antagonismo producao x funcionalidade (correlacoes geneticas)")
    print(f"    - Desvio de dominancia (heterozigosidade)")
    print(f"    - Distancia ao perfil ideal")
    print(f"    - Analise de diversificacao por grupo de pai-da-vaca")


if __name__ == "__main__":
    main()
