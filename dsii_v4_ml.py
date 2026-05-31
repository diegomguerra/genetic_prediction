"""
DSII v4 — Machine Learning Genetic Prediction Engine
=====================================================
Trains ML models on real genomic data (genotyped daughters with known pedigree)
to learn calibration factors beyond simple Parent Average.

Data sources:
  - Filhas: May 2026.xlsx (1,974 genotyped daughters — gold standard)
  - Pais:   bulls.csv (40k sires)
  - Maes:   DAM1-DAM20.xlsx (1,301 unique dams)

Linkage:
  - Filha -> Pai:  SIRENAME (May 2026) -> Registration Name (bulls.csv)  [99.8%]
  - Filha -> Mae:  DAMREGNUM2 (May 2026) -> Naab Code (DAM1-20)          [86.7%]
"""

import warnings
warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd
import openpyxl
import csv
import json
from pathlib import Path
from dataclasses import dataclass
from typing import Optional

from sklearn.model_selection import KFold, cross_val_score
from sklearn.linear_model import Ridge, Lasso, ElasticNet
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.preprocessing import StandardScaler

try:
    from xgboost import XGBRegressor
    HAS_XGB = True
except ImportError:
    HAS_XGB = False
    print("XGBoost nao disponivel, usando GradientBoosting como substituto.")

try:
    from lightgbm import LGBMRegressor
    HAS_LGBM = True
except ImportError:
    HAS_LGBM = False
    print("LightGBM nao disponivel.")

# ============================================================
# PATHS
# ============================================================
DOWNLOADS = Path("C:/Users/DiegoGuerra/Downloads")
BULLS_CSV = Path("C:/Users/DiegoGuerra/gen-genie-advisor/sms-engine/data/bulls.csv")
OUTPUT_DIR = Path("C:/Users/DiegoGuerra/Documents/Projetos/genetic_prediction/dsii_v4_results")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ============================================================
# TRAIT MAPPING across 3 data sources
# ============================================================
# Each tuple: (trait_name, daughter_col, dam_col, bull_col)
# daughter_col = column header in May 2026
# dam_col = column header in DAM files
# bull_col = column header in bulls.csv

TRAIT_MAP = [
    # Indices economicos
    ('TPI',    'GTPI',       'TPI',          'TPI'),
    ('NM$',    'NM$',        'Net Merit',    'NM$'),
    ('CM$',    'CM$',        'CM$',          'CM$'),
    ('HHP$',   'HHP$',       None,           'HHP$\u00ae'),  # DAM nao tem HHP$

    # Producao
    ('MILK',   'MILK',       'PTA Milk',     'PTAM'),
    ('FAT',    'FAT',        'PTA Fat',      'PTAF'),
    ('FAT%',   '%F',         '% Fat',        'PTAF%'),
    ('PRO',    'PRO',        'PTA Pro',      'PTAP'),
    ('PRO%',   '%P',         '% Pro',        'PTAP%'),
    ('CFP',    'CFP',        'CFP',          'CFP'),

    # Saude e longevidade
    ('PL',     'PL',         'PL',           'PL'),
    ('SCS',    'SCS',        'SCS',          'SCS'),
    ('DPR',    'DPR',        'PTA DPR',      'DPR'),
    ('LIV',    'LIV',        'PTA LIV',      'LIV'),
    ('FI',     'FI',         'Fertil Index',  'FI'),
    ('HCR',    'HCR',        'HCR',          'HCR'),
    ('CCR',    'CCR',        'CCR',          'CCR'),
    ('MAST',   'MAST',       'Mastitis',     'MAST'),
    ('FSAV',   'FSAV',       'Feed Saved',   'F SAV'),

    # Tipo
    ('PTAT',   'PTAT',       'PTA Type',     'PTAT'),
    ('UDC',    'UDC',        'UDC',          'UDC'),
    ('FLC',    'FLC',        'FLC',          'FLC'),

    # Parto
    ('SCE',    'SCE',        'SCE',          'SCE'),
    ('DCE',    'DCE',        'DCE',          'DCE'),
    ('SSB',    'SSB',        'SSB',          'SSB'),
    ('DSB',    'DSB',        'DSB',          'DSB'),

    # GFI
    ('GFI',    'GFI',        None,           'GFI'),
]

# Traits we'll actually model (need all 3 sources)
MODELABLE_TRAITS = [t for t in TRAIT_MAP if t[2] is not None]

# Key traits for summary comparison
KEY_TRAITS = ['TPI', 'NM$', 'MILK', 'FAT', 'PRO', 'PL', 'SCS', 'DPR', 'PTAT', 'UDC', 'FLC']


def safe_float(val) -> Optional[float]:
    if val is None:
        return None
    try:
        v = float(val)
        return v if not np.isnan(v) else None
    except (ValueError, TypeError):
        return None


# ============================================================
# DATA LOADING
# ============================================================

def load_daughters() -> pd.DataFrame:
    """Load May 2026 genotyped daughters."""
    print("Carregando filhas (May 2026)...")
    wb = openpyxl.load_workbook(DOWNLOADS / "May 2026.xlsx", read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    hdr = [str(c).strip() if c else f'col_{i}' for i, c in enumerate(rows[0])]
    data = []
    for r in rows[1:]:
        row_dict = {}
        for i, h in enumerate(hdr):
            row_dict[h] = r[i] if i < len(r) else None
        data.append(row_dict)

    df = pd.DataFrame(data)
    print(f"  {len(df)} filhas carregadas")
    return df


def load_dams() -> dict:
    """Load DAM1-20 files, indexed by Naab Code."""
    print("Carregando maes (DAM1-20)...")
    dams = {}

    for i in range(1, 21):
        f = DOWNLOADS / f"DAM{i}.xlsx"
        if not f.exists():
            continue
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        hdr = [str(c).strip() if c else f'col_{j}' for j, c in enumerate(rows[0])]
        naab_idx = next(j for j, h in enumerate(hdr) if 'naab' in h.lower() and 'code' in h.lower())

        for r in rows[1:]:
            naab = r[naab_idx]
            if not naab:
                continue
            naab = str(naab).strip()
            row_dict = {}
            for j, h in enumerate(hdr):
                row_dict[h] = r[j] if j < len(r) else None
            dams[naab] = row_dict

    print(f"  {len(dams)} maes unicas carregadas")
    return dams


def load_bulls() -> dict:
    """Load bulls.csv indexed by Registration Name."""
    print("Carregando pais (bulls.csv)...")
    bulls_by_regname = {}
    bulls_by_name = {}

    with open(BULLS_CSV, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            regname = row.get('Registration Name', '').strip().upper()
            name = row.get('Name', '').strip().upper()
            if regname:
                bulls_by_regname[regname] = row
            if name:
                bulls_by_name[name] = row

    print(f"  {len(bulls_by_regname)} pais por Registration Name")
    return bulls_by_regname, bulls_by_name


# ============================================================
# TRIO ASSEMBLY
# ============================================================

def assemble_trios(daughters_df, dams_dict, bulls_regname, bulls_name):
    """Build feature matrix: sire PTAs + dam PTAs -> daughter genomic values."""
    print("\nMontando trios (filha + pai + mae)...")

    records = []
    stats = {'total': 0, 'no_sire': 0, 'no_dam': 0, 'complete': 0}

    for _, daughter in daughters_df.iterrows():
        stats['total'] += 1

        # Find sire
        sirename = str(daughter.get('SIRENAME', '')).strip().upper()
        sire = bulls_regname.get(sirename) or bulls_name.get(sirename)
        if not sire:
            stats['no_sire'] += 1
            continue

        # Find dam
        damreg = str(daughter.get('DAMREGNUM2', '')).strip()
        dam = dams_dict.get(damreg)
        if not dam:
            stats['no_dam'] += 1
            continue

        stats['complete'] += 1

        # Extract trait values for all 3
        rec = {'sirename': sirename, 'damreg': damreg}

        for trait_name, daughter_col, dam_col, bull_col in MODELABLE_TRAITS:
            d_val = safe_float(daughter.get(daughter_col))
            s_val = safe_float(sire.get(bull_col))
            m_val = safe_float(dam.get(dam_col))

            rec[f'd_{trait_name}'] = d_val   # daughter (target)
            rec[f's_{trait_name}'] = s_val   # sire (feature)
            rec[f'm_{trait_name}'] = m_val   # dam (feature)

            # Parent Average
            if s_val is not None and m_val is not None:
                rec[f'pa_{trait_name}'] = (s_val + m_val) / 2
            else:
                rec[f'pa_{trait_name}'] = None

        # Also extract traits only in sire+daughter (HHP$, GFI)
        for trait_name, daughter_col, dam_col, bull_col in TRAIT_MAP:
            if dam_col is not None:
                continue  # already handled
            d_val = safe_float(daughter.get(daughter_col))
            s_val = safe_float(sire.get(bull_col))
            rec[f'd_{trait_name}'] = d_val
            rec[f's_{trait_name}'] = s_val

        records.append(rec)

    print(f"  Total filhas: {stats['total']}")
    print(f"  Sem pai: {stats['no_sire']}")
    print(f"  Sem mae: {stats['no_dam']}")
    print(f"  Trios completos: {stats['complete']}")

    return pd.DataFrame(records)


# ============================================================
# ML TRAINING
# ============================================================

def train_single_trait(trios_df, trait_name, feature_cols=None):
    """Train ML models for a single trait. Returns results dict."""
    target_col = f'd_{trait_name}'
    sire_col = f's_{trait_name}'
    dam_col = f'm_{trait_name}'
    pa_col = f'pa_{trait_name}'

    # Filter rows with complete data
    needed = [target_col, sire_col, dam_col, pa_col]
    mask = trios_df[needed].notna().all(axis=1)
    df = trios_df[mask].copy()

    if len(df) < 50:
        return None

    y = df[target_col].values

    # --- Baseline: Parent Average ---
    pa_pred = df[pa_col].values
    pa_mae = mean_absolute_error(y, pa_pred)
    pa_rmse = np.sqrt(mean_squared_error(y, pa_pred))
    pa_r2 = r2_score(y, pa_pred)

    # --- Feature matrix for ML ---
    # Core features: sire PTA, dam PTA
    if feature_cols is None:
        # Use all available sire and dam PTAs as features (cross-trait learning)
        s_cols = [c for c in trios_df.columns if c.startswith('s_') and c != sire_col]
        m_cols = [c for c in trios_df.columns if c.startswith('m_') and c != dam_col]
        feature_cols_full = [sire_col, dam_col] + s_cols + m_cols
    else:
        feature_cols_full = feature_cols

    # Only keep columns with enough data
    good_cols = []
    for c in feature_cols_full:
        if c in df.columns and df[c].notna().sum() > len(df) * 0.5:
            good_cols.append(c)

    X = df[good_cols].copy()
    # Fill remaining NaN with column median
    for c in good_cols:
        if X[c].isna().any():
            X[c] = X[c].fillna(X[c].median())

    X = X.values

    # --- Models ---
    results = {
        'trait': trait_name,
        'n_samples': len(df),
        'n_features': len(good_cols),
        'PA': {'MAE': pa_mae, 'RMSE': pa_rmse, 'R2': pa_r2},
    }

    kf = KFold(n_splits=5, shuffle=True, random_state=42)

    models = {
        'Ridge': Ridge(alpha=1.0),
        'RF': RandomForestRegressor(n_estimators=200, max_depth=8, min_samples_leaf=5, random_state=42, n_jobs=-1),
    }

    if HAS_XGB:
        models['XGBoost'] = XGBRegressor(
            n_estimators=300, max_depth=5, learning_rate=0.05,
            subsample=0.8, colsample_bytree=0.8,
            min_child_weight=5, random_state=42, verbosity=0
        )

    if HAS_LGBM:
        models['LightGBM'] = LGBMRegressor(
            n_estimators=300, max_depth=5, learning_rate=0.05,
            subsample=0.8, colsample_bytree=0.8,
            min_child_samples=5, random_state=42, verbose=-1
        )

    models['GBR'] = GradientBoostingRegressor(
        n_estimators=200, max_depth=5, learning_rate=0.05,
        min_samples_leaf=5, random_state=42
    )

    best_model_name = 'PA'
    best_mae = pa_mae

    for name, model in models.items():
        try:
            mae_scores = -cross_val_score(model, X, y, cv=kf, scoring='neg_mean_absolute_error')
            rmse_scores = np.sqrt(-cross_val_score(model, X, y, cv=kf, scoring='neg_mean_squared_error'))
            r2_scores = cross_val_score(model, X, y, cv=kf, scoring='r2')

            avg_mae = mae_scores.mean()
            avg_rmse = rmse_scores.mean()
            avg_r2 = r2_scores.mean()

            results[name] = {
                'MAE': avg_mae,
                'RMSE': avg_rmse,
                'R2': avg_r2,
                'MAE_std': mae_scores.std(),
            }

            if avg_mae < best_mae:
                best_mae = avg_mae
                best_model_name = name

        except Exception as e:
            results[name] = {'error': str(e)}

    results['best_model'] = best_model_name
    results['improvement_vs_PA'] = (
        (pa_mae - best_mae) / pa_mae * 100 if best_model_name != 'PA' else 0
    )

    return results


def train_single_trait_simple(trios_df, trait_name):
    """Train with only sire+dam PTAs for that trait (2 features)."""
    target_col = f'd_{trait_name}'
    sire_col = f's_{trait_name}'
    dam_col = f'm_{trait_name}'
    pa_col = f'pa_{trait_name}'

    needed = [target_col, sire_col, dam_col, pa_col]
    mask = trios_df[needed].notna().all(axis=1)
    df = trios_df[mask].copy()

    if len(df) < 50:
        return None

    y = df[target_col].values
    pa_pred = df[pa_col].values
    pa_mae = mean_absolute_error(y, pa_pred)
    pa_r2 = r2_score(y, pa_pred)

    X = df[[sire_col, dam_col]].values

    # Weighted PA: learn optimal weights via Ridge
    ridge = Ridge(alpha=0.1)
    kf = KFold(n_splits=5, shuffle=True, random_state=42)
    mae_scores = -cross_val_score(ridge, X, y, cv=kf, scoring='neg_mean_absolute_error')
    r2_scores = cross_val_score(ridge, X, y, cv=kf, scoring='r2')

    # Fit on full data to get weights
    ridge.fit(X, y)
    sire_w = ridge.coef_[0]
    dam_w = ridge.coef_[1]
    intercept = ridge.intercept_

    return {
        'trait': trait_name,
        'n': len(df),
        'sire_weight': round(sire_w, 4),
        'dam_weight': round(dam_w, 4),
        'intercept': round(intercept, 4),
        'PA_MAE': round(pa_mae, 2),
        'PA_R2': round(pa_r2, 4),
        'Ridge_MAE': round(mae_scores.mean(), 2),
        'Ridge_R2': round(r2_scores.mean(), 4),
    }


# ============================================================
# DSII v3 SIMULATION (simplified for comparison)
# ============================================================

DSII_V3_WEIGHTS = {
    'sire': 0.50,
    'dam': 0.50,
}


def compute_dsii_v3_pa(sire_val, dam_val):
    """DSII v3 baseline: simple PA (before deficiency adjustments)."""
    if sire_val is None or dam_val is None:
        return None
    return sire_val * 0.5 + dam_val * 0.5


# ============================================================
# MAIN PIPELINE
# ============================================================

def run_pipeline():
    print("=" * 70)
    print("  DSII v4 — Machine Learning Genetic Prediction Pipeline")
    print("=" * 70)

    # 1. Load data
    daughters_df = load_daughters()
    dams_dict = load_dams()
    bulls_regname, bulls_name = load_bulls()

    # 2. Assemble trios
    trios = assemble_trios(daughters_df, dams_dict, bulls_regname, bulls_name)
    trios.to_csv(OUTPUT_DIR / "trios_assembled.csv", index=False)
    print(f"\nTrios salvos: {OUTPUT_DIR / 'trios_assembled.csv'}")

    # 3. Optimal weights analysis (sire vs dam)
    print("\n" + "=" * 70)
    print("  FASE 1: Pesos Otimos (Sire vs Dam)")
    print("=" * 70)

    weight_results = []
    for trait_name, _, dam_col, _ in MODELABLE_TRAITS:
        if dam_col is None:
            continue
        res = train_single_trait_simple(trios, trait_name)
        if res:
            weight_results.append(res)
            print(f"  {trait_name:>6}: Sire={res['sire_weight']:.3f}  Dam={res['dam_weight']:.3f}  "
                  f"Intercept={res['intercept']:.1f}  "
                  f"PA_MAE={res['PA_MAE']:.1f}  Ridge_MAE={res['Ridge_MAE']:.1f}  "
                  f"PA_R2={res['PA_R2']:.3f}  Ridge_R2={res['Ridge_R2']:.3f}")

    weights_df = pd.DataFrame(weight_results)
    weights_df.to_csv(OUTPUT_DIR / "optimal_weights.csv", index=False)

    # 4. Full ML training per trait
    print("\n" + "=" * 70)
    print("  FASE 2: Treinamento ML Multi-Feature")
    print("=" * 70)

    all_results = []
    for trait_name, _, dam_col, _ in MODELABLE_TRAITS:
        if dam_col is None:
            continue
        print(f"\n--- {trait_name} ---")
        res = train_single_trait(trios, trait_name)
        if res:
            all_results.append(res)
            print(f"  N={res['n_samples']}, Features={res['n_features']}")
            print(f"  PA:       MAE={res['PA']['MAE']:.2f}  R2={res['PA']['R2']:.4f}")
            for model_name in ['Ridge', 'RF', 'XGBoost', 'LightGBM', 'GBR']:
                if model_name in res and 'MAE' in res[model_name]:
                    m = res[model_name]
                    print(f"  {model_name:>8}: MAE={m['MAE']:.2f}  R2={m['R2']:.4f}")
            print(f"  MELHOR: {res['best_model']} ({res['improvement_vs_PA']:.1f}% melhor que PA)")

    # 5. Summary comparison table
    print("\n" + "=" * 70)
    print("  RESUMO: PA vs ML (Melhor Modelo)")
    print("=" * 70)
    print(f"{'Trait':>8} | {'N':>5} | {'PA MAE':>8} | {'PA R2':>7} | {'ML MAE':>8} | {'ML R2':>7} | {'Best':>8} | {'Melhoria':>8}")
    print("-" * 80)

    summary_rows = []
    for res in all_results:
        trait = res['trait']
        best = res['best_model']
        pa = res['PA']

        if best != 'PA' and best in res and 'MAE' in res[best]:
            ml = res[best]
            improvement = res['improvement_vs_PA']
        else:
            ml = pa
            best = 'PA'
            improvement = 0

        print(f"{trait:>8} | {res['n_samples']:>5} | {pa['MAE']:>8.2f} | {pa['R2']:>7.4f} | "
              f"{ml['MAE']:>8.2f} | {ml['R2']:>7.4f} | {best:>8} | {improvement:>7.1f}%")

        summary_rows.append({
            'Trait': trait,
            'N': res['n_samples'],
            'PA_MAE': round(pa['MAE'], 2),
            'PA_R2': round(pa['R2'], 4),
            'ML_MAE': round(ml['MAE'], 2),
            'ML_R2': round(ml['R2'], 4),
            'Best_Model': best,
            'Improvement_%': round(improvement, 1),
        })

    summary_df = pd.DataFrame(summary_rows)
    summary_df.to_csv(OUTPUT_DIR / "dsii_v4_summary.csv", index=False)

    # 6. Save full results as JSON
    # Convert numpy types for JSON serialization
    def convert(obj):
        if isinstance(obj, (np.integer,)):
            return int(obj)
        if isinstance(obj, (np.floating,)):
            return float(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        return obj

    with open(OUTPUT_DIR / "dsii_v4_full_results.json", 'w') as f:
        json.dump(all_results, f, indent=2, default=convert)

    # 7. Final verdict
    print("\n" + "=" * 70)
    print("  VEREDICTO FINAL")
    print("=" * 70)

    key_improvements = [r for r in summary_rows if r['Trait'] in KEY_TRAITS]
    if key_improvements:
        avg_improvement = np.mean([r['Improvement_%'] for r in key_improvements])
        avg_pa_r2 = np.mean([r['PA_R2'] for r in key_improvements])
        avg_ml_r2 = np.mean([r['ML_R2'] for r in key_improvements])

        print(f"  Traits-chave analisados: {len(key_improvements)}")
        print(f"  R2 medio PA:  {avg_pa_r2:.4f}")
        print(f"  R2 medio ML:  {avg_ml_r2:.4f}")
        print(f"  Melhoria media MAE: {avg_improvement:.1f}%")
        print()

        if avg_improvement > 5:
            print("  >> DSII v4 (ML) SUPERA Parent Average de forma significativa.")
            print("  >> Recomendacao: Implementar modelo ML calibrado na producao.")
        elif avg_improvement > 0:
            print("  >> DSII v4 (ML) oferece ganho marginal sobre Parent Average.")
            print("  >> Os pesos otimos Sire/Dam (Fase 1) podem ser o melhor custo-beneficio.")
        else:
            print("  >> Parent Average ja e um bom preditor. ML nao agrega valor significativo.")
            print("  >> Foco deve ser em dados adicionais (genomica direta, mais filhas).")

    print(f"\nResultados salvos em: {OUTPUT_DIR}")
    print("Arquivos gerados:")
    print("  - trios_assembled.csv")
    print("  - optimal_weights.csv")
    print("  - dsii_v4_summary.csv")
    print("  - dsii_v4_full_results.json")

    return all_results, summary_df, weights_df


if __name__ == '__main__':
    run_pipeline()
