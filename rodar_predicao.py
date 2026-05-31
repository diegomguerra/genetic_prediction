"""
================================================================================
EXECUÇÃO DA PREDIÇÃO GENÔMICA — DADOS REAIS
Carrega Lista de touros.xlsx e Lista de fêmeas.xlsx
================================================================================
"""

import os
import sys
import openpyxl
from predicao_genomica_holstein import (
    Animal, GenomicPredictionEngine, TRAITS_CONFIG
)

DOWNLOADS = os.path.expanduser("~/Downloads")
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))


# =============================================================================
# MAPEAMENTO DE COLUNAS DAS PLANILHAS PARA TRAITS DO MOTOR
# =============================================================================

# Confiabilidades genômicas padrão para Holstein (quando não disponíveis na planilha)
# Touros genômicos jovens: ~70-80% para produção, ~55-65% para fertilidade
# Fêmeas genômicas: ~50-65% para produção, ~35-50% para fertilidade

DEFAULT_REL_SIRE = {
    "MILK": 0.78, "FAT": 0.78, "PROT": 0.78,
    "SCS": 0.72, "PL": 0.68, "LIV": 0.58, "HLIV": 0.52,
    "DPR": 0.62, "CCR": 0.58, "HCR": 0.52, "EFC": 0.58,
    "RFI": 0.55, "BWC": 0.72, "UDC": 0.75, "FLC": 0.68,
    "CA": 0.58, "HTH": 0.55, "FS": 0.55,
}

DEFAULT_REL_DAM = {
    "MILK": 0.58, "FAT": 0.58, "PROT": 0.58,
    "SCS": 0.52, "PL": 0.48, "LIV": 0.38, "HLIV": 0.32,
    "DPR": 0.42, "CCR": 0.38, "HCR": 0.32, "EFC": 0.38,
    "RFI": 0.35, "BWC": 0.52, "UDC": 0.55, "FLC": 0.48,
    "CA": 0.38, "HTH": 0.35, "FS": 0.35,
}

# Endogamia media estimada por parentesco de pai (sire)
# Se pai é muito usado na raça, endogamia tende a ser maior
SIRE_INBREEDING_ESTIMATES = {
    # Pais populares → endogamia estimada mais alta
    "Sheepster": 0.09, "Rimbot": 0.10, "Felix": 0.08,
    "Easton": 0.09, "Einstein": 0.08,
    # Pais das fêmeas
    "Daniels": 0.08, "Winterfell": 0.09, "Rubygold": 0.07,
    "Mosaic": 0.10, "Fulham": 0.07, "Prophecy": 0.08,
    "Touchdown": 0.08, "TRilogy": 0.07, "BLack Jack": 0.08,
    "Altabluedevil": 0.08, "Honda": 0.09, "Sherpa": 0.08,
    "Frenzy": 0.07,
}
DEFAULT_INBREEDING = 0.085  # Média Holstein atual


def safe_float(val, default=0.0):
    """Converte valor para float com fallback."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def detect_shared_ancestry(sire, dam):
    """
    Detecta ancestralidade compartilhada entre touro e vaca.

    Retorna coeficiente de parentesco estimado (a_SD).
    Quanto maior, maior a endogamia esperada da cria.
    """
    relationship = 0.0

    # Se compartilham o mesmo pai → meio-irmãos paternos
    if sire.sire_id and dam.sire_id:
        if sire.sire_id == dam.sire_id:
            relationship += 0.25  # Meio-irmãos paternos

    # Se pai do touro é avô materno da vaca (ou vice-versa)
    if sire.sire_id and hasattr(dam, 'mgs_id') and dam.mgs_id:
        if sire.sire_id == dam.mgs_id:
            relationship += 0.125

    # Relação base da raça Holstein (background relatedness)
    # Holstein atual: ~6.25% de parentesco médio entre animais não aparentados
    background = 0.0625

    return max(relationship, background)


def load_sires(filepath):
    """Carrega touros da planilha."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    sires = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        naab = row[1].value  # B
        if not naab:
            continue

        name = row[2].value or "?"       # C
        reg_name = row[3].value or name   # D
        sire_of = row[68].value           # BQ - Sire
        mgs = row[69].value              # BR - MGS

        # Haplótipos letais
        haplo = row[4].value or ""        # E

        # Beta/Kappa caseína
        beta = row[5].value or ""         # F
        kappa = row[6].value or ""        # G

        ptas = {
            "MILK": safe_float(row[18].value),   # S
            "FAT": safe_float(row[19].value),    # T
            "PROT": safe_float(row[21].value),   # V
            "SCS": safe_float(row[24].value),    # Y
            "DPR": safe_float(row[25].value),    # Z
            "CCR": safe_float(row[26].value),    # AA
            "PL": safe_float(row[27].value),     # AB
            "LIV": safe_float(row[28].value),    # AC
            "HCR": safe_float(row[36].value),    # AK
            "EFC": safe_float(row[39].value),    # AN
            "HLIV": safe_float(row[40].value),   # AO
            "RFI": safe_float(row[41].value),    # AP
            "UDC": safe_float(row[44].value),    # AS
            "FLC": safe_float(row[45].value),    # AT
            "BWC": safe_float(row[46].value),    # AU
            "CA": safe_float(row[65].value),     # BN (SCE como proxy de CA)
            "HTH": safe_float(row[37].value, 0), # AL (Health Index)
        }

        # Feed Saved (usado para calcular NM$ completo)
        feed_saved = safe_float(row[42].value)  # AQ

        # Dados complementares
        tpi = safe_float(row[13].value)   # N
        nm_published = safe_float(row[14].value)  # O

        # Estima endogamia baseado no pai
        f_est = SIRE_INBREEDING_ESTIMATES.get(sire_of, DEFAULT_INBREEDING)

        animal = Animal(
            id=naab,
            name=name,
            sex="M",
            ptas=ptas,
            reliabilities=DEFAULT_REL_SIRE.copy(),
            inbreeding_coef=f_est,
            sire_id=sire_of,
        )

        # Metadados extras
        animal.reg_name = reg_name
        animal.tpi = tpi
        animal.nm_published = nm_published
        animal.beta_casein = beta
        animal.kappa_casein = kappa
        animal.haplotypes = haplo
        animal.feed_saved = feed_saved
        animal.mgs_id = mgs

        sires.append(animal)

    return sires


def load_dams(filepath):
    """Carrega fêmeas da planilha."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    dams = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        naab = row[2].value  # C
        if not naab:
            continue

        num = row[0].value           # A
        reg = row[3].value or "?"    # D
        sire_of = row[66].value      # BO - Sire
        mgs = row[67].value          # BP - MGS

        haplo = row[4].value or ""   # E
        beta = row[5].value or ""    # F
        kappa = row[6].value or ""   # G

        ptas = {
            "MILK": safe_float(row[17].value),   # R
            "FAT": safe_float(row[18].value),    # S
            "PROT": safe_float(row[20].value),   # U
            "SCS": safe_float(row[23].value),    # X
            "DPR": safe_float(row[24].value),    # Y
            "CCR": safe_float(row[25].value),    # Z
            "PL": safe_float(row[26].value),     # AA
            "LIV": safe_float(row[27].value),    # AB
            "HCR": safe_float(row[34].value),    # AI
            "EFC": safe_float(row[37].value),    # AL
            "HLIV": safe_float(row[38].value),   # AM
            "RFI": safe_float(row[39].value),    # AN
            "UDC": safe_float(row[42].value),    # AQ
            "FLC": safe_float(row[43].value),    # AR
            "BWC": safe_float(row[44].value),    # AS
            "CA": safe_float(row[63].value),     # BL (SCE)
            "HTH": safe_float(row[35].value, 0), # AJ (Health Index)
        }

        feed_saved = safe_float(row[40].value)  # AO

        tpi = safe_float(row[14].value)   # O
        nm_published = safe_float(row[15].value)  # P

        f_est = SIRE_INBREEDING_ESTIMATES.get(sire_of, DEFAULT_INBREEDING)

        animal = Animal(
            id=naab,
            name=f"#{num} Reg.{reg}",
            sex="F",
            ptas=ptas,
            reliabilities=DEFAULT_REL_DAM.copy(),
            inbreeding_coef=f_est,
            sire_id=sire_of,
        )

        animal.reg = reg
        animal.num = num
        animal.tpi = tpi
        animal.nm_published = nm_published
        animal.beta_casein = beta
        animal.kappa_casein = kappa
        animal.haplotypes = haplo
        animal.feed_saved = feed_saved
        animal.mgs_id = mgs

        dams.append(animal)

    return dams


def check_haplotype_conflict(sire, dam):
    """
    Verifica conflito de haplótipos letais.

    Se ambos são portadores do mesmo haplótipo letal,
    25% dos embriões serão homozigotos → morte embrionária.

    Retorna: lista de haplótipos em conflito
    """
    sire_haplo = set(str(getattr(sire, 'haplotypes', '') or '').split())
    dam_haplo = set(str(getattr(dam, 'haplotypes', '') or '').split())

    conflicts = sire_haplo & dam_haplo
    # Remove strings vazias
    conflicts.discard('')

    return list(conflicts)


def check_casein_compatibility(sire, dam):
    """
    Verifica compatibilidade de caseínas.

    Beta-caseína A2A2 é preferível (melhor digestibilidade).
    Kappa-caseína BB é preferível (melhor rendimento queijeiro).

    Retorna: (beta_offspring_options, kappa_offspring_options, score)
    """
    score = 0

    # Beta-caseína
    sb = str(getattr(sire, 'beta_casein', '') or '')
    db = str(getattr(dam, 'beta_casein', '') or '')

    beta_info = f"{sb}×{db}"
    if sb == "A2A2" and db == "A2A2":
        score += 2  # 100% A2A2
        beta_info += " → 100% A2A2"
    elif "A2A2" in sb and "A1A2" in db:
        score += 1  # 50% A2A2, 50% A1A2
        beta_info += " → 50% A2A2"
    elif "A1A2" in sb and "A2A2" in db:
        score += 1
        beta_info += " → 50% A2A2"
    elif "A1A2" in sb and "A1A2" in db:
        score += 0  # 25% A2A2
        beta_info += " → 25% A2A2"
    elif "A1A1" in sb or "A1A1" in db:
        score -= 1
        beta_info += " → 0% A2A2"

    # Kappa-caseína
    sk = str(getattr(sire, 'kappa_casein', '') or '')
    dk = str(getattr(dam, 'kappa_casein', '') or '')

    kappa_info = f"{sk}×{dk}"
    if sk == "BB" and dk == "BB":
        score += 2
        kappa_info += " → 100% BB"
    elif "BB" in sk and "AB" in dk:
        score += 1
        kappa_info += " → 50% BB"
    elif "AB" in sk and "BB" in dk:
        score += 1
        kappa_info += " → 50% BB"

    return beta_info, kappa_info, score


# =============================================================================
# MAIN
# =============================================================================

def main():
    import sys
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    print("=" * 100)
    print("PREDICAO GENOMICA AVANCADA -- RACA HOLANDESA")
    print("Dados reais: Lista de touros.xlsx + Lista de fêmeas.xlsx")
    print("=" * 100)
    print()

    # Carrega dados
    sires_file = os.path.join(DOWNLOADS, "Lista de touros.xlsx")
    dams_file = os.path.join(DOWNLOADS, "Lista de fêmeas.xlsx")

    sires = load_sires(sires_file)
    dams = load_dams(dams_file)

    print(f"Touros carregados: {len(sires)}")
    for s in sires:
        print(f"  {s.name:<15} NAAB={s.id} TPI={s.tpi:.0f} NM$={s.nm_published:.0f} "
              f"Milk={s.ptas['MILK']:.0f} Fat={s.ptas['FAT']:.0f}")
    print()

    print(f"Fêmeas carregadas: {len(dams)}")
    for d in dams[:5]:
        print(f"  {d.name:<25} TPI={d.tpi:.0f} NM$={d.nm_published:.0f} "
              f"Milk={d.ptas['MILK']:.0f} Fat={d.ptas['FAT']:.0f}")
    if len(dams) > 5:
        print(f"  ... e mais {len(dams)-5} fêmeas")
    print()

    total = len(sires) * len(dams)
    print(f"Total de combinações: {total}")
    print()

    # Calcula relações de parentesco conhecidas
    relationships = {}
    for s in sires:
        for d in dams:
            a_sd = detect_shared_ancestry(s, d)
            relationships[(s.id, d.id)] = a_sd

    # Cria motor e roda predição
    engine = GenomicPredictionEngine()

    # Ajusta media e SD da população baseado nos NM$ publicados das fêmeas
    nm_values = [d.nm_published for d in dams if hasattr(d, 'nm_published')]
    if nm_values:
        engine.population_avg_nm = sum(nm_values) / len(nm_values)
        if len(nm_values) > 1:
            mean = engine.population_avg_nm
            engine.population_sd_nm = max(
                (sum((x - mean)**2 for x in nm_values) / (len(nm_values) - 1)) ** 0.5,
                100  # Mínimo SD
            )
    print(f"Média NM$ da população de fêmeas: ${engine.population_avg_nm:.0f}")
    print(f"Desvio padrão NM$: ${engine.population_sd_nm:.0f}")
    print()

    # Roda predição
    print("Calculando predições...")
    predictions = engine.predict_all_matings(sires, dams, relationships)

    # Adiciona informações extras (haplótipos, caseínas)
    for pred in predictions:
        pred.haplo_conflicts = check_haplotype_conflict(pred.sire, pred.dam)
        pred.beta_info, pred.kappa_info, pred.casein_score = check_casein_compatibility(
            pred.sire, pred.dam
        )

        # Penaliza score final se há conflito de haplótipos letais
        if pred.haplo_conflicts:
            penalty = len(pred.haplo_conflicts) * 10  # -10 pontos por haplótipo
            pred.final_score = max(0, pred.final_score - penalty)

        # Bonifica score por caseínas favoráveis
        pred.final_score += pred.casein_score * 1.5  # +1.5 pontos por ponto de caseína

    # Re-ordena e re-rankeia
    predictions.sort(key=lambda p: p.final_score, reverse=True)
    for i, pred in enumerate(predictions):
        pred.rank = i + 1

    # =========================================================================
    # RELATÓRIO PRINCIPAL
    # =========================================================================
    print()
    print("=" * 130)
    print("RANKING COMPLETO — MELHORES ACASALAMENTOS")
    print("=" * 130)

    header = (
        f"{'Rk':>3} | {'Touro':<12} | {'Fêmea':<22} | "
        f"{'NM$Pub':>6} | {'NM$Pred':>7} | {'NM$Corr':>7} | "
        f"{'F%':>5} | {'DepEnd$':>7} | "
        f"{'P>Med':>5} | {'PT25':>5} | "
        f"{'Comp':>4} | {'Score':>5} | "
        f"{'Haplo':>6} | {'BCas':>12}"
    )
    print(header)
    print("-" * 130)

    for pred in predictions[:50]:  # Top 50
        haplo_flag = ",".join(pred.haplo_conflicts) if pred.haplo_conflicts else "OK"
        beta_short = getattr(pred.sire, 'beta_casein', '?') or '?'

        # NM$ publicado do touro como referência
        nm_pub_sire = getattr(pred.sire, 'nm_published', 0)

        print(
            f"{pred.rank:>3} | "
            f"{pred.sire.name:<12.12} | "
            f"{pred.dam.name:<22.22} | "
            f"${nm_pub_sire:>5.0f} | "
            f"${pred.nm_dollar:>6.0f} | "
            f"${pred.nm_dollar_corrected:>6.0f} | "
            f"{pred.expected_inbreeding*100:>4.1f}% | "
            f"${pred.inbreeding_depression_dollars:>6.0f} | "
            f"{pred.prob_above_avg*100:>4.0f}% | "
            f"{pred.prob_top25*100:>4.0f}% | "
            f"{pred.complementarity_score:>4.0f} | "
            f"{pred.final_score:>5.1f} | "
            f"{haplo_flag:>6.6} | "
            f"{beta_short}"
        )

    # =========================================================================
    # MELHOR TOURO POR FÊMEA
    # =========================================================================
    print()
    print("=" * 130)
    print("MELHOR TOURO PARA CADA FÊMEA")
    print("=" * 130)

    # Agrupa por fêmea
    best_per_dam = {}
    for pred in predictions:
        dam_id = pred.dam.id
        if dam_id not in best_per_dam:
            best_per_dam[dam_id] = pred

    # Ordena por TPI da fêmea (decrescente)
    sorted_dams = sorted(best_per_dam.values(),
                          key=lambda p: getattr(p.dam, 'tpi', 0), reverse=True)

    print(f"{'Fêmea':<25} | {'TPI':>4} | {'Melhor Touro':<12} | "
          f"{'NM$Corr':>7} | {'F%':>5} | {'Score':>5} | "
          f"{'2º Melhor':<12} | {'Score2':>6} | {'Haplo':>8}")
    print("-" * 130)

    for pred in sorted_dams:
        dam_id = pred.dam.id
        # Encontra 2º melhor
        second = None
        for p in predictions:
            if p.dam.id == dam_id and p.sire.id != pred.sire.id:
                second = p
                break

        haplo_flag = ",".join(pred.haplo_conflicts) if pred.haplo_conflicts else "OK"

        print(
            f"{pred.dam.name:<25.25} | "
            f"{getattr(pred.dam, 'tpi', 0):>4.0f} | "
            f"{pred.sire.name:<12.12} | "
            f"${pred.nm_dollar_corrected:>6.0f} | "
            f"{pred.expected_inbreeding*100:>4.1f}% | "
            f"{pred.final_score:>5.1f} | "
            f"{(second.sire.name if second else '-'):<12.12} | "
            f"{(f'{second.final_score:.1f}' if second else '-'):>6} | "
            f"{haplo_flag:>8.8}"
        )

    # =========================================================================
    # DETALHAMENTO DOS TOP 5
    # =========================================================================
    print()
    print("=" * 130)
    print("DETALHAMENTO — TOP 5 MELHORES ACASALAMENTOS")
    print("=" * 130)

    for pred in predictions[:5]:
        print()
        print(f">>> #{pred.rank}: {pred.sire.name} ({pred.sire.id}) × "
              f"{pred.dam.name} ({pred.dam.id})")
        print(f"    Score Final: {pred.final_score:.1f}/100")
        print(f"    NM$ Predito: ${pred.nm_dollar:.0f} → Corrigido: ${pred.nm_dollar_corrected:.0f}")
        print(f"    Endogamia Esperada: {pred.expected_inbreeding*100:.2f}%")
        print(f"    Depressao Endogamica: -${abs(pred.inbreeding_depression_dollars):.0f}")
        print(f"    P(acima media): {pred.prob_above_avg*100:.1f}% | P(Top 25%): {pred.prob_top25*100:.1f}%")
        print(f"    Complementaridade: {pred.complementarity_score:.1f}")
        print(f"    Haplot. letais: {','.join(pred.haplo_conflicts) if pred.haplo_conflicts else 'Nenhum conflito'}")
        print(f"    Caseinas: Beta={pred.beta_info} | Kappa={pred.kappa_info}")
        print()

        # Tabela de traits
        print(f"    {'Trait':<28} | {'Touro':>8} | {'Vaca':>8} | {'Pred.':>8} | "
              f"{'IC 5%':>8} | {'IC 95%':>8} | {'Unid':<8}")
        print("    " + "-" * 90)

        for code in ['MILK', 'FAT', 'PROT', 'SCS', 'PL', 'LIV', 'DPR', 'CCR',
                      'HCR', 'EFC', 'HLIV', 'RFI', 'UDC', 'FLC', 'BWC']:
            tc = TRAITS_CONFIG.get(code)
            if not tc:
                continue
            pta_s = pred.sire.ptas.get(code, 0)
            pta_d = pred.dam.ptas.get(code, 0)
            pa = pred.parent_avg.get(code, 0)
            lo = pred.ci_lower.get(code, 0)
            hi = pred.ci_upper.get(code, 0)
            print(
                f"    {tc.description:<28} | "
                f"{pta_s:>8.1f} | {pta_d:>8.1f} | {pa:>8.1f} | "
                f"{lo:>8.1f} | {hi:>8.1f} | {tc.unit:<8}"
            )

    # =========================================================================
    # ALERTAS DE HAPLÓTIPOS
    # =========================================================================
    conflicts_found = [(p.rank, p.sire.name, p.dam.name, p.haplo_conflicts)
                        for p in predictions if p.haplo_conflicts]

    if conflicts_found:
        print()
        print("=" * 80)
        print("ALERTAS DE HAPLOTIPOS LETAIS")
        print("=" * 80)
        print("Combinacoes com risco de 25% de morte embrionaria:")
        print()
        for rank, sire, dam, haplos in conflicts_found[:30]:
            print(f"  Rank #{rank}: {sire} × {dam} → Conflito: {', '.join(haplos)}")
        if len(conflicts_found) > 30:
            print(f"  ... e mais {len(conflicts_found)-30} combinacoes com conflito")

    # =========================================================================
    # EXPORTA CSV COMPLETO
    # =========================================================================
    import csv

    csv_path = os.path.join(OUTPUT_DIR, "resultado_predicao_real.csv")
    fieldnames = [
        "Rank", "Touro_NAAB", "Touro_Nome", "Touro_TPI", "Touro_NM$",
        "Femea_NAAB", "Femea_Num", "Femea_Reg", "Femea_TPI", "Femea_NM$",
        "NM$_Predito", "NM$_Corrigido", "F_Esperada_%",
        "Depressao_Endo_$", "P_Acima_Media_%", "P_Top25_%",
        "Complementaridade", "Risk_Score", "Score_Final",
        "Haplo_Conflitos", "Beta_Caseina", "Kappa_Caseina",
    ]

    # Traits preditos
    trait_codes = ['MILK', 'FAT', 'PROT', 'SCS', 'PL', 'LIV', 'DPR', 'CCR',
                   'HCR', 'EFC', 'HLIV', 'RFI', 'UDC', 'FLC', 'BWC']
    for code in trait_codes:
        fieldnames.extend([f"PTA_{code}_Pred", f"IC_Low_{code}", f"IC_High_{code}"])

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()

        for pred in predictions:
            row = {
                "Rank": pred.rank,
                "Touro_NAAB": pred.sire.id,
                "Touro_Nome": pred.sire.name,
                "Touro_TPI": f"{getattr(pred.sire, 'tpi', 0):.0f}",
                "Touro_NM$": f"{getattr(pred.sire, 'nm_published', 0):.0f}",
                "Femea_NAAB": pred.dam.id,
                "Femea_Num": getattr(pred.dam, 'num', ''),
                "Femea_Reg": getattr(pred.dam, 'reg', ''),
                "Femea_TPI": f"{getattr(pred.dam, 'tpi', 0):.0f}",
                "Femea_NM$": f"{getattr(pred.dam, 'nm_published', 0):.0f}",
                "NM$_Predito": f"{pred.nm_dollar:.0f}",
                "NM$_Corrigido": f"{pred.nm_dollar_corrected:.0f}",
                "F_Esperada_%": f"{pred.expected_inbreeding*100:.2f}",
                "Depressao_Endo_$": f"{pred.inbreeding_depression_dollars:.0f}",
                "P_Acima_Media_%": f"{pred.prob_above_avg*100:.1f}",
                "P_Top25_%": f"{pred.prob_top25*100:.1f}",
                "Complementaridade": f"{pred.complementarity_score:.1f}",
                "Risk_Score": f"{pred.risk_score:.1f}",
                "Score_Final": f"{pred.final_score:.1f}",
                "Haplo_Conflitos": ",".join(pred.haplo_conflicts) if pred.haplo_conflicts else "",
                "Beta_Caseina": pred.beta_info,
                "Kappa_Caseina": pred.kappa_info,
            }

            for code in trait_codes:
                row[f"PTA_{code}_Pred"] = f"{pred.parent_avg.get(code, 0):.2f}"
                row[f"IC_Low_{code}"] = f"{pred.ci_lower.get(code, 0):.2f}"
                row[f"IC_High_{code}"] = f"{pred.ci_upper.get(code, 0):.2f}"

            writer.writerow(row)

    print()
    print(f"[OK] CSV exportado: {csv_path}")

    # =========================================================================
    # RESUMO ESTATÍSTICO
    # =========================================================================
    print()
    print("=" * 80)
    print("RESUMO ESTATÍSTICO")
    print("=" * 80)
    scores = [p.final_score for p in predictions]
    nms = [p.nm_dollar_corrected for p in predictions]
    fs = [p.expected_inbreeding * 100 for p in predictions]

    print(f"Total de combinacoes avaliadas: {len(predictions)}")
    print(f"Score Final:  min={min(scores):.1f}  max={max(scores):.1f}  media={sum(scores)/len(scores):.1f}")
    print(f"NM$ Corrigido: min=${min(nms):.0f}  max=${max(nms):.0f}  media=${sum(nms)/len(nms):.0f}")
    print(f"F% Esperada:  min={min(fs):.2f}%  max={max(fs):.2f}%  media={sum(fs)/len(fs):.2f}%")

    n_conflicts = sum(1 for p in predictions if p.haplo_conflicts)
    print(f"Combinacoes com conflito de haplotipos: {n_conflicts} ({n_conflicts/len(predictions)*100:.1f}%)")

    print()
    print("Metodologia: ssGBLUP-derived PA + REL-weighted Mendelian Sampling Variance +")
    print("             Inbreeding Depression (CDCB/EFI) + NM$ 2025 + Risk Analysis +")
    print("             Haplotype Lethal Screening + Casein Compatibility")
    print("Referencias: VanRaden(2008), Aguilar/Misztal(2010), Cole/VanRaden(2025)")


if __name__ == "__main__":
    main()
