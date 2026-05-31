"""
================================================================================
PREDICAO GENOMICA AVANCADA v2 -- RACA HOLANDESA
Usa NM$ publicado (CDCB) como indice primario + crivos avancados
================================================================================

Diferenciais sobre PA simples:
1. NM$ publicado como ancora (CDCB 2025, ja calibrado)
2. Variancia de Mendelian Sampling por REL dos pais
3. Correcao de depressao endogamica (CDCB/EFI)
4. Screening de haplotipos letais (HH1-HH6, HCD, HHR, HMW3/4)
5. Compatibilidade de caseinas (Beta A2A2, Kappa BB)
6. Complementaridade de traits (Mate Allocation Index)
7. Analise probabilistica (P>media, P Top25%, IC 95%)
8. Multi-trait: balanceamento producao x saude x fertilidade x tipo
================================================================================
"""

import os
import sys
import math
import csv
import openpyxl
from dataclasses import dataclass, field
from typing import Optional
from datetime import datetime

# =============================================================================
# CONFIGURACAO
# =============================================================================

DOWNLOADS = os.path.expanduser("~/Downloads")
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# Depressao endogamica por 1% de F (Holstein, CDCB/literatura)
INBREEDING_DEPRESSION = {
    "NM$": -24.0,     # $/1% F (estimativa baseada em componentes)
    "MILK": -29.6,    # lbs/1% F
    "FAT": -1.41,     # lbs/1% F
    "PROT": -1.13,    # lbs/1% F
    "PL": -0.26,      # meses/1% F
    "DPR": -0.24,     # %/1% F
    "SCS": +0.008,    # score/1% F (pior)
}

# Haplotipos letais conhecidos no Holstein
LETHAL_HAPLOTYPES = {
    "HH1", "HH2", "HH3", "HH4", "HH5", "HH6",
    "HCD", "HHR", "HMW3", "HMW4", "HHP",
}

# Desvio padrao genetico aditivo por trait (Holstein)
GENETIC_SD = {
    "MILK": 1100, "FAT": 42, "PROT": 30,
    "SCS": 0.18, "PL": 1.7, "LIV": 1.2, "HLIV": 0.8,
    "DPR": 1.3, "CCR": 1.5, "HCR": 1.4, "EFC": 5.0,
    "UDC": 0.8, "FLC": 0.7, "BWC": 1.0,
    "NM$": 200, "TPI": 250,
}

# REL default por sexo (quando nao disponivel na planilha)
DEFAULT_REL = {
    "M": {"NM$": 0.78, "TPI": 0.78, "MILK": 0.78, "FAT": 0.78, "PROT": 0.78,
           "SCS": 0.72, "PL": 0.68, "LIV": 0.58, "DPR": 0.62, "CCR": 0.58,
           "HCR": 0.52, "EFC": 0.58, "HLIV": 0.52, "UDC": 0.75, "FLC": 0.68,
           "BWC": 0.72, "RFI": 0.55},
    "F": {"NM$": 0.55, "TPI": 0.55, "MILK": 0.55, "FAT": 0.55, "PROT": 0.55,
           "SCS": 0.50, "PL": 0.45, "LIV": 0.38, "DPR": 0.42, "CCR": 0.38,
           "HCR": 0.32, "EFC": 0.38, "HLIV": 0.32, "UDC": 0.52, "FLC": 0.48,
           "BWC": 0.52, "RFI": 0.35},
}


# =============================================================================
# ESTRUTURAS
# =============================================================================

@dataclass
class Animal:
    id: str
    name: str
    sex: str
    num: int = 0
    reg: str = ""
    naab: str = ""
    tpi: float = 0.0
    nm_dollar: float = 0.0
    ptas: dict = field(default_factory=dict)
    rels: dict = field(default_factory=dict)
    inbreeding: float = 0.085
    sire: str = ""
    mgs: str = ""
    mggs: str = ""
    beta_casein: str = ""
    kappa_casein: str = ""
    haplotypes: str = ""
    feed_saved: float = 0.0


@dataclass
class Prediction:
    rank: int = 0
    sire: Animal = None
    dam: Animal = None
    # Valores preditos
    nm_pred: float = 0.0         # PA simples de NM$
    nm_corrected: float = 0.0    # Apos correcao endogamia
    tpi_pred: float = 0.0        # PA simples de TPI
    # Endogamia
    expected_f: float = 0.0
    endo_depression_nm: float = 0.0
    # Variancia e IC
    nm_sd: float = 0.0
    nm_ic_lower: float = 0.0
    nm_ic_upper: float = 0.0
    # Probabilidades
    prob_above_avg: float = 0.0
    prob_top25: float = 0.0
    prob_top10: float = 0.0
    # Traits individuais (PA)
    trait_pred: dict = field(default_factory=dict)
    trait_ic_low: dict = field(default_factory=dict)
    trait_ic_high: dict = field(default_factory=dict)
    # Haplotipos
    haplo_conflicts: list = field(default_factory=list)
    # Caseinas
    beta_info: str = ""
    kappa_info: str = ""
    casein_score: int = 0
    # Complementaridade
    complementarity: float = 0.0
    # Balanceamento (producao vs saude/fertilidade)
    balance_score: float = 0.0
    # Score final
    final_score: float = 0.0


# =============================================================================
# FUNCOES AUXILIARES
# =============================================================================

def sf(val, default=0.0):
    if val is None: return default
    try: return float(val)
    except: return default


def normal_cdf(z):
    if z > 6: return 1.0
    if z < -6: return 0.0
    a = abs(z)
    t = 1.0 / (1.0 + 0.2316419 * a)
    pdf = math.exp(-0.5 * a * a) / math.sqrt(2 * math.pi)
    cdf = 1.0 - pdf * t * (0.319381530 + t * (-0.356563782 + t * (
        1.781477937 + t * (-1.821255978 + t * 1.330274429))))
    return cdf if z >= 0 else 1.0 - cdf


# =============================================================================
# CARGA DE DADOS
# =============================================================================

def load_sires(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    animals = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        naab = row[1].value
        if not naab: continue

        a = Animal(
            id=str(naab), name=str(row[2].value or "?"), sex="M",
            naab=str(naab),
            reg=str(row[3].value or ""),
            tpi=sf(row[13].value),
            nm_dollar=sf(row[14].value),
            ptas={
                "MILK": sf(row[18].value), "FAT": sf(row[19].value),
                "PROT": sf(row[21].value), "SCS": sf(row[24].value),
                "DPR": sf(row[25].value), "CCR": sf(row[26].value),
                "PL": sf(row[27].value), "LIV": sf(row[28].value),
                "HCR": sf(row[36].value), "EFC": sf(row[39].value),
                "HLIV": sf(row[40].value), "RFI": sf(row[41].value),
                "UDC": sf(row[44].value), "FLC": sf(row[45].value),
                "BWC": sf(row[46].value),
            },
            rels=DEFAULT_REL["M"].copy(),
            sire=str(row[68].value or ""),
            mgs=str(row[69].value or ""),
            mggs=str(row[70].value or "") if len(row) > 70 else "",
            beta_casein=str(row[5].value or ""),
            kappa_casein=str(row[6].value or ""),
            haplotypes=str(row[4].value or ""),
            feed_saved=sf(row[42].value),
        )
        animals.append(a)
    return animals


def load_dams(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    animals = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        naab = row[2].value
        if not naab: continue

        a = Animal(
            id=str(naab), name=f"#{row[0].value} Reg.{row[3].value or '?'}",
            sex="F", num=int(row[0].value or 0),
            naab=str(naab),
            reg=str(row[3].value or ""),
            tpi=sf(row[14].value),
            nm_dollar=sf(row[15].value),
            ptas={
                "MILK": sf(row[17].value), "FAT": sf(row[18].value),
                "PROT": sf(row[20].value), "SCS": sf(row[23].value),
                "DPR": sf(row[24].value), "CCR": sf(row[25].value),
                "PL": sf(row[26].value), "LIV": sf(row[27].value),
                "HCR": sf(row[34].value), "EFC": sf(row[37].value),
                "HLIV": sf(row[38].value), "RFI": sf(row[39].value),
                "UDC": sf(row[42].value), "FLC": sf(row[43].value),
                "BWC": sf(row[44].value),
            },
            rels=DEFAULT_REL["F"].copy(),
            sire=str(row[66].value or ""),
            mgs=str(row[67].value or ""),
            beta_casein=str(row[5].value or ""),
            kappa_casein=str(row[6].value or ""),
            haplotypes=str(row[4].value or ""),
            feed_saved=sf(row[40].value),
        )
        animals.append(a)
    return animals


# =============================================================================
# MOTOR DE PREDICAO
# =============================================================================

def estimate_inbreeding(sire, dam):
    """Estima F da cria baseado em parentesco conhecido."""
    relationship = 0.0625  # background Holstein (~6.25%)

    # Meio-irmaos paternos (mesmo pai)
    if sire.sire and dam.sire and sire.sire == dam.sire and sire.sire != "None":
        relationship = max(relationship, 0.25)

    # Pai do touro = avo materno da vaca
    if sire.sire and dam.mgs and sire.sire == dam.mgs and sire.sire != "None":
        relationship = max(relationship, 0.125)

    # Touro e pai da vaca
    if sire.name.lower() == dam.sire.lower() and dam.sire != "None":
        relationship = max(relationship, 0.50)

    # MGS do touro = pai da vaca
    if sire.mgs and dam.sire and sire.mgs == dam.sire and sire.mgs != "None":
        relationship = max(relationship, 0.125)

    f = relationship / 2.0
    # Ajuste pela endogamia parental
    f += (sire.inbreeding + dam.inbreeding) * 0.05
    return min(f, 0.25)


def check_haplotypes(sire, dam):
    """Verifica conflitos de haplotipos letais."""
    s_hap = set(str(sire.haplotypes or "").replace(",", " ").split()) & LETHAL_HAPLOTYPES
    d_hap = set(str(dam.haplotypes or "").replace(",", " ").split()) & LETHAL_HAPLOTYPES
    return sorted(s_hap & d_hap)


def check_caseins(sire, dam):
    """Avalia compatibilidade de caseinas."""
    sb = sire.beta_casein or ""
    db = dam.beta_casein or ""
    sk = sire.kappa_casein or ""
    dk = dam.kappa_casein or ""

    score = 0
    beta_info = f"{sb}x{db}"
    kappa_info = f"{sk}x{dk}"

    # Beta-caseina
    if "A2A2" in sb and "A2A2" in db:
        score += 3; beta_info += "=100%A2A2"
    elif ("A2A2" in sb and "A1A2" in db) or ("A1A2" in sb and "A2A2" in db):
        score += 1; beta_info += "=50%A2A2"
    elif "A1A2" in sb and "A1A2" in db:
        score += 0; beta_info += "=25%A2A2"
    elif "A1A1" in sb or "A1A1" in db:
        score -= 2; beta_info += "=0%A2A2"

    # Kappa-caseina
    if "BB" in sk and "BB" in dk:
        score += 2; kappa_info += "=100%BB"
    elif ("BB" in sk and "AB" in dk) or ("AB" in sk and "BB" in dk):
        score += 1; kappa_info += "=50%BB"
    elif ("BB" in sk and "AA" in dk) or ("AA" in sk and "BB" in dk):
        score += 1; kappa_info += "=100%AB"

    return beta_info, kappa_info, score


def compute_complementarity(sire, dam):
    """
    Score de complementaridade: penaliza fraquezas compartilhadas,
    bonifica quando um compensa o outro.
    """
    # Categorias de traits com pesos de importancia
    categories = {
        "producao": (["MILK", "FAT", "PROT"], 0.35),
        "saude": (["SCS", "PL", "LIV"], 0.25),
        "fertilidade": (["DPR", "CCR", "HCR"], 0.20),
        "tipo": (["UDC", "FLC"], 0.10),
        "eficiencia": (["BWC"], 0.10),
    }

    # Traits onde menor e melhor
    lower_better = {"SCS", "BWC"}

    score = 50.0  # base

    for cat_name, (traits, weight) in categories.items():
        for trait in traits:
            s_val = sire.ptas.get(trait, 0)
            d_val = dam.ptas.get(trait, 0)
            sd = GENETIC_SD.get(trait, 1)

            if trait in lower_better:
                s_z = -s_val / sd
                d_z = -d_val / sd
            else:
                s_z = s_val / sd
                d_z = d_val / sd

            if s_z < -0.5 and d_z < -0.5:
                # Ambos fracos: penalidade
                score -= abs(s_z + d_z) * weight * 30
            elif (s_z < -0.5 and d_z > 0.5) or (s_z > 0.5 and d_z < -0.5):
                # Complementar: bonus
                score += min(abs(s_z), abs(d_z)) * weight * 15
            elif s_z > 0.5 and d_z > 0.5:
                # Ambos fortes: bonus moderado
                score += min(s_z, d_z) * weight * 10

    return max(0, min(100, score))


def compute_balance(sire, dam):
    """
    Avalia o balanceamento producao vs funcionalidade.
    Penaliza animais muito unilaterais.
    """
    # Producao
    milk_z = ((sire.ptas.get("MILK", 0) + dam.ptas.get("MILK", 0)) / 2) / GENETIC_SD["MILK"]
    fat_z = ((sire.ptas.get("FAT", 0) + dam.ptas.get("FAT", 0)) / 2) / GENETIC_SD["FAT"]
    prod_score = (milk_z + fat_z * 2) / 3  # Fat pesa mais no NM$2025

    # Fertilidade
    dpr_z = ((sire.ptas.get("DPR", 0) + dam.ptas.get("DPR", 0)) / 2) / GENETIC_SD["DPR"]
    ccr_z = ((sire.ptas.get("CCR", 0) + dam.ptas.get("CCR", 0)) / 2) / GENETIC_SD["CCR"]
    fert_score = (dpr_z + ccr_z) / 2

    # Longevidade
    pl_z = ((sire.ptas.get("PL", 0) + dam.ptas.get("PL", 0)) / 2) / GENETIC_SD["PL"]

    # Balance: penaliza se producao e alta mas fertilidade/PL e baixa
    if prod_score > 1.0 and (fert_score < -0.5 or pl_z < -0.5):
        return max(0, 50 - abs(prod_score - fert_score) * 15)
    elif prod_score > 0.5 and fert_score > 0 and pl_z > 0:
        return min(100, 60 + prod_score * 10 + fert_score * 10)
    else:
        return 50 + (prod_score + fert_score + pl_z) * 8


def predict_mating(sire, dam, pop_avg_nm, pop_sd_nm):
    """Predicao completa de um acasalamento."""
    p = Prediction(sire=sire, dam=dam)

    # 1. Parent Average de NM$ e TPI (usando valores publicados CDCB)
    p.nm_pred = (sire.nm_dollar + dam.nm_dollar) / 2.0
    p.tpi_pred = (sire.tpi + dam.tpi) / 2.0

    # 2. Endogamia esperada
    p.expected_f = estimate_inbreeding(sire, dam)

    # 3. Correcao de depressao endogamica no NM$
    f_pct = p.expected_f * 100.0
    p.endo_depression_nm = INBREEDING_DEPRESSION["NM$"] * f_pct
    p.nm_corrected = p.nm_pred + p.endo_depression_nm

    # 4. Variancia de Mendelian Sampling + incerteza dos PTAs
    sigma2_nm = GENETIC_SD["NM$"] ** 2
    rel_s = sire.rels.get("NM$", 0.78)
    rel_d = dam.rels.get("NM$", 0.55)

    var_mendelian = 0.5 * (1.0 - (sire.inbreeding + dam.inbreeding) / 2) * sigma2_nm
    var_uncertainty = 0.25 * ((1 - rel_s) + (1 - rel_d)) * sigma2_nm
    total_var = var_mendelian + var_uncertainty
    p.nm_sd = math.sqrt(total_var)

    # IC 95%
    p.nm_ic_lower = p.nm_corrected - 1.96 * p.nm_sd
    p.nm_ic_upper = p.nm_corrected + 1.96 * p.nm_sd

    # 5. Probabilidades
    z_avg = (p.nm_corrected - pop_avg_nm) / p.nm_sd if p.nm_sd > 0 else 0
    z_25 = (pop_avg_nm + 0.674 * pop_sd_nm - p.nm_corrected) / p.nm_sd if p.nm_sd > 0 else 0
    z_10 = (pop_avg_nm + 1.282 * pop_sd_nm - p.nm_corrected) / p.nm_sd if p.nm_sd > 0 else 0
    p.prob_above_avg = normal_cdf(z_avg)
    p.prob_top25 = 1.0 - normal_cdf(z_25)
    p.prob_top10 = 1.0 - normal_cdf(z_10)

    # 6. Traits individuais (PA + IC)
    for trait in ["MILK", "FAT", "PROT", "SCS", "PL", "LIV", "DPR", "CCR",
                  "HCR", "EFC", "HLIV", "UDC", "FLC", "BWC"]:
        s_val = sire.ptas.get(trait, 0)
        d_val = dam.ptas.get(trait, 0)
        pa = (s_val + d_val) / 2.0

        # Correcao endogamica por trait
        dep = INBREEDING_DEPRESSION.get(trait, 0) * f_pct
        pa_corr = pa + dep

        p.trait_pred[trait] = pa_corr

        # IC baseado em Mendelian Sampling
        sd = GENETIC_SD.get(trait, 1)
        rel_s_t = sire.rels.get(trait, 0.70)
        rel_d_t = dam.rels.get(trait, 0.50)
        var_t = 0.5 * sd**2 + 0.25 * ((1-rel_s_t) + (1-rel_d_t)) * sd**2
        sd_t = math.sqrt(var_t)
        p.trait_ic_low[trait] = pa_corr - 1.96 * sd_t
        p.trait_ic_high[trait] = pa_corr + 1.96 * sd_t

    # 7. Haplotipos
    p.haplo_conflicts = check_haplotypes(sire, dam)

    # 8. Caseinas
    p.beta_info, p.kappa_info, p.casein_score = check_caseins(sire, dam)

    # 9. Complementaridade
    p.complementarity = compute_complementarity(sire, dam)

    # 10. Balanceamento
    p.balance_score = compute_balance(sire, dam)

    # 11. Score Final Composto
    # Normaliza NM$ para escala 0-100
    nm_norm = min(max((p.nm_corrected - pop_avg_nm) / pop_sd_nm * 20 + 50, 0), 100)

    # Penalidade de endogamia (0-100, onde F>10% = 0)
    endo_score = max(0, min(100, 100 - p.expected_f * 1000))

    # Penalidade de haplotipos letais
    haplo_penalty = len(p.haplo_conflicts) * 15

    # Bonus de caseinas (max +6)
    casein_bonus = p.casein_score * 1.5

    p.final_score = (
        0.40 * nm_norm +                    # Merito economico
        0.15 * p.complementarity +           # Complementaridade
        0.15 * p.balance_score +             # Balanceamento prod/func
        0.10 * min(100, p.prob_top25 * 100) + # Probabilidade top 25%
        0.10 * endo_score +                  # Penalidade endogamia
        0.10 * min(100, (p.tpi_pred / 35) * 1)  # TPI (bonus)
        - haplo_penalty
        + casein_bonus
    )
    p.final_score = max(0, min(100, p.final_score))

    return p


# =============================================================================
# MAIN
# =============================================================================

def main():
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    print("=" * 120)
    print("  PREDICAO GENOMICA AVANCADA v2 -- RACA HOLANDESA")
    print("  Motor: PA(NM$ CDCB) + Mendelian Sampling + Endogamia + Haplotipos + Caseinas")
    print("  Data:", datetime.now().strftime("%Y-%m-%d %H:%M"))
    print("=" * 120)

    # Carrega dados
    sires = load_sires(os.path.join(DOWNLOADS, "Lista de touros.xlsx"))
    dams = load_dams(os.path.join(DOWNLOADS, "Lista de f\u00eameas.xlsx"))

    print(f"\nTouros: {len(sires)}")
    for s in sires:
        print(f"  {s.name:<15} TPI={s.tpi:>5.0f}  NM$={s.nm_dollar:>5.0f}  "
              f"Milk={s.ptas['MILK']:>5.0f}  Fat={s.ptas['FAT']:>4.0f}  "
              f"DPR={s.ptas['DPR']:>+5.1f}  PL={s.ptas['PL']:>4.1f}  "
              f"Beta={s.beta_casein}  Kappa={s.kappa_casein}")

    print(f"\nFemeas: {len(dams)}")
    print(f"  TPI range: {min(d.tpi for d in dams):.0f} - {max(d.tpi for d in dams):.0f}")
    print(f"  NM$ range: {min(d.nm_dollar for d in dams):.0f} - {max(d.nm_dollar for d in dams):.0f}")
    print(f"\nTotal combinacoes: {len(sires) * len(dams)}")

    # Populacao de referencia
    all_nm = [d.nm_dollar for d in dams]
    pop_avg = sum(all_nm) / len(all_nm)
    pop_sd = max((sum((x - pop_avg)**2 for x in all_nm) / (len(all_nm)-1))**0.5, 100)
    print(f"Media NM$ femeas: ${pop_avg:.0f}  |  SD: ${pop_sd:.0f}")

    # Roda predicoes
    predictions = []
    for s in sires:
        for d in dams:
            p = predict_mating(s, d, pop_avg, pop_sd)
            predictions.append(p)

    predictions.sort(key=lambda x: x.final_score, reverse=True)
    for i, p in enumerate(predictions):
        p.rank = i + 1

    # =========================================================================
    # RANKING TOP 50
    # =========================================================================
    print("\n" + "=" * 140)
    print("  RANKING -- TOP 50 MELHORES ACASALAMENTOS")
    print("=" * 140)

    hdr = (f"{'Rk':>3} | {'Touro':<12} | {'Femea':<20} | "
           f"{'NM$T':>5} | {'NM$F':>5} | {'NM$Pred':>7} | {'NM$Corr':>7} | "
           f"{'IC95-':>7} | {'IC95+':>7} | "
           f"{'F%':>5} | {'DepEnd':>6} | "
           f"{'P>Med':>5} | {'PT25':>4} | "
           f"{'Comp':>4} | {'Bal':>4} | "
           f"{'Score':>5} | {'Haplo':>6} | {'BetaCas':>13}")
    print(hdr)
    print("-" * 140)

    for p in predictions[:50]:
        haplo = ",".join(p.haplo_conflicts) if p.haplo_conflicts else "OK"
        print(
            f"{p.rank:>3} | "
            f"{p.sire.name:<12.12} | "
            f"{p.dam.name:<20.20} | "
            f"${p.sire.nm_dollar:>4.0f} | ${p.dam.nm_dollar:>4.0f} | "
            f"${p.nm_pred:>6.0f} | ${p.nm_corrected:>6.0f} | "
            f"${p.nm_ic_lower:>6.0f} | ${p.nm_ic_upper:>6.0f} | "
            f"{p.expected_f*100:>4.1f}% | "
            f"${p.endo_depression_nm:>5.0f} | "
            f"{p.prob_above_avg*100:>4.0f}% | "
            f"{p.prob_top25*100:>3.0f}% | "
            f"{p.complementarity:>4.0f} | "
            f"{p.balance_score:>4.0f} | "
            f"{p.final_score:>5.1f} | "
            f"{haplo:>6.6} | "
            f"{p.beta_info[:13]}"
        )

    # =========================================================================
    # MELHOR TOURO POR FEMEA
    # =========================================================================
    print("\n" + "=" * 140)
    print("  MELHOR TOURO PARA CADA FEMEA (ordenado por TPI da femea)")
    print("=" * 140)

    best_per_dam = {}
    all_per_dam = {}
    for p in predictions:
        did = p.dam.id
        if did not in all_per_dam:
            all_per_dam[did] = []
        all_per_dam[did].append(p)
        if did not in best_per_dam:
            best_per_dam[did] = p

    sorted_best = sorted(best_per_dam.values(), key=lambda x: x.dam.tpi, reverse=True)

    print(f"{'Femea':<22} | {'TPI':>4} | {'NM$F':>5} | "
          f"{'1o Touro':<12} | {'NM$Corr':>7} | {'Score':>5} | "
          f"{'2o Touro':<12} | {'Score2':>5} | "
          f"{'3o Touro':<12} | {'Score3':>5} | "
          f"{'Haplo':>8} | {'BetaCas':>13}")
    print("-" * 140)

    for bp in sorted_best:
        did = bp.dam.id
        ranked = sorted(all_per_dam[did], key=lambda x: x.final_score, reverse=True)
        p1 = ranked[0]
        p2 = ranked[1] if len(ranked) > 1 else None
        p3 = ranked[2] if len(ranked) > 2 else None
        haplo = ",".join(p1.haplo_conflicts) if p1.haplo_conflicts else "OK"

        print(
            f"{p1.dam.name:<22.22} | "
            f"{p1.dam.tpi:>4.0f} | ${p1.dam.nm_dollar:>4.0f} | "
            f"{p1.sire.name:<12.12} | ${p1.nm_corrected:>6.0f} | {p1.final_score:>5.1f} | "
            f"{(p2.sire.name if p2 else '-'):<12.12} | {(f'{p2.final_score:.1f}' if p2 else '-'):>5} | "
            f"{(p3.sire.name if p3 else '-'):<12.12} | {(f'{p3.final_score:.1f}' if p3 else '-'):>5} | "
            f"{haplo:>8.8} | {p1.beta_info[:13]}"
        )

    # =========================================================================
    # DETALHAMENTO TOP 10
    # =========================================================================
    print("\n" + "=" * 120)
    print("  DETALHAMENTO -- TOP 10")
    print("=" * 120)

    for p in predictions[:10]:
        print(f"\n  #{p.rank}: {p.sire.name} ({p.sire.naab}) x {p.dam.name} ({p.dam.naab})")
        print(f"  Score Final: {p.final_score:.1f}/100")
        print(f"  NM$ Predito: ${p.nm_pred:.0f}  |  Corrigido: ${p.nm_corrected:.0f}  |  IC95%: [${p.nm_ic_lower:.0f} , ${p.nm_ic_upper:.0f}]")
        print(f"  TPI Predito: {p.tpi_pred:.0f}")
        print(f"  F esperada: {p.expected_f*100:.2f}%  |  Depressao: ${p.endo_depression_nm:.0f}")
        print(f"  P(>media): {p.prob_above_avg*100:.1f}%  |  P(Top25%): {p.prob_top25*100:.1f}%  |  P(Top10%): {p.prob_top10*100:.1f}%")
        print(f"  Complementaridade: {p.complementarity:.0f}  |  Balanceamento: {p.balance_score:.0f}")
        print(f"  Haplotipos: {','.join(p.haplo_conflicts) if p.haplo_conflicts else 'Nenhum conflito'}")
        print(f"  Caseinas: Beta={p.beta_info}  |  Kappa={p.kappa_info}")

        print(f"\n  {'Trait':<22} | {'Touro':>7} | {'Vaca':>7} | {'Pred':>7} | {'IC 5%':>7} | {'IC95%':>7}")
        print("  " + "-" * 72)
        for t in ["MILK", "FAT", "PROT", "SCS", "PL", "LIV", "DPR", "CCR",
                   "HCR", "EFC", "HLIV", "UDC", "FLC", "BWC"]:
            sv = p.sire.ptas.get(t, 0)
            dv = p.dam.ptas.get(t, 0)
            pv = p.trait_pred.get(t, 0)
            lo = p.trait_ic_low.get(t, 0)
            hi = p.trait_ic_high.get(t, 0)
            print(f"  {t:<22} | {sv:>7.1f} | {dv:>7.1f} | {pv:>7.1f} | {lo:>7.1f} | {hi:>7.1f}")

    # =========================================================================
    # ALERTAS
    # =========================================================================
    conflicts = [p for p in predictions if p.haplo_conflicts]
    if conflicts:
        print("\n" + "=" * 80)
        print("  ALERTAS DE HAPLOTIPOS LETAIS")
        print("=" * 80)
        for p in conflicts[:30]:
            print(f"  Rank #{p.rank}: {p.sire.name} x {p.dam.name} -> {','.join(p.haplo_conflicts)}")
        if len(conflicts) > 30:
            print(f"  ... e mais {len(conflicts)-30}")

    # =========================================================================
    # EXPORT CSV
    # =========================================================================
    csv_path = os.path.join(OUTPUT_DIR, "predicao_genomica_resultado.csv")
    fields = [
        "Rank", "Touro_NAAB", "Touro_Nome", "Touro_TPI", "Touro_NM$",
        "Femea_NAAB", "Femea_Num", "Femea_Reg", "Femea_TPI", "Femea_NM$",
        "NM$_PA", "NM$_Corrigido", "NM$_IC_Lower", "NM$_IC_Upper",
        "TPI_PA", "F_Esperada_%", "Depressao_$",
        "P_Acima_Media_%", "P_Top25_%", "P_Top10_%",
        "Complementaridade", "Balanceamento", "Score_Final",
        "Haplo_Conflitos", "Beta_Caseina", "Kappa_Caseina", "Casein_Score",
    ]
    traits_list = ["MILK", "FAT", "PROT", "SCS", "PL", "LIV", "DPR", "CCR",
                   "HCR", "EFC", "HLIV", "UDC", "FLC", "BWC"]
    for t in traits_list:
        fields.extend([f"PA_{t}", f"IC_Low_{t}", f"IC_High_{t}"])

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fields, delimiter=";")
        w.writeheader()
        for p in predictions:
            row = {
                "Rank": p.rank,
                "Touro_NAAB": p.sire.naab, "Touro_Nome": p.sire.name,
                "Touro_TPI": f"{p.sire.tpi:.0f}", "Touro_NM$": f"{p.sire.nm_dollar:.0f}",
                "Femea_NAAB": p.dam.naab, "Femea_Num": p.dam.num,
                "Femea_Reg": p.dam.reg, "Femea_TPI": f"{p.dam.tpi:.0f}",
                "Femea_NM$": f"{p.dam.nm_dollar:.0f}",
                "NM$_PA": f"{p.nm_pred:.0f}", "NM$_Corrigido": f"{p.nm_corrected:.0f}",
                "NM$_IC_Lower": f"{p.nm_ic_lower:.0f}", "NM$_IC_Upper": f"{p.nm_ic_upper:.0f}",
                "TPI_PA": f"{p.tpi_pred:.0f}",
                "F_Esperada_%": f"{p.expected_f*100:.2f}",
                "Depressao_$": f"{p.endo_depression_nm:.0f}",
                "P_Acima_Media_%": f"{p.prob_above_avg*100:.1f}",
                "P_Top25_%": f"{p.prob_top25*100:.1f}",
                "P_Top10_%": f"{p.prob_top10*100:.1f}",
                "Complementaridade": f"{p.complementarity:.1f}",
                "Balanceamento": f"{p.balance_score:.1f}",
                "Score_Final": f"{p.final_score:.1f}",
                "Haplo_Conflitos": ",".join(p.haplo_conflicts) if p.haplo_conflicts else "",
                "Beta_Caseina": p.beta_info,
                "Kappa_Caseina": p.kappa_info,
                "Casein_Score": p.casein_score,
            }
            for t in traits_list:
                row[f"PA_{t}"] = f"{p.trait_pred.get(t,0):.1f}"
                row[f"IC_Low_{t}"] = f"{p.trait_ic_low.get(t,0):.1f}"
                row[f"IC_High_{t}"] = f"{p.trait_ic_high.get(t,0):.1f}"
            w.writerow(row)

    print(f"\n[OK] CSV: {csv_path}")

    # =========================================================================
    # RESUMO
    # =========================================================================
    print("\n" + "=" * 80)
    print("  RESUMO ESTATISTICO")
    print("=" * 80)
    scores = [p.final_score for p in predictions]
    nms = [p.nm_corrected for p in predictions]
    fs = [p.expected_f*100 for p in predictions]
    print(f"  Combinacoes: {len(predictions)}")
    print(f"  Score:  min={min(scores):.1f}  max={max(scores):.1f}  media={sum(scores)/len(scores):.1f}")
    print(f"  NM$:   min=${min(nms):.0f}  max=${max(nms):.0f}  media=${sum(nms)/len(nms):.0f}")
    print(f"  F%:    min={min(fs):.2f}%  max={max(fs):.2f}%  media={sum(fs)/len(fs):.2f}%")
    print(f"  Conflitos haplot.: {sum(1 for p in predictions if p.haplo_conflicts)} ({sum(1 for p in predictions if p.haplo_conflicts)/len(predictions)*100:.1f}%)")
    print()
    print("  Metodologia: PA(NM$ CDCB 2025) + Mendelian Sampling Variance +")
    print("               Inbreeding Depression (EFI) + Haplotype Screening +")
    print("               Casein Compatibility + Trait Complementarity + Balance Index")
    print("  Refs: VanRaden(2008), Aguilar/Misztal(2010), Cole/VanRaden(2025)")


if __name__ == "__main__":
    main()
