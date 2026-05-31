"""Test: does Banco2 sire stats improve DSII v5?"""
import warnings
warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd
import openpyxl, csv, re
from pathlib import Path
from collections import defaultdict
from sklearn.model_selection import KFold
from sklearn.metrics import mean_absolute_error, r2_score
from lightgbm import LGBMRegressor

DOWNLOADS = Path("C:/Users/DiegoGuerra/Downloads")
BULLS_CSV = Path("C:/Users/DiegoGuerra/gen-genie-advisor/sms-engine/data/bulls.csv")

# Load bulls
bulls_by_naab, bulls_by_name, bulls_by_regname = {}, {}, {}
with open(BULLS_CSV, 'r', encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        naab = row.get('NAAB', '').strip().upper()
        nm = row.get('Name', '').strip().upper()
        rn = row.get('Registration Name', '').strip().upper()
        if naab: bulls_by_naab[naab] = row
        if nm: bulls_by_name[nm] = row
        if rn: bulls_by_regname[rn] = row

def normalize_naab(code):
    code = code.strip().upper()
    m = re.match(r'^(\d+)(HO\d+)$', code)
    if m: return m.group(1).zfill(3) + m.group(2)
    m2 = re.match(r'^(\d{3})H(\d+)$', code)
    if m2: return m2.group(1) + 'HO' + m2.group(2)
    return code

def find_sire_bull(s1, s2=''):
    if s1:
        norm = normalize_naab(s1)
        for d in [bulls_by_naab, bulls_by_name, bulls_by_regname]:
            if norm in d: return d[norm]
            if s1 in d: return d[s1]
    if s2:
        for d in [bulls_by_name, bulls_by_regname]:
            if s2 in d: return d[s2]
    return None

def sf(v):
    if v is None: return None
    try:
        x = float(v)
        return x if not np.isnan(x) else None
    except (ValueError, TypeError): return None

# Trait mappings
BULL_COL = {'NM$':'NM$','MILK':'PTAM','FAT':'PTAF','PRO':'PTAP','SCS':'SCS',
            'CFP':'CFP','PL':'PL','DPR':'DPR','PTAT':'PTAT','LIV':'LIV',
            'MAST':'MAST','UDC':'UDC'}
B2_COL = {'NM$':4,'MILK':5,'FAT':6,'PRO':7,'SCS':8,'CFP':9,'PL':10,
          'DPR':11,'PTAT':14,'LIV':20,'MAST':25,'UDC':36}
DAUGHTER_COL = {'NM$':'NM$','MILK':'MILK','FAT':'FAT','PRO':'PRO','SCS':'SCS',
                'CFP':'CFP','PL':'PL','DPR':'DPR','PTAT':'PTAT','LIV':'LIV',
                'MAST':'MAST','UDC':'UDC'}
DAM_COL = {'NM$':'Net Merit','MILK':'PTA Milk','FAT':'PTA Fat','PRO':'PTA Pro',
           'SCS':'SCS','CFP':'CFP','PL':'PL','DPR':'PTA DPR','PTAT':'PTA Type',
           'LIV':'PTA LIV','MAST':'Mastitis','UDC':'UDC'}

TRAITS = list(B2_COL.keys())

# ---- Load Banco2 and build sire performance stats ----
print("Loading Banco2...")
wb = openpyxl.load_workbook(DOWNLOADS / "Banco2.xlsx", read_only=True, data_only=True)
ws = wb.active
b2_rows = list(ws.iter_rows(values_only=True))
wb.close()

sire_stats = defaultdict(lambda: defaultdict(list))
b2_matched = 0
for r in b2_rows[1:]:
    s1 = str(r[1]).strip().upper() if r[1] else ''
    s2 = str(r[2]).strip().upper() if r[2] else ''
    sire = find_sire_bull(s1, s2)
    if not sire:
        continue
    b2_matched += 1
    sire_regname = sire.get('Registration Name', '').strip().upper()
    sire_name = sire.get('Name', '').strip().upper()
    sire_key = sire_regname or sire_name
    for trait, col_idx in B2_COL.items():
        dv = sf(r[col_idx])
        if dv is not None:
            sire_stats[sire_key][trait].append(dv)

print(f"  Banco2 matched: {b2_matched}")
print(f"  Sires with stats: {len(sire_stats)}")
avg_daughters = np.mean([len(v.get('NM$', [])) for v in sire_stats.values() if 'NM$' in v])
print(f"  Avg daughters/sire: {avg_daughters:.1f}")

# ---- Load May 2026 trios ----
print("\nLoading May 2026 trios...")
wb = openpyxl.load_workbook(DOWNLOADS / "May 2026.xlsx", read_only=True, data_only=True)
ws = wb.active
may_rows = list(ws.iter_rows(values_only=True))
wb.close()
hdr = [str(c).strip() if c else f'col_{i}' for i, c in enumerate(may_rows[0])]

dams = {}
for i in range(1, 21):
    f = DOWNLOADS / f"DAM{i}.xlsx"
    if not f.exists(): continue
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb.active
    dr = list(ws.iter_rows(values_only=True))
    wb.close()
    dh = [str(c).strip() if c else f'col_{j}' for j, c in enumerate(dr[0])]
    ni = next(j for j, h in enumerate(dh) if 'naab' in h.lower() and 'code' in h.lower())
    for r in dr[1:]:
        if r[ni]:
            dams[str(r[ni]).strip()] = {dh[j]: r[j] for j in range(min(len(dh), len(r)))}

records = []
for r in may_rows[1:]:
    rd = {hdr[i]: r[i] for i in range(min(len(hdr), len(r)))}
    sn = str(rd.get('SIRENAME', '')).strip().upper()
    sire = bulls_by_regname.get(sn) or bulls_by_name.get(sn)
    if not sire: continue
    dr2 = str(rd.get('DAMREGNUM2', '')).strip()
    dam = dams.get(dr2)
    if not dam: continue

    sire_key = sire.get('Registration Name', '').strip().upper() or sn
    rec = {'sire_key': sire_key}
    for t in TRAITS:
        rec[f'd_{t}'] = sf(rd.get(DAUGHTER_COL[t]))
        rec[f's_{t}'] = sf(sire.get(BULL_COL[t]))
        rec[f'm_{t}'] = sf(dam.get(DAM_COL[t]))
    records.append(rec)

may_df = pd.DataFrame(records)
print(f"  Trios: {len(may_df)}")

# ---- Compare V5 vs V5+Banco2 ----
print(f"\n{'Trait':>6} | {'V5 R2':>7} {'V5 MAE':>8} | {'V5+B2 R2':>8} {'V5+B2 MAE':>9} | "
      "{'Gain R2':>7} {'Gain MAE':>8} | B2 coverage")
print("-" * 100)

lgbm_kw = dict(n_estimators=500, max_depth=6, learning_rate=0.03,
               subsample=0.8, colsample_bytree=0.7, min_child_samples=3,
               reg_alpha=0.1, reg_lambda=1.0, random_state=42, verbose=-1)

for trait in TRAITS:
    target = f'd_{trait}'
    s_col = f's_{trait}'
    m_col = f'm_{trait}'

    needed = [target, s_col, m_col]
    mask = may_df[needed].notna().all(axis=1)
    sub = may_df[mask].reset_index(drop=True)
    if len(sub) < 50: continue

    y = sub[target].values

    # Base features
    sv = sub[s_col].values
    mv = sub[m_col].values
    pa = (sv + mv) / 2
    feats_base = [sv, mv, pa, sv - mv, sv * mv, sv**2, mv**2]

    # Cross-trait features
    for ot in TRAITS:
        if ot == trait: continue
        osv = sub[f's_{ot}'].fillna(0).values.astype(float)
        omv = sub[f'm_{ot}'].fillna(0).values.astype(float)
        feats_base.extend([osv, omv])

    X_base = np.column_stack(feats_base)

    # Banco2 sire stats features
    b2_feats = []
    for _, row in sub.iterrows():
        sk = row['sire_key']
        stats = sire_stats.get(sk, {}).get(trait, [])
        if len(stats) >= 2:
            b2_feats.append([np.mean(stats), np.std(stats), np.median(stats),
                            len(stats), np.percentile(stats, 25), np.percentile(stats, 75)])
        else:
            b2_feats.append([np.nan] * 6)

    b2_arr = np.array(b2_feats)
    for c in range(b2_arr.shape[1]):
        col = b2_arr[:, c]
        m = np.nanmean(col) if not np.all(np.isnan(col)) else 0
        col[np.isnan(col)] = m
        b2_arr[:, c] = col

    X_enriched = np.column_stack([X_base, b2_arr])

    coverage = sum(1 for f in b2_feats if not np.isnan(f[0])) if b2_feats else 0

    # Cross-validate
    kf = KFold(n_splits=5, shuffle=True, random_state=42)

    v5_r2s, v5_maes = [], []
    v5b_r2s, v5b_maes = [], []

    for tr, te in kf.split(X_base):
        m1 = LGBMRegressor(**lgbm_kw)
        m1.fit(X_base[tr], y[tr])
        p1 = m1.predict(X_base[te])
        v5_r2s.append(r2_score(y[te], p1))
        v5_maes.append(mean_absolute_error(y[te], p1))

        m2 = LGBMRegressor(**lgbm_kw)
        m2.fit(X_enriched[tr], y[tr])
        p2 = m2.predict(X_enriched[te])
        v5b_r2s.append(r2_score(y[te], p2))
        v5b_maes.append(mean_absolute_error(y[te], p2))

    v5r = np.mean(v5_r2s)
    v5m = np.mean(v5_maes)
    vbr = np.mean(v5b_r2s)
    vbm = np.mean(v5b_maes)

    gr2 = (vbr - v5r) / max(abs(v5r), 0.001) * 100
    gm = (v5m - vbm) / v5m * 100

    print(f"  {trait:>5} | {v5r:>7.4f} {v5m:>8.3f} | {vbr:>8.4f} {vbm:>9.3f} | "
          f"{gr2:>+7.2f}% {gm:>+7.2f}% | {coverage}/{len(sub)} ({100*coverage/len(sub):.0f}%)")

print("\nDone.")
