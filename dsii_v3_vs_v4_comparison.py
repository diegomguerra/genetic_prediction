"""
DSII v3 vs v4 Comparison
========================
Aplica DSII v3 (deficiency-weighted PA) e DSII v4 (ML) nos mesmos 1.709 trios
e compara qual se aproxima mais dos valores genomicos reais das filhas.
"""

import warnings
warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd
import openpyxl
import csv
import math
from pathlib import Path
from sklearn.model_selection import KFold
from sklearn.linear_model import Ridge
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score

try:
    from xgboost import XGBRegressor
    HAS_XGB = True
except ImportError:
    HAS_XGB = False

DOWNLOADS = Path("C:/Users/DiegoGuerra/Downloads")
BULLS_CSV = Path("C:/Users/DiegoGuerra/gen-genie-advisor/sms-engine/data/bulls.csv")
OUTPUT_DIR = Path("C:/Users/DiegoGuerra/Documents/Projetos/genetic_prediction/dsii_v4_results")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ============================================================
# DSII v3 PARAMETERS (from rodar_predicao_v3.py)
# ============================================================

ECON_WEIGHTS = {
    "FAT": 2.60, "PRO": 1.40, "MILK": -0.02, "PL": 35.00,
    "SCS": -120.0, "DPR": 28.00, "CCR": 11.00, "HCR": 6.00,
    "LIV": 25.00, "UDC": 30.00, "FLC": 5.00,
    "SCE": -3.50, "MAST": -25.0, "FSAV": -1.00,
}

GENETIC_SD = {
    "MILK": 675, "FAT": 29, "PRO": 19, "CFP": 25,
    "SCS": 0.14, "PL": 1.85, "LIV": 1.2,
    "DPR": 1.3, "CCR": 1.65, "HCR": 1.4,
    "UDC": 0.75, "FLC": 0.65,
    "NM$": 275, "TPI": 250,
    "SCE": 1.5, "SSB": 1.0, "DSB": 1.0,
    "MAST": 1.0, "FSAV": 50, "FI": 1.0,
    "PTAT": 0.70, "FAT%": 0.05, "PRO%": 0.02,
    "CM$": 280,
}

BREED_IDEAL = {
    "MILK": 1800, "FAT": 90, "PRO": 60, "CFP": 70,
    "DPR": 1.5, "CCR": 2.0, "HCR": 2.0,
    "PL": 5.0, "LIV": 2.0,
    "UDC": 1.5, "FLC": 1.0,
    "MAST": -1.0,
}

LOWER_IS_BETTER = {"SCS", "SCE", "SSB", "DSB", "MAST"}

INBREEDING_DEPRESSION = {
    "NM$": -25.0, "MILK": -28.5, "FAT": -1.0, "PRO": -0.9,
    "PL": -0.35, "DPR": -0.03, "SCS": 0.007, "LIV": -0.15,
    "CCR": -0.025, "HCR": -0.02,
}

CRITICAL_THRESHOLDS = {
    "DPR": -2.0, "PL": -1.0, "SCS": 3.30,
    "UDC": -1.0, "FLC": -1.0, "SCE": 5.0,
}

GENETIC_CORRELATIONS = {
    ("MILK", "DPR"): -0.35, ("MILK", "CCR"): -0.30,
    ("MILK", "SCS"): 0.10, ("MILK", "PL"): -0.15,
    ("MILK", "UDC"): -0.15, ("FAT", "PRO"): 0.65,
    ("FAT", "DPR"): -0.25, ("FAT", "CCR"): -0.20,
    ("PRO", "DPR"): -0.30, ("PRO", "CCR"): -0.25,
    ("DPR", "CCR"): 0.55, ("DPR", "PL"): 0.45,
    ("CCR", "PL"): 0.30, ("UDC", "PL"): 0.25,
    ("FLC", "PL"): 0.10, ("SCS", "PL"): -0.30,
    ("SCS", "UDC"): -0.25, ("SCS", "MAST"): 0.70,
}

# ============================================================
# TRAIT MAPPING (same as v4)
# ============================================================
TRAIT_MAP = [
    ('TPI',    'GTPI',       'TPI',          'TPI'),
    ('NM$',    'NM$',        'Net Merit',    'NM$'),
    ('CM$',    'CM$',        'CM$',          'CM$'),
    ('MILK',   'MILK',       'PTA Milk',     'PTAM'),
    ('FAT',    'FAT',        'PTA Fat',      'PTAF'),
    ('FAT%',   '%F',         '% Fat',        'PTAF%'),
    ('PRO',    'PRO',        'PTA Pro',      'PTAP'),
    ('PRO%',   '%P',         '% Pro',        'PTAP%'),
    ('CFP',    'CFP',        'CFP',          'CFP'),
    ('PL',     'PL',         'PL',           'PL'),
    ('SCS',    'SCS',        'SCS',          'SCS'),
    ('DPR',    'DPR',        'PTA DPR',      'DPR'),
    ('LIV',    'LIV',        'PTA LIV',      'LIV'),
    ('FI',     'FI',         'Fertil Index',  'FI'),
    ('HCR',    'HCR',        'HCR',          'HCR'),
    ('CCR',    'CCR',        'CCR',          'CCR'),
    ('MAST',   'MAST',       'Mastitis',     'MAST'),
    ('FSAV',   'FSAV',       'Feed Saved',   'F SAV'),
    ('PTAT',   'PTAT',       'PTA Type',     'PTAT'),
    ('UDC',    'UDC',        'UDC',          'UDC'),
    ('FLC',    'FLC',        'FLC',          'FLC'),
    ('SCE',    'SCE',        'SCE',          'SCE'),
    ('DCE',    'DCE',        'DCE',          'DCE'),
    ('SSB',    'SSB',        'SSB',          'SSB'),
    ('DSB',    'DSB',        'DSB',          'DSB'),
]


def safe_float(val):
    if val is None:
        return None
    try:
        v = float(val)
        return v if not np.isnan(v) else None
    except (ValueError, TypeError):
        return None


# ============================================================
# DATA LOADING
# ============================================================

def load_all_data():
    print("Carregando dados...")

    # Daughters
    wb = openpyxl.load_workbook(DOWNLOADS / "May 2026.xlsx", read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr_d = [str(c).strip() if c else f'col_{i}' for i, c in enumerate(rows[0])]
    daughters = []
    for r in rows[1:]:
        d = {hdr_d[i]: r[i] for i in range(len(hdr_d)) if i < len(r)}
        daughters.append(d)
    print(f"  Filhas: {len(daughters)}")

    # Dams
    dams = {}
    for i in range(1, 21):
        f = DOWNLOADS / f"DAM{i}.xlsx"
        if not f.exists():
            continue
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        hdr = [str(c).strip() if c else f'col_{j}' for j, c in enumerate(rows[0])]
        naab_idx = next(j for j, h in enumerate(hdr) if 'naab' in h.lower() and 'code' in h.lower())
        for r in rows[1:]:
            naab = r[naab_idx]
            if not naab:
                continue
            naab = str(naab).strip()
            dams[naab] = {hdr[j]: r[j] for j in range(len(hdr)) if j < len(r)}
    print(f"  Maes: {len(dams)}")

    # Bulls
    bulls_regname = {}
    bulls_name = {}
    with open(BULLS_CSV, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            rn = row.get('Registration Name', '').strip().upper()
            nm = row.get('Name', '').strip().upper()
            if rn:
                bulls_regname[rn] = row
            if nm:
                bulls_name[nm] = row
    print(f"  Pais: {len(bulls_regname)}")

    return daughters, dams, bulls_regname, bulls_name


# ============================================================
# DSII v3 PREDICTION ENGINE
# ============================================================

def compute_dam_deficiency(dam_ptas):
    """Perfil de deficiencias da mae."""
    profile = {}
    for trait, ideal in BREED_IDEAL.items():
        val = dam_ptas.get(trait, 0) or 0
        sd = GENETIC_SD.get(trait, 1)
        if trait in LOWER_IS_BETTER:
            deficit = (val - ideal) / sd
        else:
            deficit = (ideal - val) / sd
        profile[trait] = max(0, deficit)
    return profile


def dsii_v3_predict_trait(sire_val, dam_val, trait, deficiency_profile, expected_f=0.085):
    """
    DSII v3 prediction for a single trait.
    PA + inbreeding depression + deficiency-weighted correction.
    """
    if sire_val is None or dam_val is None:
        return None

    # Base PA
    pa = (sire_val + dam_val) / 2.0

    # Inbreeding depression
    ib_dep = INBREEDING_DEPRESSION.get(trait, 0) * expected_f
    pa_corrected = pa + ib_dep

    # DSII deficiency correction:
    # If dam is deficient in this trait, shift prediction toward sire
    deficiency = deficiency_profile.get(trait, 0)
    if deficiency > 0:
        # DeltaG component: how much sire improves over dam
        delta_g = (sire_val - dam_val) / 2.0
        # Amplify based on deficiency (more deficient = more sire influence)
        amplification = min(0.15, deficiency * 0.05)  # max 15% shift
        if trait in LOWER_IS_BETTER:
            correction = -abs(delta_g) * amplification if delta_g < 0 else 0
        else:
            correction = abs(delta_g) * amplification if delta_g > 0 else 0
        pa_corrected += correction

    return pa_corrected


def compute_critical_penalty(trait_preds):
    penalty = 0.0
    for trait, threshold in CRITICAL_THRESHOLDS.items():
        pred = trait_preds.get(trait, 0)
        if pred is None:
            continue
        if trait in LOWER_IS_BETTER:
            if pred > threshold:
                penalty += (pred - threshold) * 20
        else:
            if pred < threshold:
                penalty += (threshold - pred) * 20
    return penalty


def compute_antagonism_penalty(sire_ptas, dam_ptas):
    penalty = 0.0
    for (t1, t2), corr in GENETIC_CORRELATIONS.items():
        if corr >= 0:
            continue
        s1 = sire_ptas.get(t1, 0) or 0
        d1 = dam_ptas.get(t1, 0) or 0
        s2 = sire_ptas.get(t2, 0) or 0
        d2 = dam_ptas.get(t2, 0) or 0
        sd1 = GENETIC_SD.get(t1, 1)
        sd2 = GENETIC_SD.get(t2, 1)
        pa1 = (s1 + d1) / 2.0
        pa2 = (s2 + d2) / 2.0
        z1 = pa1 / sd1
        z2 = pa2 / sd2
        if z1 > 1.0 and z2 < -0.5:
            severity = abs(corr) * (z1 - z2) * 3
            penalty += severity
    return penalty


def compute_distance_to_ideal(trait_preds):
    dist_sq = 0.0
    for trait, ideal in BREED_IDEAL.items():
        pred = trait_preds.get(trait, 0)
        if pred is None:
            pred = 0
        sd = GENETIC_SD.get(trait, 1)
        w = abs(ECON_WEIGHTS.get(trait, 1))
        if trait in LOWER_IS_BETTER:
            diff = max(0, pred - ideal) / sd
        else:
            diff = max(0, ideal - pred) / sd
        dist_sq += (diff ** 2) * (w / 30)
    return math.sqrt(dist_sq)


# ============================================================
# MAIN COMPARISON
# ============================================================

def run_comparison():
    print("=" * 80)
    print("  COMPARACAO COMPLETA: PA vs DSII v3 vs DSII v4 (ML)")
    print("=" * 80)

    daughters, dams, bulls_regname, bulls_name = load_all_data()

    # Assemble trios with all trait values
    print("\nMontando trios...")
    trios = []

    for daughter in daughters:
        sirename = str(daughter.get('SIRENAME', '')).strip().upper()
        sire = bulls_regname.get(sirename) or bulls_name.get(sirename)
        if not sire:
            continue

        damreg = str(daughter.get('DAMREGNUM2', '')).strip()
        dam = dams.get(damreg)
        if not dam:
            continue

        rec = {'sirename': sirename, 'damreg': damreg}

        # Extract sire PTAs (mapped to our trait names)
        sire_ptas = {}
        dam_ptas = {}

        for trait_name, daughter_col, dam_col, bull_col in TRAIT_MAP:
            d_val = safe_float(daughter.get(daughter_col))
            s_val = safe_float(sire.get(bull_col))
            m_val = safe_float(dam.get(dam_col)) if dam_col else None

            rec[f'd_{trait_name}'] = d_val
            rec[f's_{trait_name}'] = s_val
            rec[f'm_{trait_name}'] = m_val

            if s_val is not None:
                sire_ptas[trait_name] = s_val
            if m_val is not None:
                dam_ptas[trait_name] = m_val

            # PA
            if s_val is not None and m_val is not None:
                rec[f'pa_{trait_name}'] = (s_val + m_val) / 2.0
            else:
                rec[f'pa_{trait_name}'] = None

        # DSII v3 predictions
        deficiency = compute_dam_deficiency(dam_ptas)
        v3_preds = {}

        for trait_name, _, dam_col, _ in TRAIT_MAP:
            if dam_col is None:
                continue
            s_val = sire_ptas.get(trait_name)
            m_val = dam_ptas.get(trait_name)
            v3_pred = dsii_v3_predict_trait(s_val, m_val, trait_name, deficiency)
            rec[f'v3_{trait_name}'] = v3_pred
            if v3_pred is not None:
                v3_preds[trait_name] = v3_pred

        # V3 adjustments (critical penalty, antagonism, distance)
        rec['v3_critical'] = compute_critical_penalty(v3_preds)
        rec['v3_antagonism'] = compute_antagonism_penalty(sire_ptas, dam_ptas)
        rec['v3_distance'] = compute_distance_to_ideal(v3_preds)

        trios.append(rec)

    print(f"  Trios completos: {len(trios)}")
    df = pd.DataFrame(trios)

    # ============================================================
    # TRAIN ML MODELS (DSII v4) AND COMPARE ALL 3
    # ============================================================
    print("\n" + "=" * 80)
    print(f"  {'Trait':>6} | {'N':>5} | {'PA MAE':>8} {'PA R2':>7} | "
          f"{'V3 MAE':>8} {'V3 R2':>7} | {'V4 MAE':>8} {'V4 R2':>7} {'V4 Best':>8} | "
          f"{'V3vsPA':>7} {'V4vsPA':>7}")
    print("-" * 115)

    all_results = []

    for trait_name, _, dam_col, _ in TRAIT_MAP:
        if dam_col is None:
            continue

        target = f'd_{trait_name}'
        pa_col = f'pa_{trait_name}'
        v3_col = f'v3_{trait_name}'
        s_col = f's_{trait_name}'
        m_col = f'm_{trait_name}'

        needed = [target, pa_col, v3_col, s_col, m_col]
        mask = df[needed].notna().all(axis=1)
        sub = df[mask].copy()

        if len(sub) < 50:
            continue

        y = sub[target].values
        pa_pred = sub[pa_col].values
        v3_pred = sub[v3_col].values

        # PA metrics
        pa_mae = mean_absolute_error(y, pa_pred)
        pa_rmse = np.sqrt(mean_squared_error(y, pa_pred))
        pa_r2 = r2_score(y, pa_pred)

        # V3 metrics
        v3_mae = mean_absolute_error(y, v3_pred)
        v3_rmse = np.sqrt(mean_squared_error(y, v3_pred))
        v3_r2 = r2_score(y, v3_pred)

        # V4: ML with all sire+dam features
        s_cols = [c for c in df.columns if c.startswith('s_')]
        m_cols = [c for c in df.columns if c.startswith('m_')]
        feat_cols = [s_col, m_col] + [c for c in s_cols + m_cols if c not in (s_col, m_col)]
        good_cols = [c for c in feat_cols if c in sub.columns and sub[c].notna().sum() > len(sub) * 0.5]

        X = sub[good_cols].copy()
        for c in good_cols:
            if X[c].isna().any():
                X[c] = X[c].fillna(X[c].median())
        X = X.values

        kf = KFold(n_splits=5, shuffle=True, random_state=42)

        # Ridge
        ridge = Ridge(alpha=1.0)
        ridge_mae = -np.mean(np.array([-mean_absolute_error(y[test], ridge.fit(X[train], y[train]).predict(X[test]))
                                        for train, test in kf.split(X)]))
        # Recalculate properly
        ridge_maes, ridge_r2s = [], []
        for train, test in kf.split(X):
            ridge.fit(X[train], y[train])
            pred = ridge.predict(X[test])
            ridge_maes.append(mean_absolute_error(y[test], pred))
            ridge_r2s.append(r2_score(y[test], pred))
        ridge_mae = np.mean(ridge_maes)
        ridge_r2 = np.mean(ridge_r2s)

        # RF
        rf = RandomForestRegressor(n_estimators=200, max_depth=8, min_samples_leaf=5,
                                   random_state=42, n_jobs=-1)
        rf_maes, rf_r2s = [], []
        for train, test in kf.split(X):
            rf.fit(X[train], y[train])
            pred = rf.predict(X[test])
            rf_maes.append(mean_absolute_error(y[test], pred))
            rf_r2s.append(r2_score(y[test], pred))
        rf_mae = np.mean(rf_maes)
        rf_r2 = np.mean(rf_r2s)

        # XGBoost
        if HAS_XGB:
            xgb = XGBRegressor(n_estimators=300, max_depth=5, learning_rate=0.05,
                               subsample=0.8, colsample_bytree=0.8,
                               min_child_weight=5, random_state=42, verbosity=0)
            xgb_maes, xgb_r2s = [], []
            for train, test in kf.split(X):
                xgb.fit(X[train], y[train])
                pred = xgb.predict(X[test])
                xgb_maes.append(mean_absolute_error(y[test], pred))
                xgb_r2s.append(r2_score(y[test], pred))
            xgb_mae = np.mean(xgb_maes)
            xgb_r2 = np.mean(xgb_r2s)
        else:
            xgb_mae, xgb_r2 = 999, 0

        # Best ML
        ml_options = {'Ridge': (ridge_mae, ridge_r2), 'RF': (rf_mae, rf_r2)}
        if HAS_XGB:
            ml_options['XGB'] = (xgb_mae, xgb_r2)
        best_name = min(ml_options, key=lambda k: ml_options[k][0])
        v4_mae, v4_r2 = ml_options[best_name]

        # Improvements
        v3_vs_pa = (pa_mae - v3_mae) / pa_mae * 100
        v4_vs_pa = (pa_mae - v4_mae) / pa_mae * 100

        print(f"  {trait_name:>6} | {len(sub):>5} | {pa_mae:>8.2f} {pa_r2:>7.4f} | "
              f"{v3_mae:>8.2f} {v3_r2:>7.4f} | {v4_mae:>8.2f} {v4_r2:>7.4f} {best_name:>8} | "
              f"{v3_vs_pa:>+6.1f}% {v4_vs_pa:>+6.1f}%")

        all_results.append({
            'Trait': trait_name,
            'N': len(sub),
            'PA_MAE': round(pa_mae, 3),
            'PA_RMSE': round(pa_rmse, 3),
            'PA_R2': round(pa_r2, 4),
            'V3_MAE': round(v3_mae, 3),
            'V3_RMSE': round(v3_rmse, 3),
            'V3_R2': round(v3_r2, 4),
            'V4_MAE': round(v4_mae, 3),
            'V4_R2': round(v4_r2, 4),
            'V4_Best_Model': best_name,
            'V3_vs_PA_%': round(v3_vs_pa, 2),
            'V4_vs_PA_%': round(v4_vs_pa, 2),
        })

    # Summary
    results_df = pd.DataFrame(all_results)
    results_df.to_csv(OUTPUT_DIR / "v3_vs_v4_comparison.csv", index=False)

    print("\n" + "=" * 80)
    print("  RESUMO GERAL")
    print("=" * 80)

    # Group by category
    production = ['MILK', 'FAT', 'FAT%', 'PRO', 'PRO%', 'CFP']
    indices = ['TPI', 'NM$', 'CM$']
    health = ['PL', 'SCS', 'DPR', 'LIV', 'FI', 'HCR', 'CCR', 'MAST', 'FSAV']
    type_traits = ['PTAT', 'UDC', 'FLC']
    calving = ['SCE', 'DCE', 'SSB', 'DSB']

    categories = [
        ('INDICES', indices),
        ('PRODUCAO', production),
        ('SAUDE/FERT', health),
        ('TIPO', type_traits),
        ('PARTO', calving),
    ]

    for cat_name, cat_traits in categories:
        cat_data = [r for r in all_results if r['Trait'] in cat_traits]
        if not cat_data:
            continue
        avg_pa_r2 = np.mean([r['PA_R2'] for r in cat_data])
        avg_v3_r2 = np.mean([r['V3_R2'] for r in cat_data])
        avg_v4_r2 = np.mean([r['V4_R2'] for r in cat_data])
        avg_v3_imp = np.mean([r['V3_vs_PA_%'] for r in cat_data])
        avg_v4_imp = np.mean([r['V4_vs_PA_%'] for r in cat_data])
        print(f"\n  {cat_name}:")
        print(f"    PA  R2 medio: {avg_pa_r2:.4f}")
        print(f"    V3  R2 medio: {avg_v3_r2:.4f}  ({avg_v3_imp:+.1f}% MAE vs PA)")
        print(f"    V4  R2 medio: {avg_v4_r2:.4f}  ({avg_v4_imp:+.1f}% MAE vs PA)")

    # Overall
    avg_pa = np.mean([r['PA_R2'] for r in all_results])
    avg_v3 = np.mean([r['V3_R2'] for r in all_results])
    avg_v4 = np.mean([r['V4_R2'] for r in all_results])
    avg_v3i = np.mean([r['V3_vs_PA_%'] for r in all_results])
    avg_v4i = np.mean([r['V4_vs_PA_%'] for r in all_results])

    print(f"\n  {'='*40}")
    print(f"  GERAL (todos {len(all_results)} traits):")
    print(f"    PA  R2 medio: {avg_pa:.4f}")
    print(f"    V3  R2 medio: {avg_v3:.4f}  ({avg_v3i:+.1f}% MAE vs PA)")
    print(f"    V4  R2 medio: {avg_v4:.4f}  ({avg_v4i:+.1f}% MAE vs PA)")

    # Winner per trait
    v3_wins = sum(1 for r in all_results if r['V3_R2'] > r['V4_R2'])
    v4_wins = sum(1 for r in all_results if r['V4_R2'] > r['V3_R2'])
    ties = sum(1 for r in all_results if abs(r['V4_R2'] - r['V3_R2']) < 0.001)

    print(f"\n  Vencedor por trait:")
    print(f"    DSII v3 vence: {v3_wins} traits")
    print(f"    DSII v4 vence: {v4_wins} traits")
    print(f"    Empate (<0.001 R2): {ties} traits")

    # Best approach per trait
    print(f"\n  Melhor abordagem por trait:")
    for r in all_results:
        approaches = {'PA': r['PA_R2'], 'V3': r['V3_R2'], 'V4': r['V4_R2']}
        best = max(approaches, key=approaches.get)
        print(f"    {r['Trait']:>6}: {best} (R2={approaches[best]:.4f})")

    print(f"\n  Resultados salvos: {OUTPUT_DIR / 'v3_vs_v4_comparison.csv'}")

    return results_df


if __name__ == '__main__':
    run_comparison()
